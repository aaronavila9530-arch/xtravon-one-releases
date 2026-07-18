import os
import calendar
import ctypes
import shutil
import stat
import subprocess
import sys
import tempfile
import threading
import time
import tkinter as tk
import webbrowser
from tkinter import filedialog, messagebox, simpledialog, ttk
from datetime import date, datetime

import openpyxl
import qrcode
import requests
from openpyxl.styles import Alignment, Font, PatternFill

from frontend import install_frontend_modules


def app_resource_path(*parts):
    base_dir = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_dir, *parts)


def app_user_data_path(*parts):
    base_dir = os.environ.get("LOCALAPPDATA") or tempfile.gettempdir()
    path = os.path.join(base_dir, "XTRAVON_ONE", *parts)
    os.makedirs(os.path.dirname(path) if parts else path, exist_ok=True)
    return path


def post_json_with_retry(url, payload, timeout=45, retries=1, retry_delay=0.8):
    last_response = None
    for intento in range(retries + 1):
        try:
            response = requests.post(url, json=payload, timeout=timeout)
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError):
            if intento >= retries:
                raise
            time.sleep(retry_delay)
            continue

        last_response = response
        if response.status_code < 500 or intento >= retries:
            return response
        time.sleep(retry_delay)

    return last_response


SPLASH_DURATION_MS = 3000
SPLASH_IMAGE_PATH = app_resource_path("assets", "xtravon_splash_desktop.png")
APP_ICON_PATH = app_resource_path("assets", "XTRAVON_seal_round_transparent.png")
APP_VERSION = "1.0.5"
API_BASE_DEFAULT = os.getenv("XTRAVON_API_BASE", "https://erp-elsurco-backend-production.up.railway.app")
WINDOWS_CREDENTIAL_SERVICE = "XTRAVON_ONE_DESKTOP"


class ERPElSurcoApp(tk.Tk):
    def __init__(self, auth_data=None):
        super().__init__()
        try:
            current_scaling = float(self.tk.call("tk", "scaling"))
            self.tk.call("tk", "scaling", current_scaling * 1.08)
        except Exception:
            pass

        self.title(f"XTRAVON ONE v{APP_VERSION}")
        self.aplicar_icono_app()
        screen_w = self.winfo_screenwidth()
        screen_h = self.winfo_screenheight()
        initial_w = min(1366, max(1024, screen_w - 40))
        initial_h = min(768, max(680, screen_h - 80))
        self.geometry(f"{initial_w}x{initial_h}")
        self.minsize(980, 640)
        self.resizable(True, True)
        self.configure(bg="#050B14")
        self.protocol("WM_DELETE_WINDOW", self.confirmar_salida)

        self.api_base = API_BASE_DEFAULT
        self.app_version = APP_VERSION
        self._update_prompted_version = None
        self._update_installing_version = None
        self.auth_data = auth_data or {}
        self.current_user = (self.auth_data or {}).get("usuario") or {}
        self.current_roles = [
            str(role.get("nombre", "")).lower()
            for role in (self.auth_data or {}).get("roles", [])
            if isinstance(role, dict)
        ]
        self.tree = None
        self.boletas_cache = []
        self.dashboard_canvas = None
        self.dashboard_body = None
        self.boletas_template_local_path = app_resource_path("backend", "Template", "base_operaciones_camiones.xlsx")
        self.boletas_template_trabajo_path = None
        self.boletas_template_trabajo_operacion_id = None
        self.operaciones_cache = []
        self.operacion_activa = None
        self.operaciones_tree = None
        self.cuotas_tree = None
        self._ui_after_jobs = {}
        self._screen_transition_active = False
        self._pending_old_content = None
        self._pending_new_content = None
        self._screen_cache = {}
        self._current_screen_key = None
        self._prebuild_queue = []
        self._prebuilding_screen = False
        self._suppress_sidebar_highlight = False
        self.portia_panel_visible = False
        self.portia_contexto = {
            "operacion_id": None,
            "operacion_label": "",
            "ultima_pregunta": "",
            "ultima_respuesta": "",
            "pantalla": "",
        }

        self.colors = {
            "bg_main": "#050B14",
            "bg_sidebar": "#07111F",
            "bg_topbar": "#0B1B2E",
            "bg_card": "#0B1B2E",
            "bg_elevated": "#14283D",
            "accent": "#00D1FF",
            "accent_light": "#2979FF",
            "teal": "#00B8A9",
            "text_dark": "#F4F8FF",
            "text_light": "#050B14",
            "text_secondary": "#DCEBFF",
            "text_aux": "#8FA4BC",
            "border": "#14283D",
            "button_hover": "#5AA9FF",
            "warning": "#FFB020",
            "success": "#1EE6A8",
            "info": "#5AA9FF",
            "danger": "#FF5A6A",
            "muted": "#8FA4BC",
        }

        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.configure_styles()

        self.sidebar_buttons = {}
        self.build_layout()
        self.show_centro_ejecutivo()
        self._screen_cache["Centro Ejecutivo"] = self.content
        self._current_screen_key = "Centro Ejecutivo"
        self.build_portia_floating_panel()
        self.after(1800, self.iniciar_portia_global_segura)
        self.after(3000, lambda: self.verificar_actualizacion_remota(silencioso=True))
        self.after(100, self.maximizar_ventana)

    def aplicar_icono_app(self, window=None):
        target = window or self
        try:
            icon = tk.PhotoImage(file=APP_ICON_PATH)
            target.iconphoto(True, icon)
            target._xtravon_icon = icon
        except Exception:
            pass

    def iniciar_portia_global_segura(self):
        try:
            if callable(getattr(self, "ia_iniciar_escucha_continua", None)):
                self.ia_iniciar_escucha_continua()
                if hasattr(self, "portia_status_var"):
                    self.portia_status_var.set("Escucha activa: diga Oye Portia, Hola Portia o Portia estas ahi.")
        except Exception:
            if hasattr(self, "portia_status_var"):
                self.portia_status_var.set("Escucha de voz no disponible. Use la caja de pregunta.")

    def usuario_tiene_rol(self, *roles):
        if not roles:
            return False
        roles_norm = {str(role).strip().lower() for role in roles}
        return bool(set(self.current_roles) & roles_norm)

    def filtrar_menu_por_rol(self, menu_items):
        if self.usuario_tiene_rol("master", "administrador", "supervisor"):
            return menu_items
        if self.usuario_tiene_rol("chofer"):
            permitidos = {"P.O.R.T.I.A", "Ayuda / Q&A"}
            return [(name, cmd) for name, cmd in menu_items if name in permitidos]
        if self.usuario_tiene_rol("operador de patio"):
            permitidos = {"SOF", "P.O.R.T.I.A", "Ayuda / Q&A"}
            return [(name, cmd) for name, cmd in menu_items if name in permitidos]
        if self.usuario_tiene_rol("cliente"):
            permitidos = {"Centro Ejecutivo", "Informes", "P.O.R.T.I.A", "Ayuda / Q&A"}
            return [(name, cmd) for name, cmd in menu_items if name in permitidos]
        return menu_items

    def configure_styles(self):
        c = self.colors
        self.style.configure(
            "Treeview",
            background=c["bg_topbar"],
            foreground=c["text_dark"],
            fieldbackground=c["bg_topbar"],
            rowheight=30,
            borderwidth=0,
            font=("Segoe UI", 10),
        )
        self.style.configure(
            "Treeview.Heading",
            background=c["accent"],
            foreground=c["text_light"],
            font=("Segoe UI", 10, "bold"),
            relief="flat",
        )
        self.style.map("Treeview.Heading", background=[("active", c["button_hover"])])
        self.style.configure(
            "Olive.TButton",
            background=c["accent"],
            foreground=c["text_light"],
            font=("Segoe UI", 10, "bold"),
            borderwidth=0,
            padding=10,
        )
        self.style.map("Olive.TButton", background=[("active", c["button_hover"])])
        self.style.configure(
            "Gray.TButton",
            background=c["bg_elevated"],
            foreground=c["text_dark"],
            font=("Segoe UI", 10, "bold"),
            borderwidth=0,
            padding=10,
        )
        self.style.map("Gray.TButton", background=[("active", c["accent_light"])])
        self.style.configure(
            "TCombobox",
            fieldbackground=c["bg_topbar"],
            background=c["bg_elevated"],
            foreground=c["text_dark"],
            arrowcolor=c["accent"],
            selectbackground=c["accent"],
            selectforeground=c["bg_main"],
            font=("Segoe UI", 10),
        )

    def build_layout(self):
        self.sidebar = tk.Frame(self, bg=self.colors["bg_sidebar"], width=250)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)

        self.main_container = tk.Frame(self, bg=self.colors["bg_main"])
        self.main_container.pack(side="right", fill="both", expand=True)
        self.main_container.pack_propagate(False)

        self.topbar = tk.Frame(self.main_container, bg=self.colors["bg_topbar"], height=70)
        self.topbar.pack(side="top", fill="x")
        self.topbar.pack_propagate(False)

        self.content = tk.Frame(self.main_container, bg=self.colors["bg_main"])
        self.content.pack(side="top", fill="both", expand=True)
        self.content.pack_propagate(False)

        self.build_sidebar()
        self.build_topbar()

    def build_sidebar(self):
        title_frame = tk.Frame(self.sidebar, bg=self.colors["bg_sidebar"])
        title_frame.pack(fill="x", padx=20, pady=(20, 10))

        try:
            logo_image = tk.PhotoImage(file=APP_ICON_PATH)
            scale = max(1, max(logo_image.width(), logo_image.height()) // 72)
            if scale > 1:
                logo_image = logo_image.subsample(scale, scale)
            self.sidebar_logo_image = logo_image
            tk.Label(
                title_frame,
                image=self.sidebar_logo_image,
                bg=self.colors["bg_sidebar"],
                bd=0,
                highlightthickness=0,
            ).pack(anchor="w", pady=(0, 10))
        except Exception:
            self.sidebar_logo_image = None

        tk.Label(
            title_frame,
            text="XTRAVON ONE",
            font=("Segoe UI", 14, "bold"),
            bg=self.colors["bg_sidebar"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w")
        tk.Label(
            title_frame,
            text="GRAIN CONTROL",
            font=("Segoe UI", 10),
            bg=self.colors["bg_sidebar"],
            fg=self.colors["accent"],
        ).pack(anchor="w", pady=(2, 0))

        tk.Frame(self.sidebar, bg=self.colors["border"], height=1).pack(fill="x", padx=20, pady=15)

        self.screen_commands = {
            "Centro Ejecutivo": self.show_centro_ejecutivo,
            "Informes": self.show_informes,
            "Liquidaciones Choferes": self.show_liquidaciones_choferes,
            "Despacho de Viajes": self.show_despacho_viajes,
            "Operaciones Buque": self.show_operaciones_buque,
            "Carga de Boletas": self.show_boletas,
            "Aprobaciones": self.show_aprobaciones,
            "SOF": self.show_issue_log,
            "Historial de Buques": self.show_buques,
            "Gestion Ejecutiva": self.show_gestion_profesional,
            "P.O.R.T.I.A": self.show_ia_ejecutiva,
            "Roles y Permisos": self.show_roles_permisos,
            "Ayuda / Q&A": self.show_ayuda_qa,
        }

        menu_items = [
            ("Centro Ejecutivo", self.show_centro_ejecutivo),
            ("Informes", self.show_informes),
            ("Liquidaciones Choferes", self.show_liquidaciones_choferes),
            ("Despacho de Viajes", self.show_despacho_viajes),
            ("Operaciones Buque", self.show_operaciones_buque),
            ("Carga de Boletas", self.show_boletas),
            ("Aprobaciones", self.show_aprobaciones),
            ("SOF", self.show_issue_log),
            ("Historial de Buques", self.show_buques),
            ("P.O.R.T.I.A", self.show_ia_ejecutiva),
            ("Roles y Permisos", self.show_roles_permisos),
            ("Ayuda / Q&A", self.show_ayuda_qa),
        ]
        menu_items = self.filtrar_menu_por_rol(menu_items)
        menu_item_names = {name for name, _cmd in menu_items}
        menu_groups = [
            ("01", "EJECUTIVO", ["Centro Ejecutivo", "Informes", "Liquidaciones Choferes"]),
            ("02", "DESPACHO", ["Despacho de Viajes"]),
            ("03", "OPERACION", ["Operaciones Buque", "Carga de Boletas", "Aprobaciones", "SOF", "Historial de Buques"]),
            ("04", "INTELIGENCIA", ["P.O.R.T.I.A", "Roles y Permisos", "Ayuda / Q&A"]),
        ]

        bottom_frame = tk.Frame(self.sidebar, bg=self.colors["bg_sidebar"])
        bottom_frame.pack(side="bottom", fill="x", padx=20, pady=20)
        tk.Label(
            bottom_frame,
            text="QORA SYSTEMS",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors["bg_sidebar"],
            fg=self.colors["accent"],
        ).pack(anchor="w")
        tk.Label(
            bottom_frame,
            text=f"XTRAVON ONE v{APP_VERSION}",
            font=("Segoe UI", 8),
            bg=self.colors["bg_sidebar"],
            fg=self.colors["text_light"],
        ).pack(anchor="w", pady=(4, 0))

        menu_container = tk.Frame(self.sidebar, bg=self.colors["bg_sidebar"])
        menu_container.pack(side="top", fill="both", expand=True, padx=(0, 6), pady=(0, 6))

        menu_canvas = tk.Canvas(menu_container, bg=self.colors["bg_sidebar"], highlightthickness=0, bd=0)
        menu_scroll = ttk.Scrollbar(menu_container, orient="vertical", command=menu_canvas.yview)
        menu_frame = tk.Frame(menu_canvas, bg=self.colors["bg_sidebar"])
        menu_window = menu_canvas.create_window((0, 0), window=menu_frame, anchor="nw")

        self.sidebar_menu_canvas = menu_canvas
        self.sidebar_menu_frame = menu_frame

        def configure_menu_scroll(_event=None):
            menu_canvas.configure(scrollregion=menu_canvas.bbox("all"))

        def configure_menu_width(event):
            menu_canvas.itemconfigure(menu_window, width=event.width)

        def on_menu_wheel(event):
            if menu_canvas.bbox("all"):
                menu_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        menu_frame.bind("<Configure>", configure_menu_scroll)
        menu_canvas.bind("<Configure>", configure_menu_width)
        menu_canvas.bind("<Enter>", lambda _event: menu_canvas.bind_all("<MouseWheel>", on_menu_wheel))
        menu_canvas.bind("<Leave>", lambda _event: menu_canvas.unbind_all("<MouseWheel>"))
        menu_canvas.configure(yscrollcommand=menu_scroll.set)
        menu_canvas.pack(side="left", fill="both", expand=True)
        menu_scroll.pack(side="right", fill="y")

        rendered = set()
        self.sidebar_group_bodies = {}
        self.sidebar_group_toggles = {}

        def toggle_sidebar_group(group_key):
            body = self.sidebar_group_bodies.get(group_key)
            toggle = self.sidebar_group_toggles.get(group_key)
            if body is None or toggle is None:
                return
            if body.winfo_ismapped():
                body.pack_forget()
                toggle.configure(text="+")
            else:
                body.pack(fill="x", padx=6, pady=(0, 8))
                toggle.configure(text="-")

        def add_group(number, title, names):
            visibles = [name for name in names if name in menu_item_names]
            if not visibles:
                return
            group_key = f"{number}_{title}"
            group = tk.Frame(
                menu_frame,
                bg=self.colors["bg_sidebar"],
                highlightbackground=self.colors["border"],
                highlightthickness=1,
            )
            group.pack(fill="x", padx=10, pady=(0, 10))
            header = tk.Frame(group, bg=self.colors["bg_sidebar"])
            header.pack(fill="x", padx=10, pady=(8, 2))
            tk.Label(
                header,
                text=number,
                font=("Segoe UI", 8, "bold"),
                bg=self.colors["accent"],
                fg=self.colors["bg_main"],
                width=3,
            ).pack(side="left")
            tk.Label(
                header,
                text=title,
                font=("Segoe UI", 8, "bold"),
                bg=self.colors["bg_sidebar"],
                fg=self.colors["accent"],
            ).pack(side="left", padx=(8, 0))
            toggle_btn = tk.Button(
                header,
                text="-",
                font=("Segoe UI", 9, "bold"),
                bg=self.colors["bg_sidebar"],
                fg=self.colors["accent"],
                activebackground=self.colors["button_hover"],
                activeforeground=self.colors["bg_main"],
                relief="flat",
                bd=0,
                width=2,
                command=lambda key=group_key: toggle_sidebar_group(key),
            )
            toggle_btn.pack(side="right")
            self.sidebar_group_toggles[group_key] = toggle_btn

            body = tk.Frame(group, bg=self.colors["bg_sidebar"])
            body.pack(fill="x", padx=6, pady=(0, 8))
            self.sidebar_group_bodies[group_key] = body

            for text in visibles:
                command = self.screen_commands.get(text)
                if command is None:
                    continue
                btn = tk.Button(
                    body,
                    text=text,
                    font=("Segoe UI", 10, "bold"),
                    bg=self.colors["bg_sidebar"],
                    fg=self.colors["text_dark"],
                    activebackground=self.colors["button_hover"],
                    activeforeground=self.colors["bg_main"],
                    relief="flat",
                    bd=0,
                    padx=14,
                    pady=10,
                    anchor="w",
                    command=lambda t=text, cmd=command: self.on_sidebar_click(t, cmd),
                )
                btn.pack(fill="x", padx=6, pady=2)
                self.sidebar_buttons[text] = btn
                rendered.add(text)

        for number, title, names in menu_groups:
            add_group(number, title, names)

        extras = [name for name, _cmd in menu_items if name not in rendered]
        if extras:
            add_group("99", "OTROS", extras)

    def build_topbar(self):
        left = tk.Frame(self.topbar, bg=self.colors["bg_topbar"])
        left.pack(side="left", fill="y", padx=20)
        tk.Label(
            left,
            text="XTRAVON ONE | GRAIN CONTROL",
            font=("Segoe UI", 16, "bold"),
            bg=self.colors["bg_topbar"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", pady=(12, 0))
        tk.Label(
            left,
            text=f"QORA SYSTEMS - Alianza con MSL Marine Surveyors and Logistics Group | v{APP_VERSION}",
            font=("Segoe UI", 10),
            bg=self.colors["bg_topbar"],
            fg=self.colors["muted"],
        ).pack(anchor="w")

        right = tk.Frame(self.topbar, bg=self.colors["bg_topbar"])
        right.pack(side="right", fill="y", padx=20)

    def build_portia_floating_panel(self):
        self.portia_float_button = tk.Button(
            self.main_container,
            text="P.O.R.T.I.A",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors["accent"],
            fg=self.colors["bg_main"],
            activebackground=self.colors["button_hover"],
            activeforeground=self.colors["bg_main"],
            relief="flat",
            bd=0,
            padx=16,
            pady=10,
            command=self.toggle_portia_panel,
        )
        self.portia_float_button.place(relx=0.985, rely=0.965, anchor="se")

        self.portia_panel = tk.Frame(
            self.main_container,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        self.portia_panel.place_forget()

        header = tk.Frame(self.portia_panel, bg=self.colors["bg_card"])
        header.pack(fill="x", padx=12, pady=(10, 6))
        tk.Label(
            header,
            text="P.O.R.T.I.A",
            font=("Segoe UI", 13, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["accent"],
        ).pack(side="left")
        tk.Button(
            header,
            text="x",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
            activebackground=self.colors["bg_card"],
            activeforeground=self.colors["danger"],
            relief="flat",
            bd=0,
            command=self.toggle_portia_panel,
        ).pack(side="right")

        self.portia_status_var = tk.StringVar(value="Copiloto listo. No consulta el backend hasta presionar Preguntar.")
        tk.Label(
            self.portia_panel,
            textvariable=self.portia_status_var,
            font=("Segoe UI", 9),
            bg=self.colors["bg_card"],
            fg=self.colors["muted"],
            wraplength=405,
            justify="left",
        ).pack(fill="x", padx=12, pady=(0, 8))

        self.portia_question_var = tk.StringVar()
        entrada = ttk.Entry(self.portia_panel, textvariable=self.portia_question_var)
        entrada.pack(fill="x", padx=12, pady=(0, 8))
        entrada.bind("<Return>", lambda _event: self.portia_preguntar_flotante())

        actions = tk.Frame(self.portia_panel, bg=self.colors["bg_card"])
        actions.pack(fill="x", padx=12, pady=(0, 8))
        ttk.Button(actions, text="Preguntar", style="Olive.TButton", command=self.portia_preguntar_flotante).pack(side="left", padx=(0, 6))
        ttk.Button(actions, text="Leer", style="Gray.TButton", command=self.portia_leer_flotante).pack(side="left", padx=(0, 6))
        ttk.Button(actions, text="Detener voz", style="Gray.TButton", command=self.portia_detener_voz_global).pack(side="left")

        text_frame = tk.Frame(self.portia_panel, bg=self.colors["bg_card"])
        text_frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        self.portia_resultado = tk.Text(
            text_frame,
            height=9,
            wrap="word",
            bg=self.colors["bg_topbar"],
            fg=self.colors["text_dark"],
            insertbackground=self.colors["text_dark"],
            relief="flat",
            font=("Segoe UI", 9),
        )
        scroll = ttk.Scrollbar(text_frame, orient="vertical", command=self.portia_resultado.yview)
        self.portia_resultado.configure(yscrollcommand=scroll.set)
        self.portia_resultado.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        self.portia_resultado.insert("1.0", "Pregunte en lenguaje natural. P.O.R.T.I.A responde por texto y puede leer la respuesta en voz.\n")

    def toggle_portia_panel(self):
        self.portia_panel_visible = not self.portia_panel_visible
        if self.portia_panel_visible:
            self.portia_panel.place(relx=0.985, rely=0.94, anchor="se", width=460, height=330)
            self.portia_panel.lift()
            self.portia_float_button.lift()
            self.portia_actualizar_contexto_visual()
        else:
            self.portia_panel.place_forget()

    def portia_actualizar_contexto_visual(self):
        contexto = self.portia_obtener_contexto_operacion()
        pantalla = self._current_screen_key or ""
        self.portia_contexto["pantalla"] = pantalla
        if contexto.get("operacion_id"):
            self.portia_status_var.set(f"Contexto: {contexto.get('operacion_label', '')} | Pantalla: {pantalla or 'N/D'}")
        else:
            self.portia_status_var.set(f"Sin operaciÃ³n seleccionada. Pantalla: {pantalla or 'N/D'}. Puede consultar clima, calado o AIS.")

    def portia_obtener_contexto_operacion(self):
        contexto = dict(getattr(self, "portia_contexto", {}) or {})

        if getattr(self, "ia_memoria", None):
            contexto["operacion_id"] = self.ia_memoria.get("operacion_id") or contexto.get("operacion_id")
            contexto["operacion_label"] = self.ia_memoria.get("operacion_label") or contexto.get("operacion_label", "")

        for tree_name in ("centro_operaciones_tree", "operaciones_tree"):
            tree = getattr(self, tree_name, None)
            try:
                selected = tree.selection()
                if selected:
                    values = tree.item(selected[0], "values")
                    if values:
                        contexto["operacion_id"] = values[0]
                        contexto["operacion_label"] = " | ".join(str(v) for v in values[:4] if v not in (None, ""))
                        break
            except Exception:
                pass

        self.portia_contexto.update(contexto)
        return contexto

    def portia_pregunta_rapida(self, pregunta):
        self.portia_question_var.set(pregunta)
        self.portia_preguntar_flotante()

    def portia_es_comando_ejecutable(self, pregunta):
        texto = (pregunta or "").lower()
        for old, new in (("Ã¡", "a"), ("Ã©", "e"), ("Ã­", "i"), ("Ã³", "o"), ("Ãº", "u"), ("Ã±", "n")):
            texto = texto.replace(old, new)
        comandos = [
            "sala de control",
            "modo comando",
            "estado de mando",
            "control room",
            "crear acciones",
            "crea acciones",
            "generar acciones",
            "genera acciones",
            "escalar criticas",
            "escala criticas",
            "escalar altas",
            "escala altas",
            "que es urgente",
            "que atiendo",
            "que hago primero",
        ]
        return any(cmd in texto for cmd in comandos)

    def portia_preguntar_flotante(self):
        pregunta = self.portia_question_var.get().strip()
        if not pregunta:
            messagebox.showwarning("P.O.R.T.I.A", "Escriba una pregunta o use un botÃ³n rÃ¡pido.")
            return

        contexto = self.portia_obtener_contexto_operacion()
        operacion_id = contexto.get("operacion_id")
        es_comando = self.portia_es_comando_ejecutable(pregunta)
        if es_comando and not operacion_id:
            messagebox.showwarning("P.O.R.T.I.A", "Seleccione una operaciÃ³n para ejecutar comandos.")
            return
        self.portia_contexto["ultima_pregunta"] = pregunta
        self.portia_status_var.set("Ejecutando comando P.O.R.T.I.A..." if es_comando else "Consultando P.O.R.T.I.A...")

        def tarea():
            if es_comando:
                return self.api_post_portia_comando(operacion_id, pregunta)
            return self.api_maritime_chat(
                pregunta,
                operacion_id=operacion_id,
                modo="Ejecutivo",
                buscar_web=True,
                pantalla=f"P.O.R.T.I.A global - {self.portia_contexto.get('pantalla', '')}",
                respuesta_breve=True,
            )

        def al_terminar(data):
            texto = data.get("text", "") if isinstance(data, dict) else str(data)
            self.portia_contexto["ultima_respuesta"] = texto
            self.portia_resultado.delete("1.0", "end")
            self.portia_resultado.insert("1.0", f"Pregunta: {pregunta}\n\n{texto}")
            self.portia_status_var.set("Respuesta lista.")

        self.ejecutar_en_segundo_plano(
            "P.O.R.T.I.A",
            "Ejecutando comando operativo..." if es_comando else "Consultando operaciÃ³n, riesgos o datos externos...",
            tarea,
            al_terminar,
        )

    def portia_detener_voz_global(self):
        try:
            self.ia_detener_voz()
        except Exception:
            pass
        try:
            self.ia_silenciar_portia()
        except Exception:
            pass
        self.portia_status_var.set("Voz detenida. P.O.R.T.I.A queda en espera.")

    def portia_leer_flotante(self):
        try:
            from frontend.ia_ejecutiva import _crear_proceso_hablar_windows, _ia_texto_para_voz
        except Exception as exc:
            messagebox.showerror("P.O.R.T.I.A", f"No se pudo activar la voz: {exc}")
            return

        texto = self.portia_resultado.get("1.0", "end").strip() if getattr(self, "portia_resultado", None) else ""
        if not texto:
            messagebox.showwarning("Sin texto", "No hay respuesta para leer.")
            return

        texto = _ia_texto_para_voz(texto)
        if not texto:
            return

        self.portia_detener_voz_global()
        self.portia_status_var.set("Hablando respuesta...")

        def worker():
            try:
                proceso = _crear_proceso_hablar_windows(texto)
                self.ia_voz_proceso = proceso
                proceso.wait(timeout=120)
            except Exception:
                pass
            finally:
                self.ia_voz_proceso = None
                self.after(0, lambda: self.portia_status_var.set("Respuesta lista."))

        threading.Thread(target=worker, daemon=True).start()

    def api_get_portia_memoria(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/ai/operacion/{operacion_id}/memoria",
            params={"limit": 12},
            timeout=45,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_post_portia_comando(self, operacion_id, comando):
        respuesta = requests.post(
            f"{self.api_base}/ai/operacion/{operacion_id}/comando",
            json={"comando": comando, "creado_por": "P.O.R.T.I.A"},
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_portia_timeline(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/ai/operacion/{operacion_id}/timeline",
            params={"limit": 120},
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_portia_briefing(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/ai/operacion/{operacion_id}/briefing",
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_portia_sala_control(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/ai/operacion/{operacion_id}/sala-control",
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_portia_plan(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/ai/operacion/{operacion_id}/plan-accion",
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_post_portia_crear_acciones_plan(self, operacion_id):
        respuesta = requests.post(
            f"{self.api_base}/ai/operacion/{operacion_id}/plan-accion/crear-acciones",
            json={"creado_por": "P.O.R.T.I.A", "limpiar_abiertas_portia": False},
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_portia_acciones(self, operacion_id, estado="ABIERTA"):
        respuesta = requests.get(
            f"{self.api_base}/ai/operacion/{operacion_id}/acciones",
            params={"estado": estado},
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_post_portia_completar_accion(self, accion_id):
        respuesta = requests.post(
            f"{self.api_base}/ai/acciones/{accion_id}/completar",
            timeout=45,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_post_portia_escalar_accion(self, accion_id, destinatario=None, canal="INTERNO"):
        payload = {
            "destinatario": destinatario or "",
            "canal": canal,
            "creado_por": "P.O.R.T.I.A",
        }
        respuesta = requests.post(
            f"{self.api_base}/ai/acciones/{accion_id}/escalar",
            json=payload,
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_portia_notificaciones(self, operacion_id, estado="PENDIENTE"):
        respuesta = requests.get(
            f"{self.api_base}/ai/operacion/{operacion_id}/notificaciones",
            params={"estado": estado},
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_post_portia_notificacion_enviada(self, notificacion_id):
        respuesta = requests.post(
            f"{self.api_base}/ai/notificaciones/{notificacion_id}/marcar-enviada",
            timeout=45,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def portia_cargar_memoria_flotante(self):
        contexto = self.portia_obtener_contexto_operacion()
        operacion_id = contexto.get("operacion_id")
        if not operacion_id:
            messagebox.showwarning("P.O.R.T.I.A", "Seleccione una operaciÃ³n para consultar memoria.")
            return

        def tarea():
            return self.api_get_portia_memoria(operacion_id)

        def al_terminar(data):
            rows = data.get("data", []) if isinstance(data, dict) else []
            self.portia_resultado.delete("1.0", "end")
            if not rows:
                self.portia_resultado.insert("1.0", "Esta operaciÃ³n aÃºn no tiene memoria de P.O.R.T.I.A.")
                self.portia_status_var.set("Sin memoria registrada.")
                return
            lines = [f"Memoria P.O.R.T.I.A | OperaciÃ³n {operacion_id}", ""]
            for row in rows:
                pregunta = str(row.get("pregunta", "")).strip()
                respuesta = str(row.get("respuesta", "")).strip().replace("\n", " ")
                lines.append(f"- {row.get('creado_en', '')} | {pregunta}")
                if respuesta:
                    lines.append(f"  {respuesta[:220]}{'...' if len(respuesta) > 220 else ''}")
            self.portia_resultado.insert("1.0", "\n".join(lines))
            self.portia_status_var.set(f"Memoria cargada: {len(rows)} interacciones.")

        self.ejecutar_en_segundo_plano(
            "P.O.R.T.I.A",
            "Cargando memoria operativa...",
            tarea,
            al_terminar,
        )

    def portia_operacion_requerida(self):
        contexto = self.portia_obtener_contexto_operacion()
        operacion_id = contexto.get("operacion_id")
        if not operacion_id:
            messagebox.showwarning("P.O.R.T.I.A", "Seleccione una operaciÃ³n para usar esta funciÃ³n.")
            return None
        return operacion_id

    def portia_cargar_timeline_flotante(self):
        operacion_id = self.portia_operacion_requerida()
        if not operacion_id:
            return

        def tarea():
            return self.api_get_portia_timeline(operacion_id)

        def al_terminar(data):
            resumen = data.get("resumen", {}) if isinstance(data, dict) else {}
            eventos = data.get("eventos", []) if isinstance(data, dict) else []
            lines = [
                f"Timeline operativo | OperaciÃ³n {operacion_id}",
                "",
                f"Eventos: {resumen.get('total_eventos', 0)} | Alertas aparentes: {resumen.get('alertas', 0)}",
                f"Primer evento: {resumen.get('primer_evento') or '-'}",
                f"Ãšltimo evento: {resumen.get('ultimo_evento') or '-'}",
                "",
            ]
            for item in eventos[:60]:
                lines.append(f"[{item.get('fecha') or '-'}] {item.get('tipo')} | {item.get('titulo')}")
                detalle = item.get("detalle")
                if detalle:
                    lines.append(f"  {detalle}")
            self.portia_resultado.delete("1.0", "end")
            self.portia_resultado.insert("1.0", "\n".join(lines))
            self.portia_status_var.set(f"Timeline cargado: {len(eventos)} eventos.")

        self.ejecutar_en_segundo_plano(
            "P.O.R.T.I.A",
            "Reconstruyendo timeline operativo...",
            tarea,
            al_terminar,
        )

    def portia_cargar_briefing_flotante(self):
        operacion_id = self.portia_operacion_requerida()
        if not operacion_id:
            return

        def tarea():
            return self.api_get_portia_briefing(operacion_id)

        def al_terminar(data):
            texto = data.get("text", "") if isinstance(data, dict) else str(data)
            self.portia_resultado.delete("1.0", "end")
            self.portia_resultado.insert("1.0", texto or "No se pudo generar briefing.")
            self.portia_status_var.set("Briefing operativo generado.")

        self.ejecutar_en_segundo_plano(
            "P.O.R.T.I.A",
            "Generando briefing operativo con timeline...",
            tarea,
            al_terminar,
        )

    def portia_cargar_sala_control_flotante(self):
        operacion_id = self.portia_operacion_requerida()
        if not operacion_id:
            return

        def tarea():
            return self.api_get_portia_sala_control(operacion_id)

        def al_terminar(data):
            texto = data.get("text", "") if isinstance(data, dict) else str(data)
            self.portia_contexto["ultima_respuesta"] = texto
            self.portia_resultado.delete("1.0", "end")
            self.portia_resultado.insert("1.0", texto or "No se pudo cargar sala de control.")
            resumen = data.get("resumen", {}) if isinstance(data, dict) else {}
            self.portia_status_var.set(
                f"Sala de control lista. Riesgo: {resumen.get('riesgo', 'N/D')} | "
                f"Acciones: {resumen.get('acciones_abiertas', 0)} | "
                f"Notificaciones: {resumen.get('notificaciones_pendientes', 0)}"
            )

        self.ejecutar_en_segundo_plano(
            "P.O.R.T.I.A",
            "Construyendo sala de control operativa...",
            tarea,
            al_terminar,
        )

    def portia_cargar_plan_flotante(self):
        operacion_id = self.portia_operacion_requerida()
        if not operacion_id:
            return

        def tarea():
            return self.api_get_portia_plan(operacion_id)

        def al_terminar(data):
            texto = data.get("text", "") if isinstance(data, dict) else str(data)
            self.portia_resultado.delete("1.0", "end")
            self.portia_resultado.insert("1.0", texto or "No se pudo generar plan de acciÃ³n.")
            acciones = data.get("acciones", []) if isinstance(data, dict) else []
            self.portia_status_var.set(f"Plan de acciÃ³n generado: {len(acciones)} prioridades.")

        self.ejecutar_en_segundo_plano(
            "P.O.R.T.I.A",
            "Generando plan de acciÃ³n operativo...",
            tarea,
            al_terminar,
        )

    def portia_crear_acciones_plan_flotante(self):
        operacion_id = self.portia_operacion_requerida()
        if not operacion_id:
            return
        confirmar = messagebox.askyesno(
            "Crear acciones operativas",
            "P.O.R.T.I.A crearÃ¡ tareas operativas ABIERTAS desde el plan de acciÃ³n.\n\nÂ¿Desea continuar?"
        )
        if not confirmar:
            return

        def tarea():
            return self.api_post_portia_crear_acciones_plan(operacion_id)

        def al_terminar(data):
            texto = data.get("text", "") if isinstance(data, dict) else str(data)
            self.portia_resultado.delete("1.0", "end")
            self.portia_resultado.insert("1.0", texto or "Acciones creadas.")
            self.portia_status_var.set(f"Acciones creadas: {data.get('creadas', 0) if isinstance(data, dict) else 0}.")

        self.ejecutar_en_segundo_plano(
            "P.O.R.T.I.A",
            "Creando acciones operativas desde el plan...",
            tarea,
            al_terminar,
        )

    def portia_abrir_notificaciones_flotante(self):
        operacion_id = self.portia_operacion_requerida()
        if not operacion_id:
            return

        ventana = tk.Toplevel(self)
        ventana.title("P.O.R.T.I.A - Notificaciones")
        ventana.geometry("1060x560")
        ventana.configure(bg=self.colors["bg_main"])
        ventana.transient(self)
        self.aplicar_icono_app(ventana)

        header = tk.Frame(ventana, bg=self.colors["bg_topbar"])
        header.pack(fill="x")
        tk.Label(
            header,
            text=f"Notificaciones operativas | OperaciÃ³n {operacion_id}",
            font=("Segoe UI", 16, "bold"),
            bg=self.colors["bg_topbar"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=18, pady=(14, 2))
        status_var = tk.StringVar(value="Presione Cargar notificaciones.")
        tk.Label(
            header,
            textvariable=status_var,
            font=("Segoe UI", 10),
            bg=self.colors["bg_topbar"],
            fg=self.colors["muted"],
        ).pack(anchor="w", padx=18, pady=(0, 12))

        actions = tk.Frame(ventana, bg=self.colors["bg_main"])
        actions.pack(fill="x", padx=18, pady=(12, 8))
        estado_var = tk.StringVar(value="PENDIENTE")
        ttk.Combobox(actions, textvariable=estado_var, values=["PENDIENTE", "ENVIADA", ""], state="readonly", width=14).pack(side="left", padx=(0, 8))

        table_frame = tk.Frame(ventana, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1)
        table_frame.pack(fill="both", expand=True, padx=18, pady=(0, 18))
        cols = ("id", "prioridad", "canal", "destinatario", "asunto", "estado", "creado")
        tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=16)
        for col, heading, width in [
            ("id", "ID", 70),
            ("prioridad", "Prioridad", 100),
            ("canal", "Canal", 100),
            ("destinatario", "Destinatario", 180),
            ("asunto", "Asunto", 420),
            ("estado", "Estado", 120),
            ("creado", "Creado", 160),
        ]:
            tree.heading(col, text=heading)
            tree.column(col, width=width, anchor="center")
        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        def cargar():
            def tarea():
                return self.api_get_portia_notificaciones(operacion_id, estado_var.get())

            def al_terminar(data):
                rows = data.get("data", []) if isinstance(data, dict) else []
                for item in tree.get_children():
                    tree.delete(item)
                for row in rows:
                    tree.insert(
                        "",
                        "end",
                        values=(
                            row.get("id", ""),
                            row.get("prioridad", ""),
                            row.get("canal", ""),
                            row.get("destinatario", ""),
                            row.get("asunto", ""),
                            row.get("estado", ""),
                            self.formatear_fecha(row.get("creado_en")),
                        ),
                    )
                status_var.set(f"Notificaciones cargadas: {len(rows)}.")

            self.ejecutar_en_segundo_plano("P.O.R.T.I.A", "Cargando notificaciones...", tarea, al_terminar)

        def marcar_enviada():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Sin selecciÃ³n", "Seleccione una notificaciÃ³n.")
                return
            notificacion_id = tree.item(selected[0], "values")[0]
            if not messagebox.askyesno("Marcar enviada", f"Â¿Marcar notificaciÃ³n {notificacion_id} como ENVIADA?"):
                return

            def tarea():
                return self.api_post_portia_notificacion_enviada(notificacion_id)

            def al_terminar(_data):
                status_var.set(f"NotificaciÃ³n {notificacion_id} marcada como enviada.")
                cargar()

            self.ejecutar_en_segundo_plano("P.O.R.T.I.A", "Actualizando notificaciÃ³n...", tarea, al_terminar)

        ttk.Button(actions, text="Cargar notificaciones", style="Olive.TButton", command=cargar).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Marcar enviada", style="Gray.TButton", command=marcar_enviada).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Cerrar", style="Gray.TButton", command=ventana.destroy).pack(side="left")

    def portia_abrir_acciones_flotante(self):
        operacion_id = self.portia_operacion_requerida()
        if not operacion_id:
            return

        ventana = tk.Toplevel(self)
        ventana.title("P.O.R.T.I.A - Acciones operativas")
        ventana.geometry("980x560")
        ventana.configure(bg=self.colors["bg_main"])
        ventana.transient(self)
        self.aplicar_icono_app(ventana)

        header = tk.Frame(ventana, bg=self.colors["bg_topbar"])
        header.pack(fill="x")
        tk.Label(
            header,
            text=f"Acciones operativas | OperaciÃ³n {operacion_id}",
            font=("Segoe UI", 16, "bold"),
            bg=self.colors["bg_topbar"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=18, pady=(14, 2))
        status_var = tk.StringVar(value="Presione Cargar acciones.")
        tk.Label(
            header,
            textvariable=status_var,
            font=("Segoe UI", 10),
            bg=self.colors["bg_topbar"],
            fg=self.colors["muted"],
        ).pack(anchor="w", padx=18, pady=(0, 12))

        actions = tk.Frame(ventana, bg=self.colors["bg_main"])
        actions.pack(fill="x", padx=18, pady=(12, 8))
        estado_var = tk.StringVar(value="ABIERTA")
        ttk.Combobox(actions, textvariable=estado_var, values=["ABIERTA", "COMPLETADA", "CERRADA", ""], state="readonly", width=14).pack(side="left", padx=(0, 8))

        table_frame = tk.Frame(ventana, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1)
        table_frame.pack(fill="both", expand=True, padx=18, pady=(0, 18))
        cols = ("id", "prioridad", "area", "titulo", "responsable", "estado", "creado")
        tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=16)
        for col, heading, width in [
            ("id", "ID", 70),
            ("prioridad", "Prioridad", 100),
            ("area", "Ãrea", 130),
            ("titulo", "AcciÃ³n", 360),
            ("responsable", "Responsable", 180),
            ("estado", "Estado", 120),
            ("creado", "Creado", 160),
        ]:
            tree.heading(col, text=heading)
            tree.column(col, width=width, anchor="center")
        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        acciones_cache = {"rows": []}

        def cargar():
            def tarea():
                return self.api_get_portia_acciones(operacion_id, estado_var.get())

            def al_terminar(data):
                rows = data.get("data", []) if isinstance(data, dict) else []
                acciones_cache["rows"] = rows
                for item in tree.get_children():
                    tree.delete(item)
                for row in rows:
                    tree.insert(
                        "",
                        "end",
                        values=(
                            row.get("id", ""),
                            row.get("prioridad", ""),
                            row.get("alerta_tipo", ""),
                            row.get("titulo", ""),
                            row.get("responsable", ""),
                            row.get("estado", ""),
                            self.formatear_fecha(row.get("creado_en")),
                        ),
                    )
                status_var.set(f"Acciones cargadas: {len(rows)}.")

            self.ejecutar_en_segundo_plano("P.O.R.T.I.A", "Cargando acciones operativas...", tarea, al_terminar)

        def completar():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Sin selecciÃ³n", "Seleccione una acciÃ³n.")
                return
            accion_id = tree.item(selected[0], "values")[0]
            if not messagebox.askyesno("Completar acciÃ³n", f"Â¿Marcar acciÃ³n {accion_id} como COMPLETADA?"):
                return

            def tarea():
                return self.api_post_portia_completar_accion(accion_id)

            def al_terminar(_data):
                status_var.set(f"AcciÃ³n {accion_id} completada.")
                cargar()

            self.ejecutar_en_segundo_plano("P.O.R.T.I.A", "Actualizando acciÃ³n...", tarea, al_terminar)

        def escalar():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Sin selecciÃ³n", "Seleccione una acciÃ³n.")
                return
            accion_id = tree.item(selected[0], "values")[0]
            destinatario = simpledialog.askstring(
                "Escalar acciÃ³n",
                "Destinatario sugerido:",
                parent=ventana,
            )
            if destinatario is None:
                return

            def tarea():
                return self.api_post_portia_escalar_accion(accion_id, destinatario.strip())

            def al_terminar(data):
                texto = data.get("text", "") if isinstance(data, dict) else str(data)
                status_var.set(f"AcciÃ³n {accion_id} escalada.")
                self.portia_resultado.delete("1.0", "end")
                self.portia_resultado.insert("1.0", texto or "AcciÃ³n escalada.")

            self.ejecutar_en_segundo_plano("P.O.R.T.I.A", "Escalando acciÃ³n...", tarea, al_terminar)

        ttk.Button(actions, text="Cargar acciones", style="Olive.TButton", command=cargar).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Marcar completada", style="Gray.TButton", command=completar).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Escalar", style="Gray.TButton", command=escalar).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Cerrar", style="Gray.TButton", command=ventana.destroy).pack(side="left")

    def maximizar_ventana(self):
        try:
            self.state("zoomed")
        except Exception:
            ancho = self.winfo_screenwidth()
            alto = self.winfo_screenheight()
            self.geometry(f"{ancho}x{alto}+0+0")

    def confirmar_salida(self):
        salir = messagebox.askyesno(
            "Salir del sistema",
            "Desea salir del sistema?"
        )
        if salir:
            self.destroy()

    def mostrar_modal_proceso(self, titulo, mensaje):
        modal = tk.Toplevel(self)
        modal.title(titulo)
        modal.geometry("420x170")
        modal.resizable(False, False)
        modal.configure(bg=self.colors["bg_card"])
        modal.transient(self)
        modal.grab_set()

        modal.update_idletasks()
        x = self.winfo_rootx() + max(0, (self.winfo_width() - 420) // 2)
        y = self.winfo_rooty() + max(0, (self.winfo_height() - 170) // 2)
        modal.geometry(f"420x170+{x}+{y}")

        tk.Label(
            modal,
            text=titulo,
            font=("Segoe UI", 14, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=22, pady=(22, 6))

        tk.Label(
            modal,
            text=mensaje,
            font=("Segoe UI", 10),
            bg=self.colors["bg_card"],
            fg=self.colors["text_secondary"],
            wraplength=360,
            justify="left",
        ).pack(anchor="w", padx=22, pady=(0, 14))

        barra = ttk.Progressbar(modal, mode="indeterminate")
        barra.pack(fill="x", padx=22, pady=(0, 18))
        barra.start(12)

        modal.protocol("WM_DELETE_WINDOW", lambda: None)
        return modal

    def ejecutar_en_segundo_plano(self, titulo, mensaje, tarea, al_terminar=None):
        modal = self.mostrar_modal_proceso(titulo, mensaje)
        self.configure(cursor="watch")

        def ui_after(callback):
            try:
                if self.winfo_exists():
                    self.after(0, callback)
            except RuntimeError:
                pass
            except tk.TclError:
                pass

        def worker():
            try:
                resultado = tarea()
                ui_after(lambda resultado=resultado: finalizar(resultado, None))
            except Exception as exc:
                ui_after(lambda exc=exc: finalizar(None, exc))

        def finalizar(resultado, error):
            try:
                if modal.winfo_exists():
                    modal.grab_release()
                    modal.destroy()
            except Exception:
                pass

            try:
                if self.winfo_exists():
                    self.configure(cursor="")
            except Exception:
                pass

            if error is not None:
                messagebox.showerror(titulo, str(error))
                return

            if al_terminar:
                al_terminar(resultado)

        threading.Thread(target=worker, daemon=True).start()

    def on_sidebar_click(self, name, command):
        if self._screen_transition_active:
            return

        self.highlight_sidebar_button(name)
        self.render_screen(name, command)

    def render_screen(self, key, command):
        self._screen_transition_active = True
        self.configure(cursor="watch")

        cached = self._screen_cache.get(key)
        if cached is not None:
            self.show_cached_screen(key, cached)
            return

        old_content = self.content
        new_content = tk.Frame(self.main_container, bg=self.colors["bg_main"])
        new_content.pack_propagate(False)
        self._pending_old_content = old_content
        self._pending_new_content = new_content
        self.content = new_content

        self.after(1, lambda: self.build_screen_after_transition(key, command, old_content, new_content))

    def build_screen_after_transition(self, key, command, old_content, new_content):
        if not self._screen_transition_active:
            return

        try:
            command()
            new_content.update_idletasks()
            self._screen_cache[key] = new_content
        except Exception:
            self.content = old_content
            try:
                new_content.destroy()
            except Exception:
                pass
            self.configure(cursor="")
            self._screen_transition_active = False
            raise

        self.after_idle(lambda: self.finish_screen_transition(key, old_content, new_content))

    def show_cached_screen(self, key, cached):
        old_content = self.content

        try:
            if old_content is not None and old_content.winfo_exists() and old_content is not cached:
                old_content.pack_forget()
        except Exception:
            pass

        self.content = cached
        self._current_screen_key = key

        try:
            cached.pack(side="top", fill="both", expand=True)
        except Exception:
            pass

        self.configure(cursor="")
        self._screen_transition_active = False
        self.lift_portia_overlay()

    def finish_screen_transition(self, key=None, old_content=None, new_content=None):
        old_content = old_content or self._pending_old_content
        new_content = new_content or self._pending_new_content or self.content

        try:
            if old_content is not None and old_content.winfo_exists():
                old_content.pack_forget()
        except Exception:
            pass

        try:
            if new_content is not None and new_content.winfo_exists():
                new_content.pack(side="top", fill="both", expand=True)
        except Exception:
            pass

        self.content = new_content
        self._current_screen_key = key
        self._pending_old_content = None
        self._pending_new_content = None
        self.configure(cursor="")
        self._screen_transition_active = False
        self.lift_portia_overlay()

    def lift_portia_overlay(self):
        try:
            if hasattr(self, "portia_panel") and self.portia_panel_visible:
                self.portia_panel.lift()
            if hasattr(self, "portia_float_button"):
                self.portia_float_button.lift()
        except Exception:
            pass

    def start_prebuilding_screens(self):
        return

    def prebuild_next_screen(self):
        if self._prebuilding_screen or self._screen_transition_active:
            self.after(500, self.prebuild_next_screen)
            return

        if not self._prebuild_queue:
            return

        key = self._prebuild_queue.pop(0)
        command = self.screen_commands.get(key)

        if command is None or key in self._screen_cache:
            self.after(250, self.prebuild_next_screen)
            return

        self._prebuilding_screen = True
        old_content = self.content
        old_key = self._current_screen_key
        new_content = tk.Frame(self.main_container, bg=self.colors["bg_main"])
        new_content.pack_propagate(False)

        try:
            self.content = new_content
            self._suppress_sidebar_highlight = True
            command()
            new_content.update_idletasks()
            self._screen_cache[key] = new_content
            self.content = old_content
            self._current_screen_key = old_key
            self._suppress_sidebar_highlight = False
            self.highlight_sidebar_button(old_key or "Centro Ejecutivo")
        except Exception:
            self._suppress_sidebar_highlight = False
            self.content = old_content
            self._current_screen_key = old_key
            try:
                new_content.destroy()
            except Exception:
                pass
        finally:
            self._prebuilding_screen = False
            self.after(650, self.prebuild_next_screen)

    def schedule_ui_job(self, key, callback, delay=35):
        previous = self._ui_after_jobs.get(key)
        if previous is not None:
            try:
                self.after_cancel(previous)
            except Exception:
                pass

        def run():
            self._ui_after_jobs.pop(key, None)
            callback()

        self._ui_after_jobs[key] = self.after(delay, run)

    def safe_canvas_exists(self, canvas):
        try:
            return bool(canvas.winfo_exists())
        except Exception:
            return False

    def schedule_canvas_scrollregion(self, canvas):
        key = f"scrollregion_{id(canvas)}"

        def apply():
            if self.safe_canvas_exists(canvas):
                canvas.configure(scrollregion=canvas.bbox("all"))

        self.schedule_ui_job(key, apply)

    def schedule_canvas_window_width(self, canvas, window_id, width):
        key = f"canvas_width_{id(canvas)}"

        def apply():
            if self.safe_canvas_exists(canvas):
                canvas.itemconfigure(window_id, width=width)

        self.schedule_ui_job(key, apply)

    def bind_scroll_canvas(self, canvas, body, window_id=None, min_width=None):
        if window_id is not None:
            canvas.bind(
                "<Configure>",
                lambda event: self.schedule_canvas_window_width(
                    canvas,
                    window_id,
                    max(event.width, min_width or event.width),
                ),
            )

        def on_mousewheel(event):
            if not self.safe_canvas_exists(canvas):
                return
            if getattr(event, "state", 0) & 0x0001:
                canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")
            else:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def on_shift_mousewheel(event):
            if self.safe_canvas_exists(canvas):
                canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")

        def bind_widget_tree(widget):
            try:
                if not getattr(widget, "_xtravon_scroll_bound", False):
                    widget.bind("<MouseWheel>", on_mousewheel, add="+")
                    widget.bind("<Shift-MouseWheel>", on_shift_mousewheel, add="+")
                    widget._xtravon_scroll_bound = True
            except Exception:
                pass
            for child in widget.winfo_children():
                bind_widget_tree(child)

        def refresh_scroll_bindings(_event=None):
            self.schedule_canvas_scrollregion(canvas)
            bind_widget_tree(body)

        body.bind("<Configure>", refresh_scroll_bindings, add="+")
        bind_widget_tree(body)
        canvas.bind("<MouseWheel>", on_mousewheel, add="+")
        canvas.bind("<Shift-MouseWheel>", on_shift_mousewheel, add="+")

        def bind_all(_event=None):
            canvas.bind_all("<MouseWheel>", on_mousewheel)
            canvas.bind_all("<Shift-MouseWheel>", on_shift_mousewheel)

        def unbind_all(_event=None):
            canvas.unbind_all("<MouseWheel>")
            canvas.unbind_all("<Shift-MouseWheel>")

        for widget in (canvas, body):
            widget.bind("<Enter>", bind_all)
            widget.bind("<Leave>", unbind_all)

    def bind_debounced_draw(self, canvas, draw_func, delay=45):
        key = f"draw_{id(canvas)}"
        canvas.bind("<Configure>", lambda _event: self.schedule_ui_job(key, draw_func, delay))

    def highlight_sidebar_button(self, active_name):
        if getattr(self, "_suppress_sidebar_highlight", False):
            return

        for name, btn in self.sidebar_buttons.items():
            if name == active_name:
                btn.configure(bg=self.colors["accent"], fg=self.colors["bg_main"])
            else:
                btn.configure(bg=self.colors["bg_sidebar"], fg=self.colors["text_dark"])

    def clear_content(self):
        for widget in self.content.winfo_children():
            widget.destroy()

    def create_page_title(self, parent, title, subtitle):
        wrapper = tk.Frame(parent, bg=self.colors["bg_main"])
        wrapper.pack(fill="x", padx=25, pady=(20, 10))
        tk.Label(
            wrapper,
            text=title,
            font=("Segoe UI", 22, "bold"),
            bg=self.colors["bg_main"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w")
        tk.Label(
            wrapper,
            text=subtitle,
            font=("Segoe UI", 10),
            bg=self.colors["bg_main"],
            fg=self.colors["text_secondary"],
        ).pack(anchor="w", pady=(3, 0))

    def create_card(self, parent, title, value, color=None):
        color = color or self.colors["accent"]
        card = tk.Frame(
            parent,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        card.pack(side="left", fill="both", expand=True, padx=8)
        tk.Frame(card, bg=color, height=6).pack(fill="x")
        body = tk.Frame(card, bg=self.colors["bg_card"])
        body.pack(fill="both", expand=True, padx=16, pady=16)
        tk.Label(
            body,
            text=title,
            font=("Segoe UI", 10, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_secondary"],
        ).pack(anchor="w")
        tk.Label(
            body,
            text=str(value),
            font=("Segoe UI", 22, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", pady=(8, 0))

    def color_bodega(self, bodega):
        try:
            numero = int(str(bodega).upper().replace("BODEGA", "").replace("B", "").strip())
        except Exception:
            numero = 0
        palette = {
            1: self.colors["teal"],
            2: self.colors["info"],
            3: self.colors["accent_light"],
            4: self.colors["success"],
            5: self.colors["accent"],
        }
        return palette.get(numero, self.colors["accent"])

    def es_label_bodega(self, label):
        texto = str(label or "").strip().upper()
        return texto.startswith("B") and texto[1:].strip().isdigit()

    def safe_number(self, value, default=0):
        try:
            if value is None or value == "":
                return default
            return float(value)
        except (TypeError, ValueError):
            return default

    def safe_int(self, value, default=0):
        try:
            if value is None or value == "":
                return default
            return int(float(value))
        except (TypeError, ValueError):
            return default

    def formatear_numero(self, value, decimals=2):
        number = self.safe_number(value)
        return f"{number:,.{decimals}f}"

    def formatear_fecha(self, valor):
        if not valor:
            return ""
        return str(valor).replace("T", " ")[:19]

    def estado_visual_qr(self, fila):
        lecturas = self.safe_int(fila.get("lecturas"))
        if lecturas >= 3 or bool(fila.get("qr_bloqueado")):
            return "BLOQUEADA"
        if lecturas == 2:
            return "TERCER_ESCANEO"
        if lecturas == 1:
            return "SEGUNDO_ESCANEO"
        return "PENDIENTE"

    def obtener_detalle_error(self, respuesta):
        try:
            data = respuesta.json()
            return data.get("detail", data)
        except Exception:
            return respuesta.text

    def api_get_boletas(self, params=None):
        respuesta = requests.get(f"{self.api_base}/base-operaciones-camiones", params=params or {}, timeout=60)
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_boletas_filtros(self, params=None):
        respuesta = requests.get(f"{self.api_base}/base-operaciones-camiones/filtros", params=params or {}, timeout=60)
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_boletas_estado_carga(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/base-operaciones-camiones/operacion/{operacion_id}/estado-carga",
            timeout=30,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_dashboard(self, params=None):
        respuesta = requests.get(
            f"{self.api_base}/dashboard/resumen",
            params=params or {},
            timeout=90
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_dashboard_filtros(self, params=None):
        respuesta = requests.get(
            f"{self.api_base}/dashboard/filtros",
            params=params or {},
            timeout=60
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_operaciones_buque(self, estado=None):
        params = {}
        if estado:
            params["estado"] = estado

        respuesta = requests.get(
            f"{self.api_base}/operaciones-buque",
            params=params,
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_operacion_activa(self):
        respuesta = requests.get(
            f"{self.api_base}/operaciones-buque/activa",
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json().get("operacion")

    def api_get_operacion_detalle(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/operaciones-buque/{operacion_id}",
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_despacho_resumen(self, operacion_id=None, filtros=None):
        params = {}
        if not operacion_id:
            activa = self.api_get_operacion_activa()
            if activa:
                operacion_id = self.safe_int(activa.get("id"), None)
        if operacion_id:
            params["operacion_id"] = operacion_id
        for key, value in (filtros or {}).items():
            if value not in (None, ""):
                params[key] = value
        respuesta = requests.get(
            f"{self.api_base}/despacho-viajes/resumen",
            params=params,
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_despacho_filtros(self, operacion_id=None):
        params = {}
        if operacion_id:
            params["operacion_id"] = operacion_id
        respuesta = requests.get(
            f"{self.api_base}/despacho-viajes/filtros",
            params=params,
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_despacho_propuesta(self, operacion_id, chofer, placa, filtros=None):
        params = {"operacion_id": operacion_id, "chofer": chofer, "placa": placa}
        for key, value in (filtros or {}).items():
            if value not in (None, ""):
                params[key] = value
        respuesta = requests.get(
            f"{self.api_base}/despacho-viajes/propuesta",
            params=params,
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_post_despacho_asignar(self, payload):
        respuesta = requests.post(
            f"{self.api_base}/despacho-viajes/asignar",
            json=payload,
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_post_despacho_liberar(self, ids, comentario=""):
        respuesta = requests.post(
            f"{self.api_base}/despacho-viajes/liberar",
            json={"ids": ids, "comentario": comentario, "usuario": "desktop"},
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def asegurar_operacion_despacho(self):
        if getattr(self, "despacho_operacion_id", None):
            return True
        try:
            activa = self.api_get_operacion_activa()
            if activa and activa.get("id"):
                self.despacho_operacion_id = self.safe_int(activa.get("id"), None)
                if hasattr(self, "despacho_estado_label"):
                    self.despacho_estado_label.configure(
                        text=f"Operacion: {activa.get('codigo_operacion', '')} | {activa.get('nombre_buque', '')} | {activa.get('estado', '')}"
                    )
                return True
        except Exception:
            pass
        messagebox.showwarning("Sin operacion", "Primero presione Buscar operacion activa.")
        return False

    def api_post_despacho_reasignar(
        self,
        ids,
        chofer,
        placa,
        comentario="",
        politica_empresa="MISMA_EMPRESA",
        empresa_destino=None,
    ):
        respuesta = requests.post(
            f"{self.api_base}/despacho-viajes/reasignar",
            json={
                "ids": ids,
                "chofer": chofer,
                "placa": placa,
                "comentario": comentario,
                "usuario": "desktop",
                "politica_empresa": politica_empresa,
                "empresa_destino": empresa_destino,
                "activar_qr": True,
            },
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_post_despacho_rechazar_solicitud(self, solicitud_id, comentario=""):
        respuesta = requests.post(
            f"{self.api_base}/despacho-viajes/solicitudes/{solicitud_id}/rechazar",
            json={"comentario": comentario, "usuario": "desktop"},
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_post_despacho_cancelar(self, ids, comentario=""):
        respuesta = requests.post(
            f"{self.api_base}/despacho-viajes/cancelar",
            json={"ids": ids, "comentario": comentario, "usuario": "desktop"},
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_post_despacho_bloquear(self, operacion_id, chofer, placa, motivo):
        respuesta = requests.post(
            f"{self.api_base}/despacho-viajes/bloquear",
            json={
                "operacion_id": operacion_id,
                "chofer": chofer,
                "placa": placa,
                "motivo": motivo,
                "creado_por": "desktop",
            },
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_post_despacho_desbloquear(self, bloqueo_id):
        respuesta = requests.post(
            f"{self.api_base}/despacho-viajes/desbloquear/{bloqueo_id}",
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_reporte_buque(self, operacion_id, params=None):
        respuesta = requests.get(
            f"{self.api_base}/reportes-buque/operacion/{operacion_id}",
            params=params or {},
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_reporte_buque_filtros(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/reportes-buque/operacion/{operacion_id}/filtros",
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_descargar_reporte_buque(self, operacion_id, formato, ruta, params=None):
        respuesta = requests.get(
            f"{self.api_base}/reportes-buque/operacion/{operacion_id}/{formato}",
            params=params or {},
            timeout=120,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        try:
            with open(ruta, "wb") as archivo:
                archivo.write(respuesta.content)
            return ruta
        except PermissionError:
            base, extension = os.path.splitext(ruta)
            alternativa = f"{base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{extension}"
            try:
                with open(alternativa, "wb") as archivo:
                    archivo.write(respuesta.content)
                return alternativa
            except PermissionError as exc:
                raise RuntimeError(
                    "No se pudo guardar el reporte. Cierre el archivo si esta abierto en PDF/Excel/Word "
                    "o seleccione otra carpeta de destino."
                ) from exc

    def api_crear_operacion_buque(self, payload):
        respuesta = requests.post(
            f"{self.api_base}/operaciones-buque",
            json=payload,
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_cerrar_operacion_buque(self, operacion_id):
        respuesta = requests.post(
            f"{self.api_base}/operaciones-buque/{operacion_id}/cerrar",
            json={},
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_reabrir_operacion_buque(self, operacion_id):
        respuesta = requests.post(
            f"{self.api_base}/operaciones-buque/{operacion_id}/reabrir",
            json={"cerrar_otras_abiertas": True},
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_crear_cuota_buque(self, payload):
        respuesta = requests.post(
            f"{self.api_base}/operaciones-buque-cuotas",
            json=payload,
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_cuotas_buque(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/operaciones-buque-cuotas/operacion/{operacion_id}",
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_eliminar_cuota_buque(self, cuota_id):
        respuesta = requests.delete(
            f"{self.api_base}/operaciones-buque-cuotas/{cuota_id}",
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_actualizar_cuota_buque(self, cuota_id, payload):
        respuesta = requests.patch(
            f"{self.api_base}/operaciones-buque-cuotas/{cuota_id}",
            json=payload,
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_qr_seguridad_estado(self):
        respuesta = requests.get(
            f"{self.api_base}/base-operaciones-camiones/qr/seguridad/estado",
            timeout=30,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def show_dashboard(self):
        self.clear_content()
        self.highlight_sidebar_button("Dashboard")

        self.dashboard_filters_loaded = False
        self.dashboard_filter_vars = {}
        self.dashboard_filter_widgets = {}

        self.create_page_title(
            self.content,
            "Dashboard",
            "Seleccione filtros y presione Crear data para cargar los indicadores.",
        )

        filtros_panel = tk.Frame(
            self.content,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        filtros_panel.pack(fill="x", padx=25, pady=(0, 10))

        filtros_grid = tk.Frame(filtros_panel, bg=self.colors["bg_card"])
        filtros_grid.pack(fill="x", padx=14, pady=12)

        campos = [
            ("buque", "Buque"),
            ("cliente", "Cliente"),
            ("producto", "Producto"),
            ("chofer", "Chofer"),
            ("placa", "Placa"),
            ("estado", "Estado"),
            ("fecha_desde", "Fecha desde"),
            ("fecha_hasta", "Fecha hasta"),
        ]

        for idx, (key, label) in enumerate(campos):
            row = idx // 4
            col = idx % 4

            box = tk.Frame(filtros_grid, bg=self.colors["bg_card"])
            box.grid(row=row, column=col, sticky="ew", padx=6, pady=5)

            tk.Label(
                box,
                text=label,
                font=("Segoe UI", 9, "bold"),
                bg=self.colors["bg_card"],
                fg=self.colors["text_dark"],
            ).pack(anchor="w")

            var = tk.StringVar()
            self.dashboard_filter_vars[key] = var

            if key in ("fecha_desde", "fecha_hasta"):
                widget = ttk.Entry(box, textvariable=var)
            else:
                widget = ttk.Combobox(box, textvariable=var, state="readonly")
                widget.bind("<<ComboboxSelected>>", lambda _e: None)

            widget.pack(fill="x")
            self.dashboard_filter_widgets[key] = widget

        for col in range(4):
            filtros_grid.grid_columnconfigure(col, weight=1)

        actions = tk.Frame(filtros_panel, bg=self.colors["bg_card"])
        actions.pack(fill="x", padx=14, pady=(0, 12))

        ttk.Button(
            actions,
            text="Crear data",
            style="Olive.TButton",
            command=self.cargar_dashboard,
        ).pack(side="left", padx=(0, 10))

        ttk.Button(
            actions,
            text="Limpiar filtros",
            style="Gray.TButton",
            command=self.limpiar_filtros_dashboard,
        ).pack(side="left")

        wrapper = tk.Frame(self.content, bg=self.colors["bg_main"])
        wrapper.pack(fill="both", expand=True, padx=25, pady=(0, 20))

        self.dashboard_canvas = tk.Canvas(wrapper, bg=self.colors["bg_main"], highlightthickness=0)
        scroll_y = ttk.Scrollbar(wrapper, orient="vertical", command=self.dashboard_canvas.yview)
        self.dashboard_body = tk.Frame(self.dashboard_canvas, bg=self.colors["bg_main"])

        canvas_window = self.dashboard_canvas.create_window(
            (0, 0),
            window=self.dashboard_body,
            anchor="nw",
        )

        self.dashboard_canvas.configure(yscrollcommand=scroll_y.set)
        self.bind_scroll_canvas(self.dashboard_canvas, self.dashboard_body, canvas_window)

        self.dashboard_canvas.pack(side="left", fill="both", expand=True)
        scroll_y.pack(side="right", fill="y")

        self.render_dashboard_placeholder()

    def obtener_params_dashboard(self):
        params = {}

        for key, var in self.dashboard_filter_vars.items():
            value = var.get().strip()
            if value:
                params[key] = value

        return params

    def limpiar_filtros_dashboard(self):
        for var in self.dashboard_filter_vars.values():
            var.set("")

        self.dashboard_filters_loaded = False
        if self.dashboard_body is None:
            self.render_dashboard_placeholder()
            return

        self.cargar_dashboard()

    def actualizar_filtros_dashboard(self):
        if not self.dashboard_filters_loaded:
            return

        try:
            data = self.api_get_dashboard_filtros(self.obtener_params_dashboard())
            self.aplicar_opciones_filtros_dashboard(data.get("opciones", {}))
        except Exception as e:
            messagebox.showerror("Error filtros dashboard", str(e))

    def aplicar_opciones_filtros_dashboard(self, opciones):
        mapa = {
            "buque": "buques",
            "cliente": "clientes",
            "producto": "productos",
            "chofer": "choferes",
            "placa": "placas",
            "estado": "estados",
        }

        for key, option_key in mapa.items():
            widget = self.dashboard_filter_widgets.get(key)
            var = self.dashboard_filter_vars.get(key)

            if widget is None or var is None:
                continue

            actual = var.get()
            valores = opciones.get(option_key, []) or []
            valores = [str(v) for v in valores if v not in (None, "")]

            widget["values"] = valores

            if actual and actual not in valores:
                var.set("")

    def cargar_dashboard(self):
        if self.dashboard_body is None:
            return

        for widget in self.dashboard_body.winfo_children():
            widget.destroy()

        tk.Label(
            self.dashboard_body,
            text="Creando data del dashboard...",
            font=("Segoe UI", 12, "bold"),
            bg=self.colors["bg_main"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", pady=15)

        self.update_idletasks()

        params = self.obtener_params_dashboard()

        def tarea():
            return {
                "filtros": self.api_get_dashboard_filtros(params),
                "resumen": self.api_get_dashboard(params),
            }

        def al_terminar(resultado):
            self.dashboard_filters_loaded = True
            self.aplicar_opciones_filtros_dashboard(resultado.get("filtros", {}).get("opciones", {}))
            self.render_dashboard(resultado.get("resumen", {}))

        self.ejecutar_en_segundo_plano(
            "Dashboard",
            "Creando data desde el backend. Por favor espere...",
            tarea,
            al_terminar,
        )

    def render_dashboard_placeholder(self):
        for widget in self.dashboard_body.winfo_children():
            widget.destroy()

        panel = tk.Frame(
            self.dashboard_body,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        panel.pack(fill="x", pady=(0, 12))

        tk.Label(
            panel,
            text="Dashboard listo para crear data",
            font=("Segoe UI", 13, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=16, pady=(14, 4))

        tk.Label(
            panel,
            text="Seleccione filtros si aplica y presione Crear data. No se cargan KPIs ni grÃ¡ficos automÃ¡ticamente.",
            font=("Segoe UI", 10),
            bg=self.colors["bg_card"],
            fg=self.colors["text_secondary"],
        ).pack(anchor="w", padx=16, pady=(0, 14))

    def render_dashboard_error(self, mensaje):
        for widget in self.dashboard_body.winfo_children():
            widget.destroy()

        panel = tk.Frame(
            self.dashboard_body,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        panel.pack(fill="x", pady=(0, 12))

        tk.Label(
            panel,
            text="No se pudo cargar el dashboard",
            font=("Segoe UI", 13, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["danger"],
        ).pack(anchor="w", padx=16, pady=(14, 4))

        tk.Label(
            panel,
            text=mensaje,
            font=("Segoe UI", 10),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
            wraplength=900,
            justify="left",
        ).pack(anchor="w", padx=16, pady=(0, 14))

    def obtener_lista_grafico(self, graficos, *keys):
        for key in keys:
            value = graficos.get(key)
            if isinstance(value, list) and value:
                return value
        return []

    def render_dashboard(self, data):
        for widget in self.dashboard_body.winfo_children():
            widget.destroy()

        kpis = data.get("kpis", {}) if isinstance(data, dict) else {}
        graficos = data.get("graficos", {}) if isinstance(data, dict) else {}

        cards_frame = tk.Frame(self.dashboard_body, bg=self.colors["bg_main"])
        cards_frame.pack(fill="x", pady=(0, 12))

        self.create_card(cards_frame, "Viajes totales", self.formatear_numero(kpis.get("total_viajes")), self.colors["accent"])
        self.create_card(cards_frame, "Completos", self.formatear_numero(kpis.get("viajes_completos")), self.colors["success"])
        self.create_card(cards_frame, "Pendientes", self.formatear_numero(kpis.get("viajes_pendientes")), self.colors["warning"])
        self.create_card(cards_frame, "Avance", f"{self.safe_number(kpis.get('avance_operativo_pct')):,.2f}%", self.colors["info"])

        cards_frame_2 = tk.Frame(self.dashboard_body, bg=self.colors["bg_main"])
        cards_frame_2.pack(fill="x", pady=(0, 18))

        self.create_card(cards_frame_2, "Peso cargado", self.formatear_numero(kpis.get("peso_cargado_total"), 2), self.colors["success"])
        self.create_card(cards_frame_2, "Peso promedio", self.formatear_numero(kpis.get("peso_cargado_promedio"), 2), self.colors["info"])
        self.create_card(cards_frame_2, "DuraciÃ³n prom.", f"{self.safe_number(kpis.get('duracion_promedio_min')):,.2f} min", self.colors["accent"])
        self.create_card(cards_frame_2, "QR bloqueados", self.formatear_numero(kpis.get("qr_bloqueados")), self.colors["warning"])

        silueta_panel = tk.Frame(
            self.dashboard_body,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        silueta_panel.pack(fill="x", pady=(0, 18))
        tk.Label(
            silueta_panel,
            text="Silueta del buque - pendiente por bodega",
            font=("Segoe UI", 13, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=14, pady=(12, 4))
        silueta_canvas = tk.Canvas(silueta_panel, bg=self.colors["bg_card"], height=230, highlightthickness=0)
        silueta_canvas.pack(fill="x", padx=14, pady=(0, 14))
        silueta_canvas.after(80, lambda c=silueta_canvas, rows=self.obtener_lista_grafico(graficos, "silueta_bodegas"): self.dibujar_silueta_pendiente_bodegas(c, rows))

        charts_grid = tk.Frame(self.dashboard_body, bg=self.colors["bg_main"])
        charts_grid.pack(fill="both", expand=True)

        self.crear_chart_barras(
            charts_grid,
            "Viajes por producto",
            self.obtener_lista_grafico(graficos, "barras_viajes_por_producto", "viajes_por_producto"),
            "producto",
            "viajes",
            0,
            0,
        )

        self.crear_chart_barras(
            charts_grid,
            "Peso por producto",
            self.obtener_lista_grafico(graficos, "barras_peso_por_producto", "peso_por_producto", "viajes_por_producto"),
            "producto",
            "peso_cargado",
            0,
            1,
        )

        self.crear_chart_lineal(
            charts_grid,
            "Tendencia por fecha",
            self.obtener_lista_grafico(graficos, "lineal_tendencia_por_fecha", "tendencia_por_fecha"),
            "fecha",
            "viajes",
            1,
            0,
        )

        self.crear_chart_circular(
            charts_grid,
            "Estado operativo",
            self.obtener_lista_grafico(graficos, "circular_estado_operativo", "estado_operativo"),
            "estado",
            "viajes",
            1,
            1,
        )

        self.crear_chart_circular(
            charts_grid,
            "Estado de escaneos",
            self.obtener_lista_grafico(graficos, "circular_estado_escaneos", "estado_escaneos"),
            "etapa",
            "viajes",
            2,
            0,
        )

        self.crear_chart_barras(
            charts_grid,
            "Viajes por cliente",
            self.obtener_lista_grafico(graficos, "barras_viajes_por_cliente", "viajes_por_cliente"),
            "cliente",
            "viajes",
            2,
            1,
        )

        self.crear_chart_barras(
            charts_grid,
            "DuraciÃ³n por camiÃ³n",
            self.obtener_lista_grafico(graficos, "barras_duracion_por_camion", "duracion_por_camion"),
            "camion",
            "duracion_min",
            3,
            0,
        )

        for col in range(2):
            charts_grid.grid_columnconfigure(col, weight=1, uniform="charts")

        for row in range(4):
            charts_grid.grid_rowconfigure(row, weight=1, minsize=290)

        self.dashboard_canvas.update_idletasks()
        self.dashboard_canvas.configure(scrollregion=self.dashboard_canvas.bbox("all"))

    def crear_panel_chart(self, parent, title, row, col):
        panel = tk.Frame(
            parent,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        panel.grid(row=row, column=col, sticky="nsew", padx=8, pady=8)

        tk.Label(
            panel,
            text=title,
            font=("Segoe UI", 12, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=14, pady=(12, 4))

        canvas = tk.Canvas(panel, bg=self.colors["bg_card"], height=230, highlightthickness=0)
        canvas.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        return canvas

    def dibujar_silueta_pendiente_bodegas(self, canvas, rows):
        canvas.delete("all")
        width = max(canvas.winfo_width(), 620)
        height = max(canvas.winfo_height(), 220)
        rows = [row for row in (rows or []) if isinstance(row, dict)]
        por_bodega = {}
        for row in rows:
            bodega = int(row.get("bodega_numero") or 0)
            if bodega <= 0:
                continue
            capacidad = self.safe_number(row.get("capacidad_mt"), 0)
            descargado = self.safe_number(row.get("descargado_mt", row.get("retirado_mt")), 0)
            pendiente = self.safe_number(row.get("pendiente_mt"), max(capacidad - descargado, 0))
            item = por_bodega.setdefault(bodega, {"capacidad": 0, "descargado": 0, "pendiente": 0, "productos": set()})
            item["capacidad"] += capacidad
            item["descargado"] += descargado
            item["pendiente"] += max(pendiente, 0)
            if row.get("producto"):
                item["productos"].add(str(row.get("producto")))

        y_top = 34
        y_bottom = height - 24
        y_mid = (y_top + y_bottom) / 2
        x0 = 18
        x1 = width - 18
        canvas.create_polygon(x0, y_top, x1 - 45, y_top, x1, y_mid, x1 - 45, y_bottom, x0, y_bottom, fill=self.colors["bg_topbar"], outline=self.colors["accent"], width=2)
        usable_w = x1 - x0 - 70
        seg_w = usable_w / 5
        palette = {1: self.colors["teal"], 2: self.colors["info"], 3: self.colors["accent_light"], 4: self.colors["success"], 5: self.colors["accent"]}

        for idx, numero in enumerate([5, 4, 3, 2, 1]):
            x = x0 + 16 + idx * seg_w
            rx0 = x
            rx1 = x + seg_w - 10
            ry0 = y_top + 16
            ry1 = y_bottom - 16
            item = por_bodega.get(numero, {})
            capacidad = max(self.safe_number(item.get("capacidad"), 0), 0)
            descargado = max(self.safe_number(item.get("descargado"), 0), 0)
            pendiente = max(self.safe_number(item.get("pendiente"), max(capacidad - descargado, 0)), 0)
            pendiente_pct = (pendiente / capacidad * 100) if capacidad else 0
            restante_ratio = min(max(pendiente / capacidad, 0), 1) if capacidad else 0
            canvas.create_rectangle(rx0, ry0, rx1, ry1, fill="#FFFFFF", outline=self.colors["border"])
            if restante_ratio > 0:
                fill_top = ry1 - (ry1 - ry0) * restante_ratio
                canvas.create_rectangle(rx0, fill_top, rx1, ry1, fill=palette.get(numero, self.colors["accent"]), outline="")
            if descargado > 0 and capacidad > 0:
                nivel_y = ry1 - (ry1 - ry0) * restante_ratio
                canvas.create_line(rx0 + 2, nivel_y, rx1 - 2, nivel_y, fill=self.colors["text_light"], width=2)
            productos = ", ".join(sorted(item.get("productos", []))[:2]) if item else ""
            label = f"B{numero}\nPend. {pendiente:,.0f} MT\n{pendiente_pct:,.1f}%"
            if capacidad:
                label += f"\nTot. {capacidad:,.0f}"
            if productos:
                label += f"\n{productos[:14]}"
            canvas.create_text((rx0 + rx1) / 2, y_mid, text=label, fill="#050B14", font=("Segoe UI", 8, "bold"), justify="center", width=max(rx1 - rx0 - 6, 62))

    def normalizar_items(self, data, label_key, value_key, limit=8):
        if not isinstance(data, list):
            return []

        items = []
        for item in data[:limit]:
            if not isinstance(item, dict):
                continue

            label = str(item.get(label_key, "SIN DATO") or "SIN DATO")
            value = self.safe_number(item.get(value_key), 0)

            if value > 0:
                items.append((label, value))

        return items

    def crear_chart_barras(self, parent, title, data, label_key, value_key, row, col):
        canvas = self.crear_panel_chart(parent, title, row, col)
        items = self.normalizar_items(data, label_key, value_key)

        if not items:
            self.dibujar_sin_datos(canvas)
            return

        def draw(_event=None):
            canvas.delete("all")

            width = max(canvas.winfo_width(), 360)
            height = max(canvas.winfo_height(), 220)

            max_label_len = max((len(label) for label, _value in items), default=12)
            left = min(max(140, max_label_len * 8 + 18), int(width * 0.42))
            right = 96
            top = 16
            bar_h = 22
            gap = 10
            max_value = max([value for _, value in items] or [1]) or 1

            for idx, (label, value) in enumerate(items):
                y = top + idx * (bar_h + gap)

                if y + bar_h > height - 8:
                    break

                chart_w = max(width - left - right, 70)
                bar_w = int(chart_w * (value / max_value))
                max_label_chars = max(12, min(28, int((left - 18) / 7)))
                short_label = label[:max_label_chars] + "..." if len(label) > max_label_chars + 3 else label
                value_text = self.formatear_numero(value)
                value_anchor = "w"
                value_x = left + max(bar_w, 2) + 6
                if value_x + (len(value_text) * 7) > width - 8:
                    value_anchor = "e"
                    value_x = width - 8

                canvas.create_text(
                    8,
                    y + bar_h / 2,
                    text=short_label,
                    anchor="w",
                    fill=self.colors["text_dark"],
                    font=("Segoe UI", 9),
                )

                canvas.create_rectangle(
                    left,
                    y,
                    left + max(bar_w, 2),
                    y + bar_h,
                    fill=self.color_bodega(label) if "bodega" in str(title).lower() or self.es_label_bodega(label) else self.colors["accent"],
                    outline="",
                )

                canvas.create_text(
                    value_x,
                    y + bar_h / 2,
                    text=value_text,
                    anchor=value_anchor,
                    fill=self.colors["text_secondary"],
                    font=("Segoe UI", 9, "bold"),
                )

        self.bind_debounced_draw(canvas, draw)
        canvas.after(50, draw)

    def crear_chart_lineal(self, parent, title, data, label_key, value_key, row, col):
        canvas = self.crear_panel_chart(parent, title, row, col)
        items = self.normalizar_items(data, label_key, value_key, limit=12)

        if not items:
            self.dibujar_sin_datos(canvas)
            return

        def draw(_event=None):
            canvas.delete("all")

            width = max(canvas.winfo_width(), 360)
            height = max(canvas.winfo_height(), 220)

            pad_x = 42
            pad_y = 28
            graph_w = width - (pad_x * 2)
            graph_h = height - (pad_y * 2)
            max_value = max([value for _, value in items] or [1]) or 1

            canvas.create_line(pad_x, height - pad_y, width - pad_x, height - pad_y, fill=self.colors["border"])
            canvas.create_line(pad_x, pad_y, pad_x, height - pad_y, fill=self.colors["border"])

            points = []

            for idx, (_label, value) in enumerate(items):
                x = pad_x + (graph_w * idx / max(len(items) - 1, 1))
                y = height - pad_y - (graph_h * value / max_value)
                points.append((x, y, value))

            for idx in range(len(points) - 1):
                canvas.create_line(
                    points[idx][0],
                    points[idx][1],
                    points[idx + 1][0],
                    points[idx + 1][1],
                    fill=self.colors["info"],
                    width=3,
                )

            for x, y, value in points:
                canvas.create_oval(x - 4, y - 4, x + 4, y + 4, fill=self.colors["accent"], outline="")
                canvas.create_text(x, y - 12, text=self.formatear_numero(value), fill=self.colors["text_secondary"], font=("Segoe UI", 8, "bold"))

            canvas.create_text(pad_x, height - 10, text=str(items[0][0])[:10], anchor="w", fill=self.colors["text_secondary"], font=("Segoe UI", 8))
            canvas.create_text(width - pad_x, height - 10, text=str(items[-1][0])[:10], anchor="e", fill=self.colors["text_secondary"], font=("Segoe UI", 8))

        self.bind_debounced_draw(canvas, draw)
        canvas.after(50, draw)

    def crear_chart_circular(self, parent, title, data, label_key, value_key, row, col):
        canvas = self.crear_panel_chart(parent, title, row, col)
        items = self.normalizar_items(data, label_key, value_key, limit=6)
        total = sum(value for _, value in items)

        if not items or total <= 0:
            self.dibujar_sin_datos(canvas)
            return

        palette = [
            self.colors["accent"],
            self.colors["success"],
            self.colors["warning"],
            self.colors["info"],
            self.colors["muted"],
            self.colors["danger"],
        ]

        def draw(_event=None):
            canvas.delete("all")

            height = max(canvas.winfo_height(), 220)
            size = min(150, height - 40)
            x0 = 24
            y0 = 28
            start = 90

            for idx, (label, value) in enumerate(items):
                extent = 360 * value / total
                color = self.color_bodega(label) if "bodega" in str(title).lower() or self.es_label_bodega(label) else palette[idx % len(palette)]
                if len(items) == 1 or abs(extent) >= 359.9:
                    canvas.create_oval(
                        x0,
                        y0,
                        x0 + size,
                        y0 + size,
                        fill=color,
                        outline=self.colors["bg_card"],
                        width=2,
                    )
                else:
                    canvas.create_arc(
                        x0,
                        y0,
                        x0 + size,
                        y0 + size,
                        start=start,
                        extent=-extent,
                        fill=color,
                        outline=self.colors["bg_card"],
                        width=2,
                    )
                start -= extent

            canvas.create_text(
                x0 + size / 2,
                y0 + size / 2,
                text=self.formatear_numero(total),
                fill=self.colors["bg_main"],
                font=("Segoe UI", 10, "bold"),
            )

            legend_x = x0 + size + 24
            legend_y = y0 + 8

            for idx, (label, value) in enumerate(items):
                y = legend_y + idx * 27
                short_label = label[:22] + "..." if len(label) > 25 else label

                canvas.create_rectangle(
                    legend_x,
                    y,
                    legend_x + 12,
                    y + 12,
                    fill=self.color_bodega(label) if "bodega" in str(title).lower() or self.es_label_bodega(label) else palette[idx % len(palette)],
                    outline="",
                )

                canvas.create_text(
                    legend_x + 18,
                    y + 6,
                    text=f"{short_label}: {self.formatear_numero(value)}",
                    anchor="w",
                    fill=self.colors["text_dark"],
                    font=("Segoe UI", 9),
                )

        self.bind_debounced_draw(canvas, draw)
        canvas.after(50, draw)

    def dibujar_sin_datos(self, canvas):
        def draw(_event=None):
            canvas.delete("all")
            width = max(canvas.winfo_width(), 360)
            height = max(canvas.winfo_height(), 220)

            canvas.create_text(
                width / 2,
                height / 2,
                text="Sin datos para mostrar",
                fill=self.colors["text_aux"],
                font=("Segoe UI", 11, "bold"),
            )

        self.bind_debounced_draw(canvas, draw)
        canvas.after(50, draw)





    def refrescar_operaciones_buque(self):
        def tarea():
            return {
                "operaciones": self.api_get_operaciones_buque(),
                "activa": self.api_get_operacion_activa(),
            }

        def al_terminar(resultado):
            data = resultado.get("operaciones", {})
            self.operaciones_cache = data.get("data", [])
            self.operacion_activa = resultado.get("activa")
            self.cargar_operaciones_en_tabla()
            self.actualizar_operacion_activa_label()

        self.ejecutar_en_segundo_plano(
            "Operaciones",
            "Buscando operaciones en el backend...",
            tarea,
            al_terminar,
        )


    def cargar_operaciones_en_tabla(self):
        if self.operaciones_tree is None:
            return

        for item in self.operaciones_tree.get_children():
            self.operaciones_tree.delete(item)

        for operacion in self.operaciones_cache:
            self.operaciones_tree.insert(
                "",
                "end",
                values=(
                    operacion.get("id", ""),
                    operacion.get("codigo_operacion", ""),
                    operacion.get("nombre_buque", ""),
                    operacion.get("fecha_inicio", ""),
                    operacion.get("fecha_cierre", ""),
                    operacion.get("producto", ""),
                    operacion.get("estado", ""),
                ),
            )

    def obtener_operacion_seleccionada_id(self):
        if self.operaciones_tree is None:
            return None

        seleccion = self.operaciones_tree.selection()
        if not seleccion:
            return None

        valores = self.operaciones_tree.item(seleccion[0], "values")
        if not valores:
            return None

        return self.safe_int(valores[0], None)

    def cerrar_operacion_seleccionada(self):
        operacion_id = self.obtener_operacion_seleccionada_id()
        if not operacion_id:
            messagebox.showwarning("Sin selecciÃ³n", "Seleccione una operaciÃ³n.")
            return

        if not messagebox.askyesno("Cerrar operaciÃ³n", "Â¿Desea cerrar y archivar esta operaciÃ³n?"):
            return

        try:
            self.api_cerrar_operacion_buque(operacion_id)
            self.refrescar_operaciones_buque()
            messagebox.showinfo("OperaciÃ³n cerrada", "La operaciÃ³n fue cerrada correctamente.")
        except Exception as e:
            messagebox.showerror("Error cerrando operaciÃ³n", str(e))

    def reabrir_operacion_seleccionada(self):
        operacion_id = self.obtener_operacion_seleccionada_id()
        if not operacion_id:
            messagebox.showwarning("Sin selecciÃ³n", "Seleccione una operaciÃ³n.")
            return

        if not messagebox.askyesno("Reabrir operaciÃ³n", "Â¿Desea reabrir esta operaciÃ³n y cerrar otras abiertas?"):
            return

        try:
            self.api_reabrir_operacion_buque(operacion_id)
            self.refrescar_operaciones_buque()
            messagebox.showinfo("OperaciÃ³n reabierta", "La operaciÃ³n fue reabierta correctamente.")
        except Exception as e:
            messagebox.showerror("Error reabriendo operaciÃ³n", str(e))

    def cargar_cuotas_operacion_seleccionada(self, silencioso=False):
        operacion_id = self.obtener_operacion_seleccionada_id()
        if not operacion_id:
            if not silencioso:
                messagebox.showwarning("Sin selecciÃ³n", "Seleccione una operaciÃ³n.")
            return

        try:
            data = self.api_get_operacion_detalle(operacion_id)
            cuotas = data.get("cuotas", [])
            self.cargar_cuotas_en_tabla(cuotas)
        except Exception as e:
            if not silencioso:
                messagebox.showerror("Error cuotas", str(e))

    def cargar_cuotas_en_tabla(self, cuotas):
        if self.cuotas_tree is None:
            return

        for item in self.cuotas_tree.get_children():
            self.cuotas_tree.delete(item)

        for cuota in cuotas:
            self.cuotas_tree.insert(
                "",
                "end",
                values=(
                    cuota.get("id", ""),
                    cuota.get("cliente", ""),
                    self.producto_cuota_visible(cuota.get("producto")),
                    self.bodega_cuota_visible(cuota.get("bodega_numeros") or cuota.get("bodega_numero")),
                    self.formatear_numero(cuota.get("cuota"), 2),
                    cuota.get("unidad", ""),
                ),
            )



    def show_boletas(self):
        self.clear_content()
        self.highlight_sidebar_button("Carga de Boletas")
        self.create_page_title(
            self.content,
            "Carga de Boletas",
            "Carga el Excel a la base de datos, consulta la tabla, genera QR puros y exporta datos.",
        )

        op_panel = tk.Frame(
            self.content,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        op_panel.pack(fill="x", padx=25, pady=(0, 10))

        self.operacion_activa = None
        self.boletas_operaciones_abiertas = []
        self.boletas_operacion_var = tk.StringVar()
        self.boletas_operacion_label = tk.Label(
            op_panel,
            text="Operacion activa: presione Buscar operacion activa para consultar.",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        )
        self.boletas_operacion_label.pack(side="left", anchor="w", padx=14, pady=10)
        self.boletas_operacion_combo = ttk.Combobox(
            op_panel,
            textvariable=self.boletas_operacion_var,
            state="readonly",
            width=42,
        )
        self.boletas_operacion_combo.pack(side="right", padx=(0, 14), pady=8)
        self.boletas_operacion_combo.bind("<<ComboboxSelected>>", lambda _e: self.seleccionar_operacion_boletas())
        ttk.Button(
            op_panel,
            text="Buscar operacion activa",
            style="Gray.TButton",
            command=self.buscar_operacion_activa_boletas,
        ).pack(side="right", padx=14, pady=8)

        self.boletas_filter_vars = {key: tk.StringVar() for key in ("empresa", "guia", "producto", "chofer", "placa")}
        self.boletas_filter_widgets = {}
        filtros_panel = tk.Frame(
            self.content,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        filtros_panel.pack(fill="x", padx=25, pady=(0, 10))
        filtros_grid = tk.Frame(filtros_panel, bg=self.colors["bg_card"])
        filtros_grid.pack(fill="x", padx=12, pady=10)
        for idx, (key, label) in enumerate([
            ("empresa", "Empresa"),
            ("guia", "Guia"),
            ("producto", "Producto"),
            ("chofer", "Chofer"),
            ("placa", "Placa"),
        ]):
            box = tk.Frame(filtros_grid, bg=self.colors["bg_card"])
            box.grid(row=idx // 5, column=idx % 5, sticky="ew", padx=5, pady=4)
            tk.Label(box, text=label, font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
            combo = ttk.Combobox(box, textvariable=self.boletas_filter_vars[key], values=[""], state="normal")
            combo.pack(fill="x")
            self.boletas_filter_widgets[key] = combo
        for col in range(5):
            filtros_grid.grid_columnconfigure(col, weight=1)
        filtros_actions = tk.Frame(filtros_panel, bg=self.colors["bg_card"])
        filtros_actions.pack(fill="x", padx=12, pady=(0, 10))
        ttk.Button(filtros_actions, text="Cargar filtros", style="Gray.TButton", command=self.cargar_filtros_boletas).pack(side="left", padx=(0, 8))
        ttk.Button(filtros_actions, text="Buscar tabla", style="Olive.TButton", command=self.refrescar_boletas).pack(side="left", padx=(0, 8))
        ttk.Button(filtros_actions, text="Limpiar", style="Gray.TButton", command=self.limpiar_filtros_boletas).pack(side="left")

        actions = tk.Frame(self.content, bg=self.colors["bg_main"])
        actions.pack(fill="x", padx=25, pady=(5, 10))
        ttk.Button(actions, text="Abrir Template", style="Gray.TButton", command=self.abrir_template_boletas_remoto).pack(side="left", padx=(0, 10))
        ttk.Button(actions, text="Cargar Excel", style="Olive.TButton", command=self.cargar_excel_template).pack(side="left", padx=(0, 10))
        ttk.Button(actions, text="Cargar Tabla", style="Gray.TButton", command=self.refrescar_boletas).pack(side="left", padx=(0, 10))
        ttk.Button(actions, text="Ver", style="Gray.TButton", command=self.ver_guia_seleccionada).pack(side="left", padx=(0, 10))
        ttk.Button(actions, text="Exportar Excel", style="Gray.TButton", command=self.exportar_excel).pack(side="left", padx=(0, 10))

        main_panel = tk.Frame(
            self.content,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        main_panel.pack(fill="both", expand=True, padx=25, pady=(5, 20))
        tk.Label(
            main_panel,
            text="Boletas en base_operaciones_camiones",
            font=("Segoe UI", 13, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=15, pady=15)

        columns = ("id", "guia", "numero_embarque", "bodega_numero", "empresa", "buque", "fecha", "producto", "chofer", "placa", "estado", "lecturas", "etapa_qr", "numero_tolva", "qr_bloqueado")
        table_frame = tk.Frame(main_panel, bg=self.colors["bg_card"])
        table_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=22, selectmode="extended")

        headings = {
            "id": "ID",
            "guia": "Guia",
            "numero_embarque": "Embarque",
            "bodega_numero": "Bodega",
            "empresa": "Empresa",
            "buque": "Buque",
            "fecha": "Fecha",
            "producto": "Producto",
            "chofer": "Chofer",
            "placa": "Placa",
            "estado": "Estado",
            "lecturas": "Lecturas",
            "etapa_qr": "Etapa QR",
            "numero_tolva": "Tolva",
            "qr_bloqueado": "Bloqueado",
        }
        widths = {
            "id": 60,
            "guia": 90,
            "numero_embarque": 110,
            "bodega_numero": 75,
            "empresa": 140,
            "buque": 130,
            "fecha": 100,
            "producto": 120,
            "chofer": 140,
            "placa": 90,
            "estado": 120,
            "lecturas": 70,
            "etapa_qr": 130,
            "numero_tolva": 90,
            "qr_bloqueado": 90,
        }

        for col in columns:
            self.tree.heading(col, text=headings[col])
            self.tree.column(col, width=widths[col], anchor="center")

        scroll_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        scroll_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        self.boletas_cache = []



    def buscar_operacion_activa_boletas(self):
        try:
            data = self.api_get_operaciones_buque(estado="ABIERTA")
            if isinstance(data, dict):
                abiertas = data.get("data", [])
            elif isinstance(data, list):
                abiertas = data
            else:
                abiertas = []

            self.boletas_operaciones_abiertas = abiertas if isinstance(abiertas, list) else []
            if not self.boletas_operaciones_abiertas:
                activa = self.api_get_operacion_activa()
                if isinstance(activa, dict) and activa.get("id"):
                    self.boletas_operaciones_abiertas = [activa]

            opciones = [
                f"{op.get('id')} | {op.get('codigo_operacion', '')} | {op.get('nombre_buque', '')}"
                for op in self.boletas_operaciones_abiertas
            ]
            if hasattr(self, "boletas_operacion_combo"):
                self.boletas_operacion_combo["values"] = opciones
            if self.boletas_operaciones_abiertas:
                self.operacion_activa = self.boletas_operaciones_abiertas[0]
                self.boletas_operacion_var.set(opciones[0])
            else:
                self.operacion_activa = None
            if self.operacion_activa:
                texto = (
                    f"Operacion activa: {self.operacion_activa.get('codigo_operacion', '')} | "
                    f"Buque: {self.operacion_activa.get('nombre_buque', '')} | "
                    f"Inicio: {self.operacion_activa.get('fecha_inicio', '')}"
                )
                color = self.colors["text_dark"]
            else:
                texto = "No hay operacion activa."
                color = self.colors["danger"]

            if hasattr(self, "boletas_operacion_label") and self.boletas_operacion_label.winfo_exists():
                self.boletas_operacion_label.configure(text=texto, fg=color)
        except Exception as e:
            if hasattr(self, "boletas_operacion_label") and self.boletas_operacion_label.winfo_exists():
                self.boletas_operacion_label.configure(
                    text=f"Error consultando operacion activa: {e}",
                    fg=self.colors["danger"],
                )
            else:
                messagebox.showerror("Error operacion", str(e))

    def seleccionar_operacion_boletas(self):
        texto = self.boletas_operacion_var.get()
        op_id = self.safe_int((texto.split("|", 1)[0] if texto else ""), None)
        for op in getattr(self, "boletas_operaciones_abiertas", []) or []:
            if self.safe_int(op.get("id"), -1) == op_id:
                self.operacion_activa = op
                if hasattr(self, "boletas_operacion_label") and self.boletas_operacion_label.winfo_exists():
                    self.boletas_operacion_label.configure(
                        text=(
                            f"Operacion seleccionada: {op.get('codigo_operacion', '')} | "
                            f"Buque: {op.get('nombre_buque', '')} | "
                            f"Inicio: {op.get('fecha_inicio', '')}"
                        ),
                        fg=self.colors["text_dark"],
                    )
                return

    def asegurar_operacion_activa_boletas(self, mostrar_error=False):
        try:
            if not getattr(self, "operacion_activa", None) or not self.operacion_activa.get("id"):
                self.operacion_activa = self.api_get_operacion_activa()
            if hasattr(self, "boletas_operacion_label") and self.boletas_operacion_label.winfo_exists():
                if self.operacion_activa:
                    self.boletas_operacion_label.configure(
                        text=(
                            f"Operacion activa: {self.operacion_activa.get('codigo_operacion', '')} | "
                            f"Buque: {self.operacion_activa.get('nombre_buque', '')} | "
                            f"Inicio: {self.operacion_activa.get('fecha_inicio', '')}"
                        ),
                        fg=self.colors["text_dark"],
                    )
                else:
                    self.boletas_operacion_label.configure(text="No hay operacion activa.", fg=self.colors["danger"])
            return self.operacion_activa
        except Exception as e:
            if mostrar_error:
                messagebox.showerror("Error operacion", str(e))
            return None

    def cargar_excel_template(self):
        api_post = f"{self.api_base}/base-operaciones-camiones/cargar-excel-archivo"

        def tarea():
            operacion = self.asegurar_operacion_activa_boletas(mostrar_error=False)
            if not operacion or not operacion.get("id"):
                return {
                    "warning": "Seleccione o busque la operacion abierta antes de cargar el Excel."
                }

            estado_carga = self.api_get_boletas_estado_carga(operacion.get("id"))
            if estado_carga.get("tiene_carga_inicial"):
                return {
                    "warning": (
                        "Carga inicial detenida.\n\n"
                        f"La operacion seleccionada ya tiene {estado_carga.get('total_boletas', 0)} "
                        "boleta(s) cargada(s).\n\n"
                        "Para cargas adicionales de esta misma operacion use Aprobaciones/Operaciones, "
                        "donde los registros quedan controlados por aprobacion y auditoria."
                    )
                }

            template_path = self.obtener_template_boletas_trabajo_path(operacion.get("id"))
            if not template_path:
                return {
                    "warning": "Primero presione Abrir Template, complete el Excel y guardelo. Luego presione Cargar Excel."
                }

            estado_qr = self.api_get_qr_seguridad_estado()
            if not estado_qr.get("seguro"):
                return {
                    "warning": (
                        "El backend no tiene QR_SECRET seguro.\n\n"
                        f"Longitud actual: {estado_qr.get('longitud', 0)} caracteres.\n"
                        f"Minimo requerido: {estado_qr.get('minimo_caracteres', 32)} caracteres.\n\n"
                        "Configure QR_SECRET en Railway con un valor aleatorio de al menos 32 caracteres "
                        "y reinicie/redeploy el servicio antes de cargar Excel."
                    )
                }

            with open(template_path, "rb") as archivo:
                respuesta_post = requests.post(
                    api_post,
                    params={"operacion_id": operacion.get("id")},
                    data=archivo.read(),
                    headers={
                        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    },
                    timeout=300,
                )

            if respuesta_post.status_code != 200:
                raise RuntimeError(
                    f"No se pudo cargar el Excel a la base de datos.\n\nDetalle:\n{self.obtener_detalle_error(respuesta_post)}"
                )

            return respuesta_post.json()

        def al_terminar(resultado):
            if not isinstance(resultado, dict):
                messagebox.showerror(
                    "Cargando Excel",
                    "La carga de Excel termino sin respuesta valida del backend. "
                    "Revise la conexion y vuelva a intentar.",
                )
                return

            if resultado.get("warning"):
                messagebox.showwarning("Carga detenida", resultado["warning"])
                return

            messagebox.showinfo(
                "Excel cargado",
                f"Archivo cargado en base_operaciones_camiones.\n\n"
                f"Archivo de trabajo:\n{self.obtener_template_boletas_trabajo_path(getattr(self, 'operacion_activa', {}).get('id')) or ''}\n\n"
                f"Insertados: {resultado.get('insertados', 0)}\n"
                f"Omitidos: {resultado.get('omitidos', 0)}\n\n"
                f"Presione Cargar Tabla para consultar la base y construir la tabla visible.",
            )

        self.ejecutar_en_segundo_plano(
            "Cargando Excel",
            "Cargando datos en la base de datos. Por favor espere...",
            tarea,
            al_terminar,
        )

    def obtener_carpeta_templates_trabajo(self):
        carpeta = app_user_data_path("templates")
        os.makedirs(carpeta, exist_ok=True)
        try:
            os.chmod(carpeta, 0o777)
        except Exception:
            pass
        return carpeta

    def obtener_template_boletas_trabajo_path(self, operacion_id=None):
        path = getattr(self, "boletas_template_trabajo_path", None)
        path_operacion = getattr(self, "boletas_template_trabajo_operacion_id", None)
        if path and os.path.exists(path):
            if operacion_id is None or str(path_operacion or "") == str(operacion_id):
                return path
        return None

    def abrir_template_boletas_remoto(self):
        try:
            operacion = self.asegurar_operacion_activa_boletas(mostrar_error=False) or {}
            operacion_id = self.safe_int(operacion.get("id"), None)
            usuario_actual = getattr(self, "current_user", {}) or {}
            usuario_raw = (
                usuario_actual.get("username")
                or usuario_actual.get("usuario")
                or os.getenv("USERNAME")
                or "usuario"
            )
            usuario = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in str(usuario_raw))[:40] or "usuario"
            destino = os.path.join(
                self.obtener_carpeta_templates_trabajo(),
                f"base_operaciones_camiones_op_{operacion_id or 'sin_operacion'}_{usuario}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            )

            descargado = False
            try:
                respuesta = requests.get(
                    f"{self.api_base}/base-operaciones-camiones/template-excel/descargar",
                    timeout=60,
                )
                if respuesta.status_code == 200 and respuesta.content:
                    with open(destino, "wb") as archivo:
                        archivo.write(respuesta.content)
                    descargado = True
            except Exception:
                descargado = False

            if not descargado:
                if os.path.exists(self.boletas_template_local_path):
                    shutil.copy2(self.boletas_template_local_path, destino)
                else:
                    messagebox.showerror(
                        "Template no disponible",
                        "No se pudo obtener el template desde backend ni desde el respaldo local.",
                    )
                    return

            try:
                os.chmod(destino, stat.S_IREAD | stat.S_IWRITE)
            except Exception:
                try:
                    os.chmod(destino, 0o666)
                except Exception:
                    pass

            self.boletas_template_trabajo_path = destino
            self.boletas_template_trabajo_operacion_id = operacion_id

            try:
                os.startfile(destino)
            except Exception:
                messagebox.showinfo(
                    "Template disponible",
                    f"Template editable preparado. Abra manualmente:\n{destino}",
                )

        except requests.exceptions.ConnectionError:
            messagebox.showerror("Backend no disponible", f"No se pudo conectar con la API.\n\nVerifica:\n{self.api_base}")
        except requests.exceptions.Timeout:
            messagebox.showerror("Tiempo agotado", "La API tardo demasiado en responder.")
        except Exception as e:
            messagebox.showerror("Error inesperado", str(e))

    def abrir_template_aprobaciones_remoto(self):
        try:
            operacion = getattr(self, "operacion_activa", None) or self.asegurar_operacion_activa_boletas(mostrar_error=False) or {}
            operacion_id = self.safe_int(operacion.get("id"), None)
            if not operacion_id:
                messagebox.showwarning("Sin operacion", "Busque o seleccione una operacion abierta antes de abrir el template.")
                return

            usuario_actual = getattr(self, "current_user", {}) or {}
            usuario_raw = (
                usuario_actual.get("username")
                or usuario_actual.get("usuario")
                or os.getenv("USERNAME")
                or "usuario"
            )
            usuario = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in str(usuario_raw))[:40] or "usuario"
            destino = os.path.join(
                self.obtener_carpeta_templates_trabajo(),
                f"aprobaciones_op_{operacion_id}_{usuario}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            )

            respuesta = requests.get(
                f"{self.api_base}/aprobaciones/template-trabajo/descargar",
                params={"operacion_id": operacion_id},
                timeout=90,
            )
            if respuesta.status_code != 200 or not respuesta.content:
                raise RuntimeError(self.obtener_detalle_error(respuesta))

            with open(destino, "wb") as archivo:
                archivo.write(respuesta.content)

            try:
                os.chmod(destino, stat.S_IREAD | stat.S_IWRITE)
            except Exception:
                pass

            self.boletas_template_trabajo_path = destino
            self.boletas_template_trabajo_operacion_id = operacion_id
            self.iniciar_sync_template_aprobaciones_backend(destino, operacion_id)

            try:
                os.startfile(destino)
            except Exception:
                messagebox.showinfo(
                    "Template disponible",
                    f"Template editable preparado. Abra manualmente:\n{destino}",
                )

        except requests.exceptions.ConnectionError:
            messagebox.showerror("Backend no disponible", f"No se pudo conectar con la API.\n\nVerifica:\n{self.api_base}")
        except requests.exceptions.Timeout:
            messagebox.showerror("Tiempo agotado", "La API tardo demasiado en responder.")
        except Exception as e:
            messagebox.showerror("Error abriendo template", str(e))

    def iniciar_sync_template_aprobaciones_backend(self, path, operacion_id):
        if not path or not operacion_id:
            return

        token = f"{operacion_id}:{path}:{datetime.now().timestamp()}"
        self.aprobaciones_template_sync_token = token
        usuario = (getattr(self, "current_user", {}) or {}).get("username") or os.getenv("USERNAME") or "desktop"

        def subir():
            with open(path, "rb") as archivo:
                contenido = archivo.read()
            respuesta = requests.post(
                f"{self.api_base}/aprobaciones/template-trabajo/guardar",
                params={"operacion_id": operacion_id, "usuario": usuario},
                data=contenido,
                headers={"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
                timeout=90,
            )
            if respuesta.status_code != 200:
                raise RuntimeError(self.obtener_detalle_error(respuesta))

        def vigilar():
            ultimo_mtime = None
            ultimo_tamano = None
            for _ in range(240):
                if getattr(self, "aprobaciones_template_sync_token", None) != token:
                    return
                try:
                    if os.path.exists(path):
                        stat_info = os.stat(path)
                        firma = (stat_info.st_mtime, stat_info.st_size)
                        if firma != (ultimo_mtime, ultimo_tamano):
                            ultimo_mtime, ultimo_tamano = firma
                            subir()
                except Exception:
                    pass
                threading.Event().wait(3)

        threading.Thread(target=vigilar, daemon=True).start()

    def obtener_params_boletas(self):
        params = {}
        for key, var in getattr(self, "boletas_filter_vars", {}).items():
            value = var.get().strip()
            if value:
                params[key] = value
        if getattr(self, "operacion_activa", None) and self.operacion_activa.get("id"):
            params.setdefault("operacion_id", self.operacion_activa.get("id"))
        return params

    def opciones_tienen_datos(self, opciones):
        return any(opciones.get(key) for key in (
            "empresas",
            "guias",
            "productos",
            "choferes",
            "placas",
        ))

    def obtener_filas_boletas_para_filtros(self, params=None):
        params = dict(params or {})
        if not params.get("operacion_id"):
            operacion = getattr(self, "operacion_activa", None)
            if operacion and operacion.get("id"):
                params["operacion_id"] = operacion.get("id")
        if not params.get("operacion_id"):
            raise RuntimeError("Seleccione una operacion antes de consultar boletas.")
        try:
            filas = self.api_get_boletas(params)
            if filas:
                return filas
        except Exception:
            return list(getattr(self, "boletas_cache", []) or [])

    def cargar_filtros_boletas(self):
        self.asegurar_operacion_activa_boletas()

        def tarea():
            try:
                data = self.api_get_boletas_filtros(self.obtener_params_boletas())
                opciones = data.get("opciones", {}) if isinstance(data, dict) else {}
                if self.opciones_tienen_datos(opciones):
                    return data
            except Exception:
                pass
            filas = self.obtener_filas_boletas_para_filtros(self.obtener_params_boletas())
            return {"ok": True, "opciones": self.opciones_filtros_desde_filas(filas)}

        def al_terminar(data):
            self.aplicar_opciones_filtros_boletas(data.get("opciones", {}))

        self.ejecutar_en_segundo_plano("Filtros boletas", "Cargando filtros dinamicos...", tarea, al_terminar)

    def opciones_filtros_desde_filas(self, filas):
        mapa = {
            "empresas": "empresa",
            "guias": "guia",
            "productos": "producto",
            "choferes": "chofer",
            "placas": "placa",
        }
        opciones = {}
        for option_key, field in mapa.items():
            valores = {
                str(fila.get(field) or "").strip()
                for fila in filas or []
                if str(fila.get(field) or "").strip()
            }
            opciones[option_key] = sorted(valores)
        return opciones

    def aplicar_opciones_filtros_boletas(self, opciones):
        mapa = {
            "empresa": "empresas",
            "guia": "guias",
            "producto": "productos",
            "chofer": "choferes",
            "placa": "placas",
        }
        for key, option_key in mapa.items():
            widget = getattr(self, "boletas_filter_widgets", {}).get(key)
            var = getattr(self, "boletas_filter_vars", {}).get(key)
            if widget is None or var is None:
                continue
            actual = var.get()
            valores = [""] + [str(v) for v in opciones.get(option_key, []) if v not in (None, "")]
            widget["values"] = valores
            if actual and actual not in valores:
                var.set("")

    def limpiar_filtros_boletas(self):
        for var in getattr(self, "boletas_filter_vars", {}).values():
            var.set("")

    def refrescar_boletas(self):
        self.asegurar_operacion_activa_boletas()

        def tarea():
            return self.api_get_boletas(self.obtener_params_boletas())

        def al_terminar(datos):
            self.boletas_cache = datos
            self.cargar_datos_en_tabla(datos)

        self.ejecutar_en_segundo_plano(
            "Cargar Tabla",
            "Consultando boletas en la base de datos...",
            tarea,
            al_terminar,
        )

    def cargar_datos_en_tabla(self, datos):
        if self.tree is None:
            return
        for item in self.tree.get_children():
            self.tree.delete(item)
        for fila in datos:
            self.tree.insert(
                "",
                "end",
                values=(
                    fila.get("id", ""),
                    fila.get("guia", ""),
                    fila.get("numero_embarque", ""),
                    fila.get("bodega_numero", ""),
                    fila.get("empresa", ""),
                    fila.get("buque", ""),
                    self.formatear_fecha(fila.get("fecha")),
                    fila.get("producto", ""),
                    fila.get("chofer", ""),
                    fila.get("placa", ""),
                    fila.get("estado", "PENDIENTE"),
                    fila.get("lecturas", 0),
                    fila.get("etapa_qr", "") or self.estado_visual_qr(fila),
                    fila.get("numero_tolva", ""),
                    "SI" if fila.get("qr_bloqueado") else "NO",
                ),
            )

    def generar_qr_backend(self):
        api_url = f"{self.api_base}/base-operaciones-camiones/generar-qr"
        operacion = self.asegurar_operacion_activa_boletas(mostrar_error=False)
        if not operacion or not operacion.get("id"):
            messagebox.showwarning("Operacion requerida", "Busque una operacion abierta antes de generar QR.")
            return
        payload = {"formato": "jpg", "operacion_id": operacion.get("id")}

        def tarea():
            respuesta = requests.post(api_url, json=payload, timeout=120)
            if respuesta.status_code != 200:
                raise RuntimeError(f"No se pudieron generar los QR.\n\nDetalle:\n{self.obtener_detalle_error(respuesta)}")

            data = respuesta.json()
            archivos = data.get("archivos", [])
            ruta_base_local = app_user_data_path("QR")
            os.makedirs(ruta_base_local, exist_ok=True)
            generados_local = 0

            for item in archivos:
                empresa = str(item.get("empresa") or "SIN_EMPRESA").strip().replace(" ", "_")
                chofer = str(item.get("chofer") or "SIN_CHOFER").strip().replace(" ", "_")
                guia = str(item.get("guia") or "SIN_GUIA").strip().replace(" ", "_")
                url_qr = item.get("url_qr")
                if not url_qr:
                    continue
                carpeta_destino = os.path.join(ruta_base_local, empresa, chofer)
                os.makedirs(carpeta_destino, exist_ok=True)
                archivo_qr = os.path.join(carpeta_destino, f"{guia}.jpg")
                qr = qrcode.QRCode(
                    version=None,
                    error_correction=qrcode.constants.ERROR_CORRECT_M,
                    box_size=12,
                    border=4,
                )
                qr.add_data(url_qr)
                qr.make(fit=True)
                img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
                img = img.resize((900, 900))
                img.save(archivo_qr, "JPEG", quality=95)
                generados_local += 1

            return {
                "data": data,
                "generados_local": generados_local,
                "ruta_base_local": ruta_base_local,
            }

        def al_terminar(resultado):
            data = resultado["data"]
            messagebox.showinfo(
                "QR generados",
                f"QR generados correctamente.\n\n"
                f"Total backend: {data.get('total_generados', 0)}\n"
                f"Total local: {resultado['generados_local']}\n\n"
                f"Ruta local:\n{resultado['ruta_base_local']}",
            )

        self.ejecutar_en_segundo_plano(
            "Generando QR",
            "Generando codigos QR. Esta operacion puede tardar si hay muchas guias. Por favor espere...",
            tarea,
            al_terminar,
        )

    def entregar_qr_backend(self):
        ventana = tk.Toplevel(self)
        ventana.title("Entrega QR")
        ventana.geometry("780x720")
        ventana.minsize(720, 640)
        ventana.configure(bg=self.colors["bg_card"])
        ventana.transient(self)
        self.aplicar_icono_app(ventana)

        self.asegurar_operacion_activa_boletas()
        seleccion_ids = []
        if self.tree is not None:
            for item in self.tree.selection():
                valores = self.tree.item(item, "values")
                registro_id = self.safe_int(valores[0], None) if valores else None
                if registro_id:
                    seleccion_ids.append(registro_id)

        canal_var = tk.StringVar(value="CARPETA")
        formato_var = tk.StringVar(value="jpg")
        email_var = tk.StringVar()
        whatsapp_var = tk.StringVar()
        carpeta_var = tk.StringVar()
        usar_seleccion_var = tk.BooleanVar(value=bool(seleccion_ids))
        guia_var = tk.StringVar(value=getattr(self, "boletas_filter_vars", {}).get("guia", tk.StringVar()).get() if hasattr(self, "boletas_filter_vars") else "")
        guias_lista_var = tk.StringVar()
        empresa_var = tk.StringVar(value=getattr(self, "boletas_filter_vars", {}).get("empresa", tk.StringVar()).get() if hasattr(self, "boletas_filter_vars") else "")
        producto_var = tk.StringVar(value=getattr(self, "boletas_filter_vars", {}).get("producto", tk.StringVar()).get() if hasattr(self, "boletas_filter_vars") else "")
        chofer_var = tk.StringVar(value=getattr(self, "boletas_filter_vars", {}).get("chofer", tk.StringVar()).get() if hasattr(self, "boletas_filter_vars") else "")
        placa_var = tk.StringVar(value=getattr(self, "boletas_filter_vars", {}).get("placa", tk.StringVar()).get() if hasattr(self, "boletas_filter_vars") else "")
        estado_var = tk.StringVar(value=getattr(self, "boletas_filter_vars", {}).get("estado", tk.StringVar()).get() if hasattr(self, "boletas_filter_vars") else "")
        etapa_var = tk.StringVar(value=getattr(self, "boletas_filter_vars", {}).get("etapa_qr", tk.StringVar()).get() if hasattr(self, "boletas_filter_vars") else "")

        canvas = tk.Canvas(ventana, bg=self.colors["bg_card"], highlightthickness=0)
        scroll_y = ttk.Scrollbar(ventana, orient="vertical", command=canvas.yview)
        body = tk.Frame(canvas, bg=self.colors["bg_card"])
        window_id = canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=scroll_y.set)
        self.bind_scroll_canvas(canvas, body, window_id)
        canvas.pack(side="left", fill="both", expand=True)
        scroll_y.pack(side="right", fill="y")

        tk.Label(
            body,
            text="Preparar entrega de QR",
            font=("Segoe UI", 15, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=18, pady=(16, 8))

        op_text = "Operacion activa: "
        if getattr(self, "operacion_activa", None):
            op_text += f"{self.operacion_activa.get('codigo_operacion', '')} | {self.operacion_activa.get('nombre_buque', '')}"
        else:
            op_text += "sin operacion activa detectada"
        tk.Label(
            body,
            text=op_text,
            font=("Segoe UI", 9, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["muted"],
        ).pack(anchor="w", padx=18, pady=(0, 8))

        form = tk.Frame(body, bg=self.colors["bg_card"])
        form.pack(fill="x", padx=18)

        canal_host = tk.Frame(form, bg=self.colors["bg_card"])
        canal_host.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
        tk.Label(canal_host, text="Canal", font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
        ttk.Combobox(canal_host, textvariable=canal_var, values=["CARPETA", "CORREO", "WHATSAPP"], state="readonly").pack(fill="x")

        formato_host = tk.Frame(form, bg=self.colors["bg_card"])
        formato_host.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        tk.Label(formato_host, text="Formato", font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
        ttk.Combobox(formato_host, textvariable=formato_var, values=["jpg", "pdf"], state="readonly").pack(fill="x")

        canal_campos_host = tk.Frame(form, bg=self.colors["bg_card"])
        canal_campos_host.grid(row=1, column=0, columnspan=2, sticky="ew")
        canal_frames = {}

        correo_frame = tk.Frame(canal_campos_host, bg=self.colors["bg_card"])
        tk.Label(correo_frame, text="Email destino", font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w", padx=5)
        ttk.Entry(correo_frame, textvariable=email_var).pack(fill="x", padx=5, pady=(0, 5))
        canal_frames["CORREO"] = correo_frame

        whatsapp_frame = tk.Frame(canal_campos_host, bg=self.colors["bg_card"])
        tk.Label(whatsapp_frame, text="WhatsApp destino", font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w", padx=5)
        ttk.Entry(whatsapp_frame, textvariable=whatsapp_var).pack(fill="x", padx=5, pady=(0, 5))
        canal_frames["WHATSAPP"] = whatsapp_frame

        carpeta_frame = tk.Frame(canal_campos_host, bg=self.colors["bg_card"])
        tk.Label(carpeta_frame, text="Carpeta local opcional", font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w", padx=5)
        carpeta_row = tk.Frame(carpeta_frame, bg=self.colors["bg_card"])
        carpeta_row.pack(fill="x", padx=5, pady=(0, 5))
        ttk.Entry(carpeta_row, textvariable=carpeta_var).pack(side="left", fill="x", expand=True, padx=(0, 5))
        ttk.Button(
            carpeta_row,
            text="Elegir",
            style="Gray.TButton",
            command=lambda: carpeta_var.set(filedialog.askdirectory(title="Guardar QR en carpeta") or carpeta_var.get()),
        ).pack(side="left")
        canal_frames["CARPETA"] = carpeta_frame

        def actualizar_campos_canal(*_args):
            for frame in canal_frames.values():
                frame.pack_forget()
            canal_frames.get(canal_var.get(), carpeta_frame).pack(fill="x")

        canal_var.trace_add("write", actualizar_campos_canal)
        actualizar_campos_canal()

        campos = [
            ("Guia contiene", guia_var, "combo_guia"),
            ("Guias exactas separadas por coma", guias_lista_var, "entry"),
            ("Cliente / empresa", empresa_var, "combo_empresa"),
            ("Producto", producto_var, "combo_producto"),
            ("Chofer", chofer_var, "combo_chofer"),
            ("Placa", placa_var, "combo_placa"),
            ("Estado", estado_var, "combo_estado"),
            ("Etapa QR", etapa_var, "combo_etapa_qr"),
        ]
        entrega_filter_widgets = {}
        for idx, (label, var, kind) in enumerate(campos):
            box = tk.Frame(form, bg=self.colors["bg_card"])
            box.grid(row=(idx // 2) + 2, column=idx % 2, sticky="ew", padx=5, pady=5)
            tk.Label(box, text=label, font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
            if kind.startswith("combo_"):
                combo = ttk.Combobox(box, textvariable=var, values=[""], state="normal")
                combo.pack(fill="x")
                entrega_filter_widgets[kind.replace("combo_", "")] = combo
            else:
                ttk.Entry(box, textvariable=var).pack(fill="x")
        for col in range(2):
            form.grid_columnconfigure(col, weight=1)

        seleccion_panel = tk.Frame(body, bg=self.colors["bg_card"])
        seleccion_panel.pack(fill="x", padx=18, pady=(8, 0))
        ttk.Checkbutton(
            seleccion_panel,
            text=f"Usar seleccion actual de tabla ({len(seleccion_ids)} guia(s))",
            variable=usar_seleccion_var,
        ).pack(anchor="w")
        tk.Label(
            seleccion_panel,
            text=(
                "Si marca esta opcion, se entregan solo las filas seleccionadas. "
                "Si escribe guias exactas separadas por coma, espacio o salto de linea, se entregan esas guias. "
                "Si no marca ni escribe guias, se usan los filtros visibles; sin filtros se prepara toda la operacion activa."
            ),
            font=("Segoe UI", 8),
            bg=self.colors["bg_card"],
            fg=self.colors["muted"],
            wraplength=690,
            justify="left",
        ).pack(anchor="w", pady=(3, 0))

        def params_entrega_filtros():
            params = {}
            if getattr(self, "operacion_activa", None) and self.operacion_activa.get("id"):
                params["operacion_id"] = self.operacion_activa.get("id")
            for key, var in {
                "guia": guia_var,
                "empresa": empresa_var,
                "producto": producto_var,
                "chofer": chofer_var,
                "placa": placa_var,
                "estado": estado_var,
                "etapa_qr": etapa_var,
            }.items():
                value = var.get().strip()
                if value:
                    params[key] = value
            return params

        def aplicar_opciones_entrega(opciones):
            mapa = {
                "guia": "guias",
                "empresa": "empresas",
                "producto": "productos",
                "chofer": "choferes",
                "placa": "placas",
                "estado": "estados",
                "etapa_qr": "etapas_qr",
            }
            for key, option_key in mapa.items():
                widget = entrega_filter_widgets.get(key)
                if widget:
                    widget["values"] = [""] + [str(v) for v in opciones.get(option_key, []) if v not in (None, "")]

        def cargar_opciones_entrega():
            try:
                try:
                    data = self.api_get_boletas_filtros(params_entrega_filtros())
                    opciones = data.get("opciones", {})
                    if not self.opciones_tienen_datos(opciones):
                        raise RuntimeError("Filtros vacios")
                except Exception:
                    filas = self.obtener_filas_boletas_para_filtros(params_entrega_filtros())
                    opciones = self.opciones_filtros_desde_filas(filas)
                aplicar_opciones_entrega(opciones)
            except Exception as e:
                messagebox.showerror("Filtros entrega QR", str(e))

        filas_cache = list(getattr(self, "boletas_cache", []) or [])
        if filas_cache:
            buque_activo = str((getattr(self, "operacion_activa", None) or {}).get("nombre_buque") or "").strip().lower()
            if buque_activo:
                filas_cache = [
                    fila for fila in filas_cache
                    if str(fila.get("buque") or "").strip().lower() == buque_activo
                ] or filas_cache
            aplicar_opciones_entrega(self.opciones_filtros_desde_filas(filas_cache))

        filtros_actions = tk.Frame(body, bg=self.colors["bg_card"])
        filtros_actions.pack(fill="x", padx=18, pady=(10, 0))
        ttk.Button(filtros_actions, text="Cargar filtros de entrega", style="Gray.TButton", command=cargar_opciones_entrega).pack(side="left", padx=(0, 8))
        ttk.Button(
            filtros_actions,
            text="Usar seleccion tabla",
            style="Gray.TButton",
            command=lambda: usar_seleccion_var.set(True),
        ).pack(side="left", padx=(0, 8))
        ttk.Button(
            filtros_actions,
            text="Limpiar filtros entrega",
            style="Gray.TButton",
            command=lambda: [var.set("") for var in (guia_var, guias_lista_var, empresa_var, producto_var, chofer_var, placa_var, estado_var, etapa_var)],
        ).pack(side="left")

        info = tk.Label(
            body,
            text="Si no hay proveedor SMTP/WhatsApp API configurado, XTRAVON prepara archivos y links para envio manual controlado.",
            font=("Segoe UI", 9),
            bg=self.colors["bg_card"],
            fg=self.colors["muted"],
            wraplength=690,
            justify="left",
        )
        info.pack(anchor="w", padx=18, pady=(8, 0))

        def ejecutar_entrega():
            api_url = f"{self.api_base}/base-operaciones-camiones/entregar-qr"
            guias_lista = [
                item.strip()
                for item in guias_lista_var.get().replace("\n", ",").replace(" ", ",").split(",")
                if item.strip()
            ]
            filtros_payload = params_entrega_filtros()
            payload = {
                "canal": canal_var.get(),
                "formato": formato_var.get(),
                "email_destino": email_var.get().strip() or None,
                "whatsapp_destino": whatsapp_var.get().strip() or None,
                "operacion_id": filtros_payload.get("operacion_id"),
                "ids": seleccion_ids if usar_seleccion_var.get() and seleccion_ids else None,
                "guia": filtros_payload.get("guia"),
                "guias": guias_lista or None,
                "empresa": filtros_payload.get("empresa"),
                "producto": filtros_payload.get("producto"),
                "chofer": filtros_payload.get("chofer"),
                "placa": filtros_payload.get("placa"),
                "estado": filtros_payload.get("estado"),
                "etapa_qr": filtros_payload.get("etapa_qr"),
                "ruta_destino_local": carpeta_var.get().strip() or None,
            }
            if not payload.get("ids") and not payload.get("guias") and not any(payload.get(k) for k in ("guia", "empresa", "producto", "chofer", "placa", "estado", "etapa_qr")):
                if not messagebox.askyesno(
                    "Confirmar entrega",
                    "No selecciono guias ni filtros especificos. Se prepararan los QR disponibles de la operacion activa. Desea continuar?",
                ):
                    return

            ventana.destroy()

            def tarea():
                respuesta = requests.post(api_url, json=payload, timeout=120)
                if respuesta.status_code != 200:
                    raise RuntimeError(self.obtener_detalle_error(respuesta))
                return respuesta.json()

            def al_terminar(data):
                generados_local = 0
                ruta_local = carpeta_var.get().strip()
                if canal_var.get() == "CARPETA" and ruta_local:
                    try:
                        os.makedirs(ruta_local, exist_ok=True)
                        for entrega in data.get("entregas", []):
                            for item in entrega.get("archivos", []):
                                url_qr = item.get("url_qr")
                                guia = str(item.get("guia") or item.get("id") or "SIN_GUIA").strip().replace(" ", "_")
                                if not url_qr:
                                    continue
                                qr = qrcode.QRCode(
                                    version=None,
                                    error_correction=qrcode.constants.ERROR_CORRECT_M,
                                    box_size=12,
                                    border=4,
                                )
                                qr.add_data(url_qr)
                                qr.make(fit=True)
                                img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
                                img.resize((900, 900)).save(os.path.join(ruta_local, f"{guia}.jpg"), "JPEG", quality=95)
                                generados_local += 1
                    except Exception as exc:
                        messagebox.showwarning("Copia local QR", f"No se pudieron guardar todos los QR localmente:\n{exc}")
                links = [entrega.get("link") for entrega in data.get("entregas", []) if entrega.get("link")]
                enviadas = len([entrega for entrega in data.get("entregas", []) if entrega.get("estado") == "ENVIADO"])
                errores = len([entrega for entrega in data.get("entregas", []) if entrega.get("estado") == "ERROR_ENVIO"])
                resumen = (
                    f"Lote: {data.get('lote_codigo')}\n"
                    f"Canal: {data.get('canal')}\n"
                    f"QR preparados: {data.get('total_qr', 0)}\n"
                    f"Entregas: {data.get('total_entregas', 0)}\n"
                    f"Pendientes contacto: {data.get('pendientes_contacto', 0)}\n\n"
                    f"Enviadas automaticamente: {enviadas}\n"
                    f"Errores de envio: {errores}\n\n"
                    f"Ruta QR backend:\n{data.get('ruta_base_qr', '')}"
                )
                if ruta_local:
                    resumen += f"\n\nRuta local:\n{ruta_local}\nQR locales generados: {generados_local}"
                if links and messagebox.askyesno("Entrega QR", f"{resumen}\n\nDesea abrir el primer link de envio?"):
                    webbrowser.open(links[0])
                else:
                    messagebox.showinfo("Entrega QR", resumen)

            self.ejecutar_en_segundo_plano(
                "Entrega QR",
                "Preparando entrega, lote, archivos QR y trazabilidad. Por favor espere...",
                tarea,
                al_terminar,
            )

        actions = tk.Frame(body, bg=self.colors["bg_card"])
        actions.pack(fill="x", padx=18, pady=16)
        ttk.Button(actions, text="Preparar entrega", style="Olive.TButton", command=ejecutar_entrega).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Cancelar", style="Gray.TButton", command=ventana.destroy).pack(side="left")
        # Los filtros se cargan solo cuando el usuario presiona el boton; abrir esta ventana no consulta el backend.

    def obtener_fila_seleccionada(self):
        if self.tree is None:
            messagebox.showwarning("Sin tabla", "La tabla de boletas no esta disponible.")
            return None
        seleccion = self.tree.selection()
        if not seleccion:
            messagebox.showwarning("Sin seleccion", "Selecciona una guia en la tabla.")
            return None
        valores = self.tree.item(seleccion[0], "values")
        if not valores:
            messagebox.showwarning("Seleccion invalida", "No se pudo leer la fila seleccionada.")
            return None
        registro_id = self.safe_int(valores[0], None)
        if registro_id is None:
            messagebox.showwarning("Seleccion invalida", "El ID de la guia no es valido.")
            return None

        for fila in self.boletas_cache:
            if self.safe_int(fila.get("id"), -1) == registro_id:
                return fila

        try:
            respuesta = requests.get(f"{self.api_base}/base-operaciones-camiones/{registro_id}", timeout=60)
            if respuesta.status_code == 200:
                return respuesta.json()
            messagebox.showerror("Error", self.obtener_detalle_error(respuesta))
        except Exception as e:
            messagebox.showerror("Error", str(e))
        return None

    def ver_guia_seleccionada(self):
        fila = self.obtener_fila_seleccionada()
        if not fila:
            return

        ventana = tk.Toplevel(self)
        ventana.title(f"Detalle guia {fila.get('guia', '')}")
        ventana.geometry("760x640")
        ventana.minsize(640, 520)
        ventana.resizable(True, True)
        ventana.configure(bg=self.colors["bg_main"])

        header = tk.Frame(ventana, bg=self.colors["bg_topbar"], height=70)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(
            header,
            text=f"Detalle de Guia {fila.get('guia', '')}",
            font=("Segoe UI", 18, "bold"),
            bg=self.colors["bg_topbar"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=20, pady=(12, 0))
        tk.Label(
            header,
            text=f"Estado: {fila.get('estado', '')} | Lecturas: {fila.get('lecturas', 0)} | Etapa QR: {fila.get('etapa_qr', '')}",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors["bg_topbar"],
            fg=self.colors["text_secondary"],
        ).pack(anchor="w", padx=20)

        body = tk.Frame(
            ventana,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        body.pack(fill="both", expand=True, padx=20, pady=20)

        campos = [
            ("ID", fila.get("id")),
            ("Guia", fila.get("guia")),
            ("Numero de embarque", fila.get("numero_embarque")),
            ("Bodega", fila.get("bodega_numero")),
            ("Empresa", fila.get("empresa")),
            ("Buque", fila.get("buque")),
            ("Fecha", self.formatear_fecha(fila.get("fecha"))),
            ("Producto", fila.get("producto")),
            ("Chofer", fila.get("chofer")),
            ("Placa", fila.get("placa")),
            ("Estado", fila.get("estado")),
            ("Lecturas / Escaneos", fila.get("lecturas")),
            ("Aprobada", fila.get("aprobada")),
            ("Ficha", fila.get("ficha")),
            ("Peso vacio", fila.get("peso_vacio")),
            ("Peso lleno", fila.get("peso_lleno")),
            ("Marchamos", fila.get("marchamos")),
            ("Etapa QR", fila.get("etapa_qr")),
            ("QR bloqueado", fila.get("qr_bloqueado")),
            ("Lectura ingreso", self.formatear_fecha(fila.get("lectura_ingreso"))),
            ("Lectura salida", self.formatear_fecha(fila.get("lectura_salida"))),
            ("Lectura tercer escaneo", self.formatear_fecha(fila.get("lectura_tercer_escaneo"))),
            ("Numero tolva", fila.get("numero_tolva")),
            ("Comentario SOF", fila.get("comentario_issue_log")),
            ("Fecha escaneo", self.formatear_fecha(fila.get("fecha_escaneo"))),
            ("Fecha cierre", self.formatear_fecha(fila.get("fecha_cierre"))),
            ("Creado en", self.formatear_fecha(fila.get("creado_en"))),
        ]

        canvas = tk.Canvas(body, bg=self.colors["bg_card"], highlightthickness=0)
        scroll = ttk.Scrollbar(body, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=self.colors["bg_card"])
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        self.bind_scroll_canvas(canvas, inner)
        canvas.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        for label, value in campos:
            row = tk.Frame(inner, bg=self.colors["bg_card"])
            row.pack(fill="x", padx=18, pady=5)
            tk.Label(
                row,
                text=f"{label}:",
                width=18,
                anchor="w",
                font=("Segoe UI", 10, "bold"),
                bg=self.colors["bg_card"],
                fg=self.colors["text_dark"],
            ).pack(side="left")
            tk.Label(
                row,
                text="" if value is None else str(value),
                anchor="w",
                justify="left",
                wraplength=500,
                font=("Segoe UI", 10),
                bg=self.colors["bg_card"],
                fg="#4C4C4C",
            ).pack(side="left", fill="x", expand=True)

    def exportar_excel(self):
        try:
            datos = self.boletas_cache or self.api_get_boletas(self.obtener_params_boletas())
        except Exception as e:
            messagebox.showerror("Error", str(e))
            return
        if not datos:
            messagebox.showwarning("Sin datos", "No hay datos para exportar.")
            return

        ruta = filedialog.asksaveasfilename(
            title="Exportar tabla a Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not ruta:
            return

        columnas = [
            "id",
            "guia",
            "numero_embarque",
            "bodega_numero",
            "empresa",
            "buque",
            "fecha",
            "producto",
            "chofer",
            "placa",
            "estado",
            "lecturas",
            "aprobada",
            "lectura_ingreso",
            "lectura_salida",
            "ficha",
            "peso_vacio",
            "peso_lleno",
            "marchamos",
            "etapa_qr",
            "lectura_tercer_escaneo",
            "numero_tolva",
            "comentario_issue_log",
            "qr_bloqueado",
            "fecha_escaneo",
            "fecha_cierre",
            "creado_en",
        ]

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "base_operaciones"
            header_fill = PatternFill("solid", fgColor="6B705C")
            header_font = Font(color="FFFFFF", bold=True)

            for col_idx, col in enumerate(columnas, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row_idx, fila in enumerate(datos, start=2):
                for col_idx, col in enumerate(columnas, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=fila.get(col))

            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

            wb.save(ruta)
            messagebox.showinfo("Exportacion completada", f"Archivo exportado correctamente:\n{ruta}")
        except Exception as e:
            messagebox.showerror("Error al exportar", str(e))

    def ver_cuotas_operacion_popup(self):
        operacion_id = self.obtener_operacion_seleccionada_id()
        if not operacion_id:
            messagebox.showwarning("Sin seleccion", "Seleccione una operacion para ver cuotas.")
            return

        try:
            data = self.api_get_reporte_buque(operacion_id)
        except Exception as e:
            messagebox.showerror("Ver cuotas", str(e))
            return

        operacion = data.get("operacion", {})
        kpis = data.get("kpis", {})
        graficos = data.get("graficos", {})
        clientes = data.get("clientes", [])
        bodegas = data.get("bodegas", [])
        plan_viajes = data.get("plan_viajes", {}) or {}

        ventana = tk.Toplevel(self)
        ventana.title(f"Cuotas y avance - {operacion.get('nombre_buque', '')}")
        ventana.geometry("1180x760")
        ventana.minsize(980, 620)
        ventana.resizable(True, True)
        ventana.configure(bg=self.colors["bg_main"])

        header = tk.Frame(ventana, bg=self.colors["bg_topbar"], height=76)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(
            header,
            text=f"Cuotas, descargado y productividad - {operacion.get('nombre_buque', '')}",
            font=("Segoe UI", 17, "bold"),
            bg=self.colors["bg_topbar"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=18, pady=(12, 0))
        tk.Label(
            header,
            text=f"Operacion {operacion.get('codigo_operacion', operacion_id)} | Estado {operacion.get('estado', '')}",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors["bg_topbar"],
            fg=self.colors["text_secondary"],
        ).pack(anchor="w", padx=18)

        filtro_cliente_var = tk.StringVar(value="")
        filtro_producto_var = tk.StringVar(value="")
        filtro_cuota_var = tk.StringVar(value="")
        exportar_formato_var = tk.StringVar(value="PDF")
        clientes_opciones = sorted({str(row.get("empresa") or "") for row in clientes if row.get("empresa")})
        productos_opciones = sorted({str(row.get("producto") or "") for row in clientes if row.get("producto")})
        cuotas_opciones = sorted({
            self.formatear_numero(row.get("cuota_mt"), 2)
            for row in clientes
            if self.safe_number(row.get("cuota_mt"), 0) > 0
        })

        filtros_panel = tk.Frame(ventana, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1)
        filtros_panel.pack(fill="x", padx=14, pady=(12, 0))

        def filtro_box(label, variable, values, width=18):
            box = tk.Frame(filtros_panel, bg=self.colors["bg_card"])
            box.pack(side="left", fill="x", expand=True, padx=8, pady=10)
            tk.Label(box, text=label, font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
            combo = ttk.Combobox(box, textvariable=variable, values=[""] + values, state="readonly", width=width)
            combo.pack(fill="x")
            return combo

        filtro_box("Cliente", filtro_cliente_var, clientes_opciones)
        filtro_box("Producto", filtro_producto_var, productos_opciones)
        filtro_box("Cuota MT", filtro_cuota_var, cuotas_opciones, width=12)

        export_box = tk.Frame(filtros_panel, bg=self.colors["bg_card"])
        export_box.pack(side="left", padx=8, pady=10)
        tk.Label(export_box, text="Exportar", font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
        ttk.Combobox(export_box, textvariable=exportar_formato_var, values=["PDF", "Excel"], state="readonly", width=10).pack(fill="x")

        action_box = tk.Frame(filtros_panel, bg=self.colors["bg_card"])
        action_box.pack(side="left", padx=8, pady=10)
        ttk.Button(action_box, text="Aplicar filtros", style="Olive.TButton", command=lambda: aplicar_filtros_popup()).pack(side="left", padx=(0, 8), pady=(16, 0))
        ttk.Button(action_box, text="Limpiar", style="Gray.TButton", command=lambda: limpiar_filtros_popup()).pack(side="left", padx=(0, 8), pady=(16, 0))
        ttk.Button(action_box, text="Descargar", style="Gray.TButton", command=lambda: exportar_popup()).pack(side="left", pady=(16, 0))

        host = tk.Frame(ventana, bg=self.colors["bg_main"])
        host.pack(fill="both", expand=True, padx=14, pady=14)
        canvas = tk.Canvas(host, bg=self.colors["bg_main"], highlightthickness=0)
        scroll_y = ttk.Scrollbar(host, orient="vertical", command=canvas.yview)
        scroll_x = ttk.Scrollbar(host, orient="horizontal", command=canvas.xview)
        body = tk.Frame(canvas, bg=self.colors["bg_main"])
        window_id = canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        self.bind_scroll_canvas(canvas, body, window_id, min_width=1040)
        canvas.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        host.grid_rowconfigure(0, weight=1)
        host.grid_columnconfigure(0, weight=1)

        kpi_frame = tk.Frame(body, bg=self.colors["bg_main"])
        kpi_frame.pack(fill="x", pady=(0, 12))
        for title, value, color in [
            ("Clientes", self.formatear_numero(len(clientes)), self.colors["accent"]),
            ("Guias", self.formatear_numero(kpis.get("total_guias")), self.colors["info"]),
            ("Completas", self.formatear_numero(kpis.get("completas")), self.colors["success"]),
            ("Pendientes", self.formatear_numero(kpis.get("pendientes")), self.colors["warning"]),
            ("Descargado MT", self.formatear_numero(kpis.get("retirado_mt"), 2), self.colors["success"]),
            ("Pendiente descarga MT", self.formatear_numero(kpis.get("faltante_mt"), 2), self.colors["warning"]),
            ("Avance", f"{self.safe_number(kpis.get('avance_descarga_pct')):,.2f}%", self.colors["accent"]),
            ("Tiempo prom.", f"{self.safe_number(kpis.get('duracion_promedio_min')):,.2f} min", self.colors["muted"]),
            ("Prom. MT/camion", self.formatear_numero(kpis.get("promedio_mt_camion"), 2), self.colors["info"]),
            ("Viajes sugeridos", self.formatear_numero(kpis.get("viajes_estimados_necesarios")), self.colors["warning"]),
        ]:
            self.create_card(kpi_frame, title, value, color)

        if plan_viajes.get("mensaje"):
            plan_panel = tk.Frame(body, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1)
            plan_panel.pack(fill="x", pady=(0, 12))
            tk.Label(
                plan_panel,
                text="Plan inteligente de viajes",
                font=("Segoe UI", 13, "bold"),
                bg=self.colors["bg_card"],
                fg=self.colors["text_dark"],
            ).pack(anchor="w", padx=14, pady=(12, 4))
            tk.Label(
                plan_panel,
                text=plan_viajes.get("mensaje", ""),
                font=("Segoe UI", 10, "bold"),
                bg=self.colors["bg_card"],
                fg=self.colors["warning"] if plan_viajes.get("estado") != "BALANCEADO" else self.colors["success"],
                wraplength=1000,
                justify="left",
            ).pack(anchor="w", padx=14, pady=(0, 12))

        buque_panel = tk.Frame(body, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1)
        buque_panel.pack(fill="x", pady=(0, 12))
        tk.Label(buque_panel, text="Avance por bodega", font=("Segoe UI", 13, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w", padx=14, pady=(12, 6))
        ship_canvas = tk.Canvas(buque_panel, bg=self.colors["bg_card"], height=220, highlightthickness=0)
        ship_canvas.pack(fill="x", expand=True, padx=14, pady=(0, 12))
        old_canvas = getattr(self, "centro_buque_canvas", None)
        self.centro_buque_canvas = ship_canvas
        try:
            self.dibujar_buque_centro(bodegas)
        finally:
            self.centro_buque_canvas = old_canvas

        charts = tk.Frame(body, bg=self.colors["bg_main"])
        charts.pack(fill="both", expand=True, pady=(0, 12))
        clientes_grafico = graficos.get("cuota_vs_retiro_cliente") or graficos.get("retiro_por_cliente") or clientes
        self.crear_chart_barras(charts, "Cuota vs descargado por cliente", clientes_grafico, "empresa", "retirado_mt", 0, 0)
        self.crear_chart_barras(charts, "Saldo por cliente", clientes_grafico, "empresa", "saldo_mt", 0, 1)
        self.crear_chart_circular(charts, "Estado de guias", graficos.get("estado_guias", []), "estado", "valor", 0, 2)
        self.crear_chart_barras(charts, "Descargado por producto", graficos.get("retiro_por_producto", []), "producto", "retirado_mt", 1, 0)
        self.crear_chart_lineal(charts, "Tendencia diaria", graficos.get("tendencia_fecha", []), "fecha", "retirado_mt", 1, 1)
        self.crear_chart_barras(charts, "Duracion por camion", graficos.get("duracion_por_camion", []), "camion", "duracion_min", 1, 2)
        self.crear_chart_barras(charts, "Descargado por chofer", graficos.get("retiro_por_chofer", []), "camion", "retirado_mt", 2, 0)
        for col in range(3):
            charts.grid_columnconfigure(col, weight=1, uniform="popup_charts")
        for row in range(3):
            charts.grid_rowconfigure(row, weight=1, minsize=285)

        tabla_panel = tk.Frame(body, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1)
        tabla_panel.pack(fill="both", expand=True)
        tk.Label(tabla_panel, text="Detalle de cuotas por cliente y producto", font=("Segoe UI", 13, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w", padx=14, pady=(12, 6))
        table_frame = tk.Frame(tabla_panel, bg=self.colors["bg_card"])
        table_frame.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        columns = ("cliente", "producto", "cuota", "retirado", "faltante", "avance", "viajes")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=10)
        headings = {
            "cliente": "Cliente",
            "producto": "Producto",
            "cuota": "Cuota MT",
            "retirado": "Descargado MT",
            "faltante": "Pendiente/Sobredescarga MT",
            "avance": "Avance %",
            "viajes": "Viajes",
        }
        for col in columns:
            tree.heading(col, text=headings[col])
            tree.column(col, width=155, anchor="center")
        sy = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        sx = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        tree.grid(row=0, column=0, sticky="nsew")
        sy.grid(row=0, column=1, sticky="ns")
        sx.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        def clientes_filtrados_popup():
            cliente = filtro_cliente_var.get().strip()
            producto = filtro_producto_var.get().strip()
            cuota = filtro_cuota_var.get().strip()
            filtrados = []
            for row in clientes:
                if cliente and str(row.get("empresa") or "") != cliente:
                    continue
                if producto and str(row.get("producto") or "") != producto:
                    continue
                if cuota and self.formatear_numero(row.get("cuota_mt"), 2) != cuota:
                    continue
                filtrados.append(row)
            return filtrados

        def llenar_tabla_popup(rows):
            for item in tree.get_children():
                tree.delete(item)
            for row in rows:
                tree.insert("", "end", values=(
                row.get("empresa", ""),
                row.get("producto", ""),
                self.formatear_numero(row.get("cuota_mt"), 2),
                self.formatear_numero(row.get("retirado_mt"), 2),
                self.formatear_numero(row.get("saldo_mt", row.get("faltante_mt")), 2),
                f"{self.safe_number(row.get('avance_pct')):,.2f}%",
                self.formatear_numero(row.get("viajes")),
                ))

        def aplicar_filtros_popup():
            llenar_tabla_popup(clientes_filtrados_popup())

        def limpiar_filtros_popup():
            filtro_cliente_var.set("")
            filtro_producto_var.set("")
            filtro_cuota_var.set("")
            llenar_tabla_popup(clientes)

        def exportar_popup():
            formato_ui = exportar_formato_var.get().strip().lower()
            formato = "excel" if formato_ui == "excel" else "pdf"
            extension = ".xlsx" if formato == "excel" else ".pdf"
            filetypes = [("Excel", "*.xlsx")] if formato == "excel" else [("PDF", "*.pdf")]
            ruta = filedialog.asksaveasfilename(
                title=f"Exportar cuotas {formato_ui.upper()}",
                defaultextension=extension,
                filetypes=filetypes,
            )
            if not ruta:
                return
            params = {}
            if filtro_cliente_var.get().strip():
                params["empresa"] = filtro_cliente_var.get().strip()
            if filtro_producto_var.get().strip():
                params["producto"] = filtro_producto_var.get().strip()
            try:
                ruta_final = self.api_descargar_reporte_buque(operacion_id, formato, ruta, params)
                extra = "\n\nEl nombre cambio automaticamente porque el archivo original estaba bloqueado." if ruta_final != ruta else ""
                messagebox.showinfo("Reporte exportado", f"Archivo generado correctamente:\n{ruta_final}{extra}")
            except Exception as e:
                messagebox.showerror("Exportar cuotas", str(e))

        llenar_tabla_popup(clientes)

    # =========================================================
    # OPERACIONES BUQUE - VERSION INTEGRADA CON PRODUCTOS/CUOTAS
    # Estas definiciones sobrescriben las anteriores dentro de la clase.
    # =========================================================

    def meses_es(self):
        return [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
        ]

    def fecha_larga_es(self, fecha_iso):
        if not fecha_iso:
            return ""

        try:
            fecha = datetime.strptime(str(fecha_iso)[:10], "%Y-%m-%d").date()
            return f"{fecha.day} de {self.meses_es()[fecha.month - 1]} de {fecha.year}"
        except Exception:
            return str(fecha_iso)

    def meses_en(self):
        return [
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December",
        ]

    def fecha_larga_en(self, fecha_iso):
        if not fecha_iso:
            return ""
        try:
            fecha = datetime.strptime(str(fecha_iso)[:10], "%Y-%m-%d").date()
            weekday = calendar.day_name[fecha.weekday()]
            month = self.meses_en()[fecha.month - 1]
            return f"{weekday}, {month} {fecha.day}, {fecha.year}"
        except Exception:
            return str(fecha_iso)

    def abrir_selector_fecha_operacion(self):
        popup = tk.Toplevel(self)
        popup.title("Seleccionar fecha")
        popup.geometry("360x210")
        popup.configure(bg=self.colors["bg_card"])
        popup.transient(self)
        popup.grab_set()

        hoy = date.today()
        actual = self.operacion_form_vars.get("fecha_inicio").get() if hasattr(self, "operacion_form_vars") else ""

        try:
            base = datetime.strptime(actual, "%Y-%m-%d").date() if actual else hoy
        except Exception:
            base = hoy

        dia_var = tk.StringVar(value=str(base.day))
        mes_var = tk.StringVar(value=self.meses_es()[base.month - 1])
        anio_var = tk.StringVar(value=str(base.year))

        contenido = tk.Frame(popup, bg=self.colors["bg_card"])
        contenido.pack(fill="both", expand=True, padx=18, pady=18)

        tk.Label(contenido, text="Fecha de inicio", font=("Segoe UI", 13, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w", pady=(0, 12))

        fila = tk.Frame(contenido, bg=self.colors["bg_card"])
        fila.pack(fill="x")

        ttk.Combobox(fila, textvariable=dia_var, values=[str(i) for i in range(1, 32)], state="readonly", width=6).pack(side="left", padx=(0, 8))
        ttk.Combobox(fila, textvariable=mes_var, values=self.meses_es(), state="readonly", width=14).pack(side="left", padx=(0, 8))
        ttk.Combobox(fila, textvariable=anio_var, values=[str(i) for i in range(hoy.year - 3, hoy.year + 6)], state="readonly", width=8).pack(side="left")

        def aplicar():
            try:
                mes = self.meses_es().index(mes_var.get()) + 1
                fecha = date(int(anio_var.get()), mes, int(dia_var.get()))
            except Exception:
                messagebox.showerror("Fecha invÃ¡lida", "Seleccione una fecha vÃ¡lida.")
                return

            self.operacion_form_vars["fecha_inicio"].set(fecha.isoformat())
            self.operacion_form_vars["fecha_inicio_larga"].set(self.fecha_larga_es(fecha.isoformat()))
            popup.destroy()

        acciones = tk.Frame(contenido, bg=self.colors["bg_card"])
        acciones.pack(fill="x", pady=(18, 0))
        ttk.Button(acciones, text="Aplicar", style="Olive.TButton", command=aplicar).pack(side="left", padx=(0, 8))
        ttk.Button(acciones, text="Cancelar", style="Gray.TButton", command=popup.destroy).pack(side="left")

    def render_productos_operacion(self):
        for widget in self.productos_frame.winfo_children():
            widget.destroy()

        for idx, var in enumerate(self.producto_vars, start=1):
            row = tk.Frame(self.productos_frame, bg=self.colors["bg_card"])
            row.pack(fill="x", pady=2)
            tk.Label(row, text=f"Producto {idx}", width=12, anchor="w", font=("Segoe UI", 9), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(side="left")
            entry = ttk.Entry(row, textvariable=var)
            entry.pack(side="left", fill="x", expand=True)
            entry.bind("<KeyRelease>", self.refrescar_productos_operacion_dependientes)

    def productos_operacion_actuales(self):
        productos = []
        for var in getattr(self, "producto_vars", []):
            valor = var.get().strip()
            if valor and valor not in productos:
                productos.append(valor)
        return productos

    def producto_operacion_default(self):
        productos = self.productos_operacion_actuales()
        return productos[0] if productos else ""

    def producto_compacto_operacion(self, producto, limite=14):
        producto = str(producto or "").strip()
        if len(producto) <= limite:
            return producto
        return producto[: max(limite - 3, 1)] + "..."

    def refrescar_productos_operacion_dependientes(self, _event=None):
        productos = self.productos_operacion_actuales()
        for combo in getattr(self, "bodega_producto_combos", {}).values():
            try:
                combo["values"] = productos
            except tk.TclError:
                pass
        for combo in getattr(self, "particion_producto_combos", []):
            try:
                combo["values"] = productos
            except tk.TclError:
                pass
        combo = getattr(self, "cuota_producto_combo", None)
        if combo is not None:
            try:
                combo["values"] = productos
            except tk.TclError:
                pass
        for combo in getattr(self, "cuota_producto_combos", []):
            try:
                combo["values"] = productos
            except tk.TclError:
                pass
        if hasattr(self, "buque_silueta_canvas"):
            self.dibujar_silueta_buque()

    def render_bodegas_operacion(self):
        for widget in self.bodegas_frame.winfo_children():
            widget.destroy()

        self.bodega_producto_combos = {}
        self.particion_producto_combos = []
        label_fg = "#F4F8FF"
        productos = self.productos_operacion_actuales()
        for idx, numero in enumerate([5, 4, 3, 2, 1]):
            box = tk.Frame(self.bodegas_frame, bg=self.colors["bg_card"])
            box.grid(row=0, column=idx, sticky="nsew", padx=(0, 8), pady=(0, 2))
            self.bodegas_frame.grid_columnconfigure(idx, weight=1, minsize=230)
            tk.Label(
                box,
                text=f"Bodega {numero}",
                font=("Segoe UI", 8, "bold"),
                bg=self.colors["bg_card"],
                fg=label_fg,
            ).pack(anchor="w")
            header_row = tk.Frame(box, bg=self.colors["bg_card"])
            header_row.pack(fill="x", pady=(0, 2))
            for col, text in enumerate(("MT", "Producto")):
                header_row.grid_columnconfigure(col, weight=1, uniform=f"bodega_{numero}")
                tk.Label(
                    header_row,
                    text=text,
                    font=("Segoe UI", 7, "bold"),
                    bg=self.colors["bg_card"],
                    fg=self.colors["text_secondary"],
                ).grid(row=0, column=col, sticky="w", padx=(0 if col == 0 else 4, 0))
            base_row = tk.Frame(box, bg=self.colors["bg_card"])
            base_row.pack(fill="x")
            base_row.grid_columnconfigure(0, weight=1, uniform=f"bodega_{numero}")
            base_row.grid_columnconfigure(1, weight=1, uniform=f"bodega_{numero}")
            entry = ttk.Entry(base_row, textvariable=self.bodega_vars[numero], width=10)
            entry.grid(row=0, column=0, sticky="ew", padx=(0, 4))
            entry.bind("<KeyRelease>", lambda _event: self.dibujar_silueta_buque())
            producto_combo = ttk.Combobox(
                base_row,
                textvariable=self.bodega_producto_vars[numero],
                values=productos,
                state="readonly",
                width=10,
            )
            producto_combo.grid(row=0, column=1, sticky="ew", padx=(0, 4))
            producto_combo.bind("<<ComboboxSelected>>", lambda _event: self.dibujar_silueta_buque())
            self.bodega_producto_combos[numero] = producto_combo
            for particion in self.bodega_particion_vars.get(numero, []):
                row = tk.Frame(box, bg=self.colors["bg_card"])
                row.pack(fill="x", pady=(4, 0))
                row.grid_columnconfigure(0, weight=1, uniform=f"bodega_{numero}")
                row.grid_columnconfigure(1, weight=1, uniform=f"bodega_{numero}")
                entry_mt = ttk.Entry(row, textvariable=particion["capacidad_mt"], width=10)
                entry_mt.grid(row=0, column=0, sticky="ew", padx=(0, 4))
                entry_mt.bind("<KeyRelease>", lambda _event: self.dibujar_silueta_buque())
                part_producto = ttk.Combobox(
                    row,
                    textvariable=particion["producto"],
                    values=productos,
                    state="readonly",
                    width=10,
                )
                part_producto.grid(row=0, column=1, sticky="ew")
                part_producto.bind("<<ComboboxSelected>>", lambda _event: self.dibujar_silueta_buque())
                self.particion_producto_combos.append(part_producto)
            part_actions = tk.Frame(box, bg=self.colors["bg_card"])
            part_actions.pack(fill="x", pady=(6, 2))
            ttk.Button(part_actions, text="+", style="Olive.TButton", command=lambda n=numero: self.agregar_particion_bodega(n)).pack(side="left", fill="x", expand=True, padx=(0, 2))
            ttk.Button(part_actions, text="-", style="Gray.TButton", command=lambda n=numero: self.eliminar_particion_bodega(n)).pack(side="left", fill="x", expand=True)

    def agregar_particion_bodega(self, numero):
        lineas = self.bodega_particion_vars.setdefault(numero, [])
        if len(lineas) >= 3:
            messagebox.showwarning("Limite de particiones", "Puede agregar hasta 3 particiones por bodega.")
            return
        producto_default = ""
        if hasattr(self, "bodega_producto_vars") and numero in self.bodega_producto_vars:
            producto_default = self.bodega_producto_vars[numero].get().strip()
        producto_default = producto_default or self.producto_operacion_default()
        lineas.append({
            "capacidad_mt": tk.StringVar(value="0"),
            "producto": tk.StringVar(value=producto_default),
        })
        self.render_bodegas_operacion()
        self.dibujar_silueta_buque()

    def eliminar_particion_bodega(self, numero):
        lineas = self.bodega_particion_vars.setdefault(numero, [])
        if lineas:
            lineas.pop()
        self.render_bodegas_operacion()
        self.dibujar_silueta_buque()

    def dibujar_silueta_buque(self):
        if not hasattr(self, "buque_silueta_canvas"):
            return

        canvas = self.buque_silueta_canvas
        canvas.delete("all")
        width = max(canvas.winfo_width(), 720)
        y_top = 18
        y_bottom = 150
        y_mid = (y_top + y_bottom) / 2
        x0 = 20
        x1 = width - 20
        canvas.create_polygon(x0, y_top, x1 - 55, y_top, x1, y_mid, x1 - 55, y_bottom, x0, y_bottom, fill=self.colors["bg_topbar"], outline=self.colors["accent"], width=2)

        total = sum(max(self.safe_number(var.get(), 0), 0) for var in self.bodega_vars.values()) or 1
        usable_w = x1 - x0 - 80
        start_x = x0 + 18
        colors = {
            1: self.colors["teal"],
            2: self.colors["info"],
            3: self.colors["accent_light"],
            4: self.colors["success"],
            5: self.colors["accent"],
        }

        for idx, numero in enumerate([5, 4, 3, 2, 1]):
            cap = max(self.safe_number(self.bodega_vars[numero].get(), 0), 0)
            seg_w = max(70, usable_w / 5)
            x = start_x + idx * seg_w
            rect_x0 = x
            rect_x1 = x + seg_w - 10
            rect_y0 = y_top + 12
            rect_y1 = y_bottom - 12
            canvas.create_rectangle(rect_x0, rect_y0, rect_x1, rect_y1, fill=self.colors["bg_main"], outline=self.colors["border"])

            segmentos = []
            producto_base = ""
            if hasattr(self, "bodega_producto_vars") and numero in self.bodega_producto_vars:
                producto_base = self.bodega_producto_vars[numero].get().strip()
            if cap > 0:
                segmentos.append((cap, producto_base or self.producto_operacion_default()))
            for item in self.bodega_particion_vars.get(numero, []):
                valor = max(self.safe_number(item["capacidad_mt"].get(), 0), 0)
                if valor > 0:
                    producto_part = item.get("producto").get().strip() if item.get("producto") else ""
                    segmentos.append((valor, producto_part))
            total_segmentos = sum(valor for valor, _producto in segmentos)

            if segmentos and total_segmentos > 0:
                cursor_y = rect_y1
                base_h = max(rect_y1 - rect_y0, 1)
                denominator = total_segmentos
                accumulated = 0
                for part_idx, (valor, _producto_part) in enumerate(segmentos):
                    accumulated = min(accumulated + valor, denominator)
                    next_y = rect_y1 - base_h * (accumulated / denominator)
                    canvas.create_rectangle(rect_x0, next_y, rect_x1, cursor_y, fill=colors[numero], outline="")
                    if part_idx > 0:
                        canvas.create_line(rect_x0 + 3, cursor_y, rect_x1 - 3, cursor_y, fill=self.colors["text_light"], width=3)
                        canvas.create_line(rect_x0 + 3, cursor_y, rect_x1 - 3, cursor_y, fill=self.colors["bg_main"], width=1)
                    cursor_y = next_y
                    if cursor_y <= rect_y0:
                        break
                detalle = []
                for valor, producto_segmento in segmentos[:3]:
                    producto_segmento = self.producto_compacto_operacion(producto_segmento, 7) or "PROD"
                    detalle.append(f"{producto_segmento} {valor:,.0f}")
                if len(segmentos) > 3:
                    detalle.append(f"+{len(segmentos) - 3} mas")
                label_text = f"B{numero}\n" + "\n".join(detalle)
                if len(segmentos) == 1:
                    label_text += " MT"
                else:
                    label_text += f"\nTot. {total_segmentos:,.0f} MT"
            else:
                label_text = f"B{numero}\n0 MT"

            text_x = rect_x0 + (rect_x1 - rect_x0) / 2
            text_fill = "#050B14" if total_segmentos > 0 else "#F4F8FF"
            shadow_fill = "#F4F8FF" if total_segmentos > 0 else "#050B14"
            canvas.create_text(
                text_x + 1,
                y_mid + 1,
                text=label_text,
                fill=shadow_fill,
                font=("Segoe UI", 8, "bold"),
                justify="center",
                width=max(rect_x1 - rect_x0 - 8, 60),
            )
            canvas.create_text(
                text_x,
                y_mid,
                text=label_text,
                fill=text_fill,
                font=("Segoe UI", 8, "bold"),
                justify="center",
                width=max(rect_x1 - rect_x0 - 8, 60),
            )

    def agregar_producto_operacion(self):
        if len(self.producto_vars) >= 5:
            messagebox.showwarning("LÃ­mite de productos", "Puede agregar hasta 5 productos por operaciÃ³n.")
            return

        self.producto_vars.append(tk.StringVar())
        self.render_productos_operacion()

    def eliminar_producto_operacion(self):
        if len(self.producto_vars) <= 1:
            self.producto_vars[0].set("")
            return

        self.producto_vars.pop()
        self.render_productos_operacion()

    def show_operaciones_buque(self):
        self.clear_content()
        self.highlight_sidebar_button("Operaciones Buque")

        self.operacion_form_vars = {
            "nombre_buque": tk.StringVar(),
            "fecha_inicio": tk.StringVar(),
            "fecha_inicio_larga": tk.StringVar(value="Seleccione una fecha"),
        }
        self.producto_vars = [tk.StringVar()]
        self.bodega_vars = {numero: tk.StringVar(value="0") for numero in range(1, 6)}
        self.bodega_producto_vars = {numero: tk.StringVar() for numero in range(1, 6)}
        self.bodega_particion_vars = {numero: [] for numero in range(1, 6)}
        self.cuota_form_vars = {
            "operacion": tk.StringVar(),
            "cliente": tk.StringVar(),
            "producto": tk.StringVar(),
            "bodega": tk.StringVar(value="Todas"),
            "cuota": tk.StringVar(),
            "unidad": tk.StringVar(value="MT"),
        }
        self.cuota_operaciones_map = {}
        self.cuota_editing_item = None
        self.cuota_lineas_vars = []
        self.cuota_producto_combos = []
        self.operaciones_tree = None
        self.cuotas_tree = None

        self.create_page_title(
            self.content,
            "Operaciones de Buque",
            "Abra una operacion completa con buque, fecha, productos y cuotas por cliente.",
        )

        scroll_host = tk.Frame(self.content, bg=self.colors["bg_main"])
        scroll_host.pack(fill="both", expand=True, padx=25, pady=(0, 20))

        canvas = tk.Canvas(scroll_host, bg=self.colors["bg_main"], highlightthickness=0)
        scroll_y = ttk.Scrollbar(scroll_host, orient="vertical", command=canvas.yview)
        scroll_x = ttk.Scrollbar(scroll_host, orient="horizontal", command=canvas.xview)
        body = tk.Frame(canvas, bg=self.colors["bg_main"])

        window_id = canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        self.bind_scroll_canvas(canvas, body, window_id, min_width=1500)

        canvas.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        scroll_host.grid_rowconfigure(0, weight=1)
        scroll_host.grid_columnconfigure(0, weight=1)

        top_panel = tk.Frame(body, bg=self.colors["bg_main"])
        top_panel.pack(fill="x", pady=(0, 10))

        form_panel = tk.Frame(top_panel, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1)
        form_panel.pack(side="left", fill="both", expand=True, padx=(0, 8))

        tk.Label(form_panel, text="Nueva operacion", font=("Segoe UI", 13, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w", padx=14, pady=(12, 6))

        form_grid = tk.Frame(form_panel, bg=self.colors["bg_card"])
        form_grid.pack(fill="x", padx=14, pady=(0, 10))

        buque_box = tk.Frame(form_grid, bg=self.colors["bg_card"])
        buque_box.grid(row=0, column=0, sticky="ew", padx=5, pady=4)
        tk.Label(buque_box, text="Buque", font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
        ttk.Entry(buque_box, textvariable=self.operacion_form_vars["nombre_buque"]).pack(fill="x")

        fecha_box = tk.Frame(form_grid, bg=self.colors["bg_card"])
        fecha_box.grid(row=0, column=1, sticky="ew", padx=5, pady=4)
        tk.Label(fecha_box, text="Fecha inicio", font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
        fecha_row = tk.Frame(fecha_box, bg=self.colors["bg_card"])
        fecha_row.pack(fill="x")
        ttk.Entry(fecha_row, textvariable=self.operacion_form_vars["fecha_inicio_larga"], state="readonly").pack(side="left", fill="x", expand=True)
        ttk.Button(fecha_row, text="Calendario", style="Gray.TButton", command=self.abrir_selector_fecha_operacion).pack(side="left", padx=(6, 0))

        productos_box = tk.Frame(form_grid, bg=self.colors["bg_card"])
        productos_box.grid(row=1, column=0, columnspan=2, sticky="ew", padx=5, pady=(8, 4))
        header_productos = tk.Frame(productos_box, bg=self.colors["bg_card"])
        header_productos.pack(fill="x")
        tk.Label(header_productos, text="Productos", font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(side="left")

        self.productos_frame = tk.Frame(productos_box, bg=self.colors["bg_card"])
        self.productos_frame.pack(fill="x", pady=(4, 0))
        self.render_productos_operacion()

        productos_actions = tk.Frame(productos_box, bg=self.colors["bg_card"])
        productos_actions.pack(fill="x", pady=(6, 0))
        ttk.Button(productos_actions, text="+", style="Olive.TButton", command=self.agregar_producto_operacion).pack(side="right", fill="x", expand=True, padx=(4, 0))
        ttk.Button(productos_actions, text="-", style="Gray.TButton", command=self.eliminar_producto_operacion).pack(side="right", fill="x", expand=True)

        bodegas_box = tk.Frame(form_grid, bg=self.colors["bg_card"])
        bodegas_box.grid(row=2, column=0, columnspan=2, sticky="ew", padx=5, pady=(10, 4))
        tk.Label(
            bodegas_box,
            text="Capacidad por bodega (MT)",
            font=("Segoe UI", 9, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w")

        self.bodegas_frame = tk.Frame(bodegas_box, bg=self.colors["bg_card"])
        self.bodegas_frame.pack(fill="x", pady=(4, 0))
        self.render_bodegas_operacion()

        self.buque_silueta_canvas = tk.Canvas(bodegas_box, bg=self.colors["bg_card"], height=172, highlightthickness=0)
        self.buque_silueta_canvas.pack(fill="x", pady=(8, 0))
        self.dibujar_silueta_buque()

        form_grid.grid_columnconfigure(0, weight=1)
        form_grid.grid_columnconfigure(1, weight=1)

        form_actions = tk.Frame(form_panel, bg=self.colors["bg_card"])
        form_actions.pack(fill="x", padx=14, pady=(0, 12))
        ttk.Button(form_actions, text="Abrir operacion", style="Olive.TButton", command=self.crear_operacion_buque_front).pack(side="left", padx=(0, 8))
        ttk.Button(form_actions, text="Limpiar", style="Gray.TButton", command=self.limpiar_form_operacion).pack(side="left")

        cuotas_panel = tk.Frame(body, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1)
        cuotas_panel.pack(fill="both", expand=True)

        tk.Label(cuotas_panel, text="Cuotas por cliente", font=("Segoe UI", 13, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w", padx=14, pady=(12, 6))

        cuota_form = tk.Frame(cuotas_panel, bg=self.colors["bg_card"])
        cuota_form.pack(fill="x", padx=14, pady=(0, 8))

        operacion_cuota_box = tk.Frame(cuota_form, bg=self.colors["bg_card"])
        operacion_cuota_box.pack(fill="x", padx=5, pady=(0, 8))
        tk.Label(
            operacion_cuota_box,
            text="Operacion para asignar cuotas",
            font=("Segoe UI", 9, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w")
        operacion_cuota_row = tk.Frame(operacion_cuota_box, bg=self.colors["bg_card"])
        operacion_cuota_row.pack(fill="x")
        self.cuota_operacion_combo = ttk.Combobox(
            operacion_cuota_row,
            textvariable=self.cuota_form_vars["operacion"],
            values=[],
            state="readonly",
        )
        self.cuota_operacion_combo.pack(side="left", fill="x", expand=True)
        self.cuota_operacion_combo.bind("<<ComboboxSelected>>", lambda _event: self.actualizar_operacion_cuota_seleccionada())
        ttk.Button(
            operacion_cuota_row,
            text="Refrescar operaciones",
            style="Gray.TButton",
            command=self.cargar_operaciones_cuota_combo,
        ).pack(side="left", padx=(8, 0))

        cuota_grid = tk.Frame(cuota_form, bg=self.colors["bg_card"])
        cuota_grid.pack(fill="x")

        for idx, (key, label) in enumerate([("cliente", "Cliente"), ("cuota", "Cuota")]):
            box = tk.Frame(cuota_grid, bg=self.colors["bg_card"])
            box.grid(row=0, column=0 if key == "cliente" else 3, sticky="ew", padx=5, pady=3)
            tk.Label(box, text=label, font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
            ttk.Entry(box, textvariable=self.cuota_form_vars[key]).pack(fill="x")

        producto_cuota_box = tk.Frame(cuota_grid, bg=self.colors["bg_card"])
        producto_cuota_box.grid(row=0, column=1, sticky="ew", padx=5, pady=3)
        tk.Label(producto_cuota_box, text="Producto", font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
        self.cuota_producto_combo = ttk.Combobox(
            producto_cuota_box,
            textvariable=self.cuota_form_vars["producto"],
            values=self.productos_operacion_actuales(),
            state="readonly",
        )
        self.cuota_producto_combo.pack(fill="x")

        bodega_cuota_box = tk.Frame(cuota_grid, bg=self.colors["bg_card"])
        bodega_cuota_box.grid(row=0, column=2, sticky="ew", padx=5, pady=3)
        tk.Label(bodega_cuota_box, text="Bodega", font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
        bodega_cuota_row = tk.Frame(bodega_cuota_box, bg=self.colors["bg_card"])
        bodega_cuota_row.pack(fill="x")
        ttk.Entry(bodega_cuota_row, textvariable=self.cuota_form_vars["bodega"], state="readonly").pack(side="left", fill="x", expand=True)
        ttk.Button(
            bodega_cuota_row,
            text="...",
            style="Gray.TButton",
            command=lambda: self.abrir_selector_bodegas_cuota(self.cuota_form_vars["bodega"]),
            width=4,
        ).pack(side="left", padx=(4, 0))

        unidad_box = tk.Frame(cuota_grid, bg=self.colors["bg_card"])
        unidad_box.grid(row=0, column=4, sticky="ew", padx=5, pady=3)
        tk.Label(unidad_box, text="Unidad", font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
        ttk.Combobox(unidad_box, textvariable=self.cuota_form_vars["unidad"], values=["MT"], state="readonly").pack(fill="x")

        for col in range(5):
            cuota_grid.grid_columnconfigure(col, weight=1)

        self.cuota_lineas_frame = tk.Frame(cuotas_panel, bg=self.colors["bg_card"])
        self.cuota_lineas_frame.pack(fill="x", padx=14, pady=(0, 8))

        cuota_actions = tk.Frame(cuotas_panel, bg=self.colors["bg_card"])
        cuota_actions.pack(fill="x", padx=14, pady=(0, 8))
        ttk.Button(cuota_actions, text="+", style="Olive.TButton", command=self.agregar_cuota_local_front).pack(side="left", padx=(0, 8))
        ttk.Button(cuota_actions, text="Crear cuotas", style="Olive.TButton", command=self.guardar_cuotas_lote_front).pack(side="left", padx=(0, 8))
        ttk.Button(cuota_actions, text="Editar", style="Gray.TButton", command=self.editar_cuota_seleccionada).pack(side="left", padx=(0, 8))
        ttk.Button(cuota_actions, text="-", style="Gray.TButton", command=self.eliminar_cuota_seleccionada).pack(side="left", padx=(0, 8))
        ttk.Button(cuota_actions, text="Cargar cuotas activas", style="Gray.TButton", command=self.cargar_cuotas_operacion_activa).pack(side="left")

        cuotas_columns = ("id", "cliente", "producto", "bodega", "cuota", "unidad")
        cuotas_table_frame = tk.Frame(cuotas_panel, bg=self.colors["bg_card"])
        cuotas_table_frame.pack(fill="both", expand=True, padx=14, pady=(0, 14))

        self.cuotas_tree = ttk.Treeview(cuotas_table_frame, columns=cuotas_columns, show="headings", height=12)
        for col, heading, width in [("id", "ID", 45), ("cliente", "Cliente", 230), ("producto", "Producto", 210), ("bodega", "Bodega", 110), ("cuota", "Cuota", 120), ("unidad", "Unidad", 90)]:
            self.cuotas_tree.heading(col, text=heading)
            self.cuotas_tree.column(col, width=width, anchor="center")

        cuotas_scroll = ttk.Scrollbar(cuotas_table_frame, orient="vertical", command=self.cuotas_tree.yview)
        cuotas_scroll_x = ttk.Scrollbar(cuotas_table_frame, orient="horizontal", command=self.cuotas_tree.xview)
        self.cuotas_tree.configure(yscrollcommand=cuotas_scroll.set, xscrollcommand=cuotas_scroll_x.set)
        self.cuotas_tree.grid(row=0, column=0, sticky="nsew")
        cuotas_scroll.grid(row=0, column=1, sticky="ns")
        cuotas_scroll_x.grid(row=1, column=0, sticky="ew")
        cuotas_table_frame.grid_rowconfigure(0, weight=1)
        cuotas_table_frame.grid_columnconfigure(0, weight=1)

        self.operacion_activa = None
        self.actualizar_operacion_activa_label()
        self.after(150, self.cargar_operaciones_cuota_combo)

    def limpiar_form_operacion(self):
        for var in self.operacion_form_vars.values():
            var.set("")

        self.operacion_form_vars["fecha_inicio_larga"].set("Seleccione una fecha")
        self.producto_vars = [tk.StringVar()]
        self.bodega_vars = {numero: tk.StringVar(value="0") for numero in range(1, 6)}
        self.bodega_producto_vars = {numero: tk.StringVar() for numero in range(1, 6)}
        self.bodega_particion_vars = {numero: [] for numero in range(1, 6)}

        if hasattr(self, "productos_frame"):
            self.render_productos_operacion()
        if hasattr(self, "bodegas_frame"):
            self.render_bodegas_operacion()
        if hasattr(self, "buque_silueta_canvas"):
            self.dibujar_silueta_buque()

        self.limpiar_form_cuota()
        if hasattr(self, "cuota_lineas_vars"):
            self.cuota_lineas_vars = []
            self.cuota_producto_combos = []
            self.render_lineas_cuota_front()

        if self.cuotas_tree is not None:
            for item in self.cuotas_tree.get_children():
                self.cuotas_tree.delete(item)

    def limpiar_form_cuota(self):
        if not hasattr(self, "cuota_form_vars"):
            return

        for key, var in self.cuota_form_vars.items():
            if key == "operacion":
                continue
            if key == "unidad":
                var.set("MT")
            elif key == "bodega":
                var.set("Todas")
            else:
                var.set("")
        productos = self.productos_operacion_actuales() if hasattr(self, "producto_vars") else []
        if productos and hasattr(self, "cuota_form_vars"):
            self.cuota_form_vars["producto"].set(productos[0])


    def cargar_operaciones_cuota_combo(self):
        if not hasattr(self, "cuota_operacion_combo"):
            return

        def tarea():
            data = self.api_get_operaciones_buque(estado="ABIERTA")
            operaciones = data.get("data", []) if isinstance(data, dict) else []
            if not operaciones:
                data_todas = self.api_get_operaciones_buque()
                todas = data_todas.get("data", []) if isinstance(data_todas, dict) else []
                operaciones = [
                    op for op in todas
                    if isinstance(op, dict) and str(op.get("estado") or "").strip().upper() == "ABIERTA"
                ]
            activa = None
            try:
                activa = self.api_get_operacion_activa()
            except Exception:
                activa = None
            return {"operaciones": operaciones, "activa": activa}

        def al_terminar(resultado):
            resultado = resultado if isinstance(resultado, dict) else {}
            operaciones = resultado.get("operaciones", []) or []
            activa = resultado.get("activa")
            if not isinstance(activa, dict):
                posible_activa = getattr(self, "operacion_activa", None)
                activa = posible_activa if isinstance(posible_activa, dict) else None
            if isinstance(activa, dict) and activa.get("id"):
                activa_id = self.safe_int(activa.get("id"), None)
                if activa_id and not any(
                    isinstance(op, dict) and self.safe_int(op.get("id"), None) == activa_id
                    for op in operaciones
                ):
                    operaciones.insert(0, activa)
            operaciones_map = {}
            valores = []

            for op in operaciones:
                if not isinstance(op, dict):
                    continue
                op_id = op.get("id")
                if not op_id:
                    continue
                etiqueta = (
                    f"{op_id} | {op.get('nombre_buque', '')} | "
                    f"{self.fecha_larga_es(op.get('fecha_inicio'))} | {op.get('estado', '')}"
                )
                operaciones_map[etiqueta] = op
                valores.append(etiqueta)

            self.cuota_operaciones_map = operaciones_map
            self.cuota_operacion_combo["values"] = valores
            seleccionado = self.cuota_form_vars["operacion"].get()

            if seleccionado in valores:
                self.operacion_activa = self.cuota_operaciones_map.get(seleccionado)
                self.actualizar_operacion_activa_label()
                return

            if isinstance(activa, dict):
                for etiqueta, op in self.cuota_operaciones_map.items():
                    if self.safe_int(op.get("id"), None) == self.safe_int(activa.get("id"), None):
                        self.cuota_form_vars["operacion"].set(etiqueta)
                        self.operacion_activa = op
                        self.actualizar_operacion_activa_label()
                        return

            if valores:
                self.cuota_form_vars["operacion"].set(valores[0])
                self.operacion_activa = self.cuota_operaciones_map.get(valores[0])
                self.actualizar_operacion_activa_label()
                return

            self.cuota_form_vars["operacion"].set("")
            self.operacion_activa = None
            self.actualizar_operacion_activa_label()
            if self.cuotas_tree is not None:
                for item in self.cuotas_tree.get_children():
                    self.cuotas_tree.delete(item)

        self.ejecutar_en_segundo_plano(
            "Operaciones para cuotas",
            "Buscando operaciones abiertas...",
            tarea,
            al_terminar,
        )

    def actualizar_operacion_cuota_seleccionada(self):
        etiqueta = self.cuota_form_vars["operacion"].get().strip() if hasattr(self, "cuota_form_vars") else ""
        self.operacion_activa = getattr(self, "cuota_operaciones_map", {}).get(etiqueta)
        self.actualizar_operacion_activa_label()
        self.actualizar_productos_cuota_desde_operacion({"operacion": self.operacion_activa})
        if self.cuotas_tree is not None:
            for item in self.cuotas_tree.get_children():
                self.cuotas_tree.delete(item)

    def obtener_operacion_cuota_id(self):
        if not hasattr(self, "cuota_form_vars"):
            return None

        etiqueta = self.cuota_form_vars["operacion"].get().strip()
        op = getattr(self, "cuota_operaciones_map", {}).get(etiqueta)
        if op:
            return self.safe_int(op.get("id"), None)

        if etiqueta:
            posible_id = etiqueta.split("|", 1)[0].strip()
            operacion_id = self.safe_int(posible_id, None)
            if operacion_id:
                return operacion_id

        activa = getattr(self, "operacion_activa", None)
        if isinstance(activa, dict):
            return self.safe_int(activa.get("id"), None)

        return None

    def productos_cuota_desde_detalle_operacion(self, detalle=None):
        productos = []

        def agregar(valor):
            for parte in str(valor or "").replace(",", "/").split("/"):
                producto = parte.strip()
                if producto and producto.upper() not in ("TODOS", "SIN PRODUCTO", "NONE", "NULL"):
                    productos.append(producto)

        detalle = detalle if isinstance(detalle, dict) else {}
        operacion = detalle.get("operacion") if isinstance(detalle.get("operacion"), dict) else detalle
        if isinstance(operacion, dict):
            agregar(operacion.get("producto"))
            for bodega in operacion.get("bodegas") or []:
                if isinstance(bodega, dict):
                    agregar(bodega.get("producto"))
        for bodega in detalle.get("bodegas") or []:
            if isinstance(bodega, dict):
                agregar(bodega.get("producto"))
        for particion in detalle.get("bodega_particiones") or []:
            if isinstance(particion, dict):
                agregar(particion.get("producto"))
        for producto in self.productos_operacion_actuales() if hasattr(self, "producto_vars") else []:
            agregar(producto)
        return sorted(dict.fromkeys(productos))

    def actualizar_productos_cuota_desde_operacion(self, detalle=None):
        productos = self.productos_cuota_desde_detalle_operacion(detalle)
        if not productos or not hasattr(self, "cuota_producto_combo"):
            return
        self.cuota_producto_combo["values"] = productos
        actual = self.cuota_form_vars["producto"].get().strip() if hasattr(self, "cuota_form_vars") else ""
        if not actual or actual.upper() in ("CUOTA GENERAL", "SIN PRODUCTO", "TODOS", "ALL"):
            self.cuota_form_vars["producto"].set(productos[0])

    def cargar_cuotas_tabla_operacion_id(self, operacion_id):
        if self.cuotas_tree is None or not operacion_id:
            return []

        data = self.api_get_cuotas_buque(operacion_id)
        cuotas = data.get("data", [])

        for item in self.cuotas_tree.get_children():
            self.cuotas_tree.delete(item)

        for cuota in cuotas:
            self.cuotas_tree.insert(
                "",
                "end",
                values=(
                    cuota.get("id", ""),
                    cuota.get("cliente", ""),
                    self.producto_cuota_visible(cuota.get("producto")),
                    self.bodega_cuota_visible(cuota.get("bodega_numeros") or cuota.get("bodega_numero")),
                    self.formatear_numero(cuota.get("cuota"), 2),
                    cuota.get("unidad", ""),
                ),
            )

        return cuotas

    def cuota_a_mt(self, cuota, unidad):
        valor = self.safe_number(cuota, 0)
        return valor

    def capacidad_total_operacion_mt(self, operacion_id):
        total = 0
        try:
            detalle = self.api_get_operacion_detalle(operacion_id)
            bodegas = detalle.get("bodegas") or detalle.get("operacion", {}).get("bodegas") or []
            for bodega in bodegas:
                total += max(self.safe_number(bodega.get("capacidad_mt"), 0), 0)
        except Exception:
            total = 0

        if total <= 0 and hasattr(self, "bodega_vars"):
            total = sum(max(self.safe_number(var.get(), 0), 0) for var in self.bodega_vars.values())
        return total

    def alerta_cobertura_cuotas(self, operacion_id, cuotas=None):
        if not operacion_id:
            return ""
        capacidad_total = self.capacidad_total_operacion_mt(operacion_id)
        if capacidad_total <= 0:
            return ""
        if cuotas is None:
            cuotas = self.api_get_cuotas_buque(operacion_id).get("data", [])
        cuotas_total = sum(self.cuota_a_mt(cuota.get("cuota"), cuota.get("unidad")) for cuota in cuotas or [])
        diferencia = capacidad_total - cuotas_total
        if abs(diferencia) <= 0.01:
            return ""
        if diferencia > 0:
            return (
                "Cuotas incompletas.\n\n"
                f"Capacidad total del buque: {self.formatear_numero(capacidad_total, 2)} MT\n"
                f"Cuotas asignadas: {self.formatear_numero(cuotas_total, 2)} MT\n"
                f"Pendiente por asignar: {self.formatear_numero(diferencia, 2)} MT"
            )
        return (
            "Cuotas excedidas.\n\n"
            f"Capacidad total del buque: {self.formatear_numero(capacidad_total, 2)} MT\n"
            f"Cuotas asignadas: {self.formatear_numero(cuotas_total, 2)} MT\n"
            f"Exceso asignado: {self.formatear_numero(abs(diferencia), 2)} MT"
        )

    def cargar_cuotas_operacion_seleccionada_combo(self):
        operacion_id = self.obtener_operacion_cuota_id()
        if not operacion_id:
            return

        try:
            data = self.api_get_operacion_detalle(operacion_id)
            self.operacion_activa = data.get("operacion") or self.operacion_activa
            self.actualizar_operacion_activa_label()
            self.actualizar_productos_cuota_desde_operacion(data)
            self.cargar_cuotas_tabla_operacion_id(operacion_id)
        except Exception as e:
            messagebox.showerror("Cuotas por operacion", str(e))

    def cargar_cuotas_operacion_activa(self):
        operacion_id = self.obtener_operacion_cuota_id()
        if not operacion_id:
            messagebox.showwarning("Sin operacion", "Seleccione una operacion abierta para consultar cuotas.")
            return

        def tarea():
            detalle = self.api_get_operacion_detalle(operacion_id)
            cuotas = self.api_get_cuotas_buque(operacion_id).get("data", [])
            return {"detalle": detalle, "cuotas": cuotas}

        def al_terminar(resultado):
            detalle = resultado.get("detalle", {})
            self.operacion_activa = detalle.get("operacion")
            self.actualizar_operacion_activa_label()
            self.actualizar_productos_cuota_desde_operacion(detalle)
            cuotas = resultado.get("cuotas", [])
            if self.cuotas_tree is not None:
                for item in self.cuotas_tree.get_children():
                    self.cuotas_tree.delete(item)
                for cuota in cuotas:
                    self.cuotas_tree.insert(
                        "",
                        "end",
                        values=(
                            cuota.get("id", ""),
                            cuota.get("cliente", ""),
                            self.producto_cuota_visible(cuota.get("producto")),
                            self.bodega_cuota_visible(cuota.get("bodega_numeros") or cuota.get("bodega_numero")),
                            self.formatear_numero(cuota.get("cuota"), 2),
                            cuota.get("unidad", ""),
                        ),
                    )
            if not cuotas:
                messagebox.showinfo("Cuotas activas", "No hay cuotas registradas para la operacion seleccionada.")
            alerta = self.alerta_cobertura_cuotas(operacion_id, cuotas)
            if alerta:
                messagebox.showwarning("Cobertura de cuotas", alerta)

        self.ejecutar_en_segundo_plano(
            "Cuotas operacion",
            "Cargando cuotas de la operacion seleccionada...",
            tarea,
            al_terminar,
        )

    def editar_cuota_seleccionada(self):
        if self.cuotas_tree is None:
            return
        seleccion = self.cuotas_tree.selection()
        if not seleccion:
            messagebox.showwarning("Sin seleccion", "Seleccione una cuota para editar.")
            return
        valores = self.cuotas_tree.item(seleccion[0], "values")
        if len(valores) >= 4:
            cuota_id, cliente, producto, bodega, cuota, unidad = self.normalizar_valores_cuota_tree(valores)
            self.cuota_editing_item = seleccion[0]
            productos_disponibles = list(self.cuota_producto_combo["values"]) if hasattr(self, "cuota_producto_combo") else []
            if str(producto or "").strip().upper() in ("CUOTA GENERAL", "SIN PRODUCTO", "TODOS", "ALL") and productos_disponibles:
                producto = productos_disponibles[0]
            self.cuota_form_vars["cliente"].set(cliente)
            self.cuota_form_vars["producto"].set(producto)
            self.cuota_form_vars["bodega"].set(bodega or "Todas")
            self.cuota_form_vars["cuota"].set(str(cuota).replace(",", ""))
            self.cuota_form_vars["unidad"].set(unidad or "MT")
            messagebox.showinfo("Editar cuota", "Modifique los campos y presione + para guardar los cambios.")

    def refrescar_operacion_activa_front(self):
        def tarea():
            return self.api_get_operacion_activa()

        def al_terminar(resultado):
            self.operacion_activa = resultado
            self.actualizar_operacion_activa_label()

        def al_error(e):
            if hasattr(self, "operacion_activa_label"):
                try:
                    if not self.operacion_activa_label.winfo_exists():
                        return
                except Exception:
                    return

                self.operacion_activa_label.configure(
                    text=f"Operacion activa:\nError consultando API\n{e}",
                    fg=self.colors["danger"],
                )

        def tarea_segura():
            try:
                return {"ok": True, "data": tarea()}
            except Exception as exc:
                return {"ok": False, "error": exc}

        def finalizar(resultado):
            if resultado.get("ok"):
                al_terminar(resultado.get("data"))
            else:
                al_error(resultado.get("error"))

        self.ejecutar_en_segundo_plano(
            "Operacion activa",
            "Consultando operacion activa...",
            tarea_segura,
            finalizar,
        )

    def actualizar_operacion_activa_label(self):
        if not hasattr(self, "operacion_activa_label"):
            return
        try:
            if not self.operacion_activa_label.winfo_exists():
                return
        except Exception:
            return

        if not self.operacion_activa:
            self.operacion_activa_label.configure(
                text="Operacion activa:\nNo hay operacion abierta.",
                fg=self.colors["danger"],
            )
            return

        fecha_larga = self.fecha_larga_es(self.operacion_activa.get("fecha_inicio"))
        self.operacion_activa_label.configure(
            text=(
                "Operacion activa:\n"
                f"Buque: {self.operacion_activa.get('nombre_buque', '')}\n"
                f"Inicio: {fecha_larga}\n"
                f"Producto: {self.operacion_activa.get('producto') or ''}\n"
                f"Estado: {self.operacion_activa.get('estado') or ''}"
            ),
            fg=self.colors["text_dark"],
        )

    def normalizar_valores_cuota_tree(self, valores):
        valores = list(valores or [])
        if len(valores) >= 6:
            return valores[0], valores[1], valores[2], valores[3], valores[4], valores[5]
        if len(valores) >= 5:
            return valores[0], valores[1], valores[2], "Todas", valores[3], valores[4]
        if len(valores) >= 4:
            return valores[0], valores[1], "", "Todas", valores[2], valores[3]
        return "", "", "", "Todas", "", "MT"

    def bodega_cuota_visible(self, bodega_numero):
        bodegas = self.bodega_cuota_payload(bodega_numero)
        if not bodegas:
            return "Todas"
        return ", ".join(f"Bodega {numero}" for numero in bodegas)

    def bodega_cuota_payload(self, valor=None):
        valor = self.cuota_form_vars["bodega"].get() if valor is None else valor
        if isinstance(valor, (list, tuple, set)):
            candidatos = valor
        else:
            texto = str(valor or "").strip()
            if not texto or texto.lower() == "todas":
                return []
            texto = texto.replace("Bodega", "").replace("bodega", "").replace("B", "").replace("b", "").replace(";", ",")
            candidatos = texto.split(",")
        bodegas = []
        for candidato in candidatos:
            numero = self.safe_int(str(candidato).strip(), None)
            if numero and 1 <= numero <= 5 and numero not in bodegas:
                bodegas.append(numero)
        return sorted(bodegas)

    def bodega_cuota_payload_api(self, bodegas):
        bodegas = self.bodega_cuota_payload(bodegas)
        return {
            "bodega_numero": bodegas[0] if len(bodegas) == 1 else None,
            "bodega_numeros": bodegas,
        }

    def abrir_selector_bodegas_cuota(self, variable):
        popup = tk.Toplevel(self)
        popup.title("Seleccionar bodegas")
        popup.configure(bg=self.colors["bg_card"])
        popup.transient(self)
        popup.grab_set()
        popup.resizable(False, False)

        actuales = set(self.bodega_cuota_payload(variable.get()))
        todas_var = tk.BooleanVar(value=not actuales)
        bodega_vars = {numero: tk.BooleanVar(value=numero in actuales) for numero in range(1, 6)}

        contenido = tk.Frame(popup, bg=self.colors["bg_card"])
        contenido.pack(fill="both", expand=True, padx=16, pady=14)
        tk.Label(contenido, text="Bodegas de la cuota", font=("Segoe UI", 11, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w", pady=(0, 8))

        def alternar_todas():
            if todas_var.get():
                for var in bodega_vars.values():
                    var.set(False)

        def alternar_bodega():
            if any(var.get() for var in bodega_vars.values()):
                todas_var.set(False)
            else:
                todas_var.set(True)

        ttk.Checkbutton(contenido, text="Todas", variable=todas_var, command=alternar_todas).pack(anchor="w", pady=2)
        for numero, var in bodega_vars.items():
            ttk.Checkbutton(contenido, text=f"Bodega {numero}", variable=var, command=alternar_bodega).pack(anchor="w", pady=2)

        acciones = tk.Frame(contenido, bg=self.colors["bg_card"])
        acciones.pack(fill="x", pady=(12, 0))

        def aplicar():
            seleccion = [numero for numero, var in bodega_vars.items() if var.get()]
            variable.set(self.bodega_cuota_visible(seleccion))
            popup.destroy()

        ttk.Button(acciones, text="Aplicar", style="Olive.TButton", command=aplicar).pack(side="left", padx=(0, 8))
        ttk.Button(acciones, text="Cancelar", style="Gray.TButton", command=popup.destroy).pack(side="left")

    def producto_cuota_visible(self, producto):
        valor = str(producto or "").strip()
        if valor.upper() in ("", "NONE", "NULL", "SIN PRODUCTO", "TODOS", "ALL"):
            return "Cuota general"
        return valor

    def obtener_datos_cuota_form(self):
        cliente = self.cuota_form_vars["cliente"].get().strip()
        producto = self.cuota_form_vars["producto"].get().strip()
        bodega_numero = self.bodega_cuota_payload()
        cuota = self.cuota_form_vars["cuota"].get().strip()
        unidad = self.cuota_form_vars["unidad"].get().strip() or "MT"

        if not cliente:
            messagebox.showwarning("Dato requerido", "Debe indicar el cliente.")
            return None

        if not producto or producto.strip().upper() in ("CUOTA GENERAL", "SIN PRODUCTO", "TODOS", "ALL"):
            messagebox.showwarning("Dato requerido", "Debe seleccionar un producto real de la cuota.")
            return None

        if not cuota:
            messagebox.showwarning("Dato requerido", "Debe indicar la cuota.")
            return None

        cuota_num = self.safe_number(cuota, None)
        if cuota_num is None or cuota_num <= 0:
            messagebox.showwarning("Cuota invalida", "La cuota debe ser mayor a cero.")
            return None

        return cliente, producto, bodega_numero, cuota_num, unidad

    def render_lineas_cuota_front(self):
        if not hasattr(self, "cuota_lineas_frame"):
            return

        for widget in self.cuota_lineas_frame.winfo_children():
            widget.destroy()

        self.cuota_producto_combos = []
        productos = self.productos_operacion_actuales()
        for item in self.cuota_lineas_vars:
            row = tk.Frame(self.cuota_lineas_frame, bg=self.colors["bg_card"])
            row.pack(fill="x", pady=3)
            ttk.Entry(row, textvariable=item["cliente"]).grid(row=0, column=0, sticky="ew", padx=5)
            producto_combo = ttk.Combobox(row, textvariable=item["producto"], values=productos, state="readonly")
            producto_combo.grid(row=0, column=1, sticky="ew", padx=5)
            self.cuota_producto_combos.append(producto_combo)
            bodega_row = tk.Frame(row, bg=self.colors["bg_card"])
            bodega_row.grid(row=0, column=2, sticky="ew", padx=5)
            ttk.Entry(bodega_row, textvariable=item["bodega"], state="readonly").pack(side="left", fill="x", expand=True)
            ttk.Button(
                bodega_row,
                text="...",
                style="Gray.TButton",
                command=lambda var=item["bodega"]: self.abrir_selector_bodegas_cuota(var),
                width=3,
            ).pack(side="left", padx=(3, 0))
            ttk.Entry(row, textvariable=item["cuota"]).grid(row=0, column=3, sticky="ew", padx=5)
            ttk.Combobox(row, textvariable=item["unidad"], values=["MT"], state="readonly").grid(row=0, column=4, sticky="ew", padx=5)
            for col in range(5):
                row.grid_columnconfigure(col, weight=1)

    def agregar_linea_cuota_front(self, cliente="", producto="", cuota="", unidad="MT", bodega="Todas"):
        if not hasattr(self, "cuota_lineas_vars"):
            self.cuota_lineas_vars = []

        producto = producto or self.producto_operacion_default()
        self.cuota_lineas_vars.append({
            "cliente": tk.StringVar(value=cliente),
            "producto": tk.StringVar(value=producto),
            "bodega": tk.StringVar(value=bodega or "Todas"),
            "cuota": tk.StringVar(value=cuota),
            "unidad": tk.StringVar(value=unidad or "MT"),
        })
        self.render_lineas_cuota_front()

    def agregar_cuota_local_front(self):
        cliente = self.cuota_form_vars["cliente"].get().strip()
        producto = self.cuota_form_vars["producto"].get().strip()
        bodega = self.cuota_form_vars["bodega"].get().strip() or "Todas"
        cuota = self.cuota_form_vars["cuota"].get().strip()
        unidad = self.cuota_form_vars["unidad"].get().strip() or "MT"
        operacion_id = self.obtener_operacion_cuota_id()

        if getattr(self, "cuota_editing_item", None):
            datos = self.obtener_datos_cuota_form()
            if not datos:
                return
            cliente, producto, bodega_numero, cuota_num, unidad = datos
            item = self.cuota_editing_item
            if self.cuotas_tree is not None and item in self.cuotas_tree.get_children():
                valores = self.cuotas_tree.item(item, "values")
                cuota_id = valores[0] if valores else "PENDIENTE"
                cuota_id_num = self.safe_int(cuota_id, None)
                if operacion_id and cuota_id_num:
                    try:
                        payload_bodega = self.bodega_cuota_payload_api(bodega_numero)
                        self.api_actualizar_cuota_buque(cuota_id_num, {
                            "cliente": cliente,
                            "producto": producto,
                            **payload_bodega,
                            "cuota": cuota_num,
                            "unidad": unidad,
                            "observaciones": None,
                        })
                        self.cargar_cuotas_tabla_operacion_id(operacion_id)
                        self.cuota_editing_item = None
                        self.limpiar_form_cuota()
                        messagebox.showinfo("Cuota actualizada", "La cuota fue actualizada correctamente.")
                        return
                    except Exception as exc:
                        messagebox.showerror("Error actualizando cuota", str(exc))
                        return
                self.cuotas_tree.item(item, values=(cuota_id or "PENDIENTE", cliente, producto, self.bodega_cuota_visible(bodega_numero), self.formatear_numero(cuota_num, 2), unidad))
                self.cuotas_tree.selection_set(item)
                self.cuotas_tree.focus(item)
                self.cuotas_tree.see(item)
                self.cuota_editing_item = None
                self.limpiar_form_cuota()
                return

        if cliente or cuota:
            datos = self.obtener_datos_cuota_form()
            if not datos:
                return
            cliente, producto, bodega_numero, cuota_num, unidad = datos
            if self.cuotas_tree is not None:
                self.cuotas_tree.insert(
                    "",
                    "end",
                    values=("PENDIENTE", cliente, producto, self.bodega_cuota_visible(bodega_numero), self.formatear_numero(cuota_num, 2), unidad),
                )
            else:
                self.agregar_linea_cuota_front(cliente, producto, self.formatear_numero(cuota_num, 2), unidad, self.bodega_cuota_visible(bodega_numero))
            self.limpiar_form_cuota()
            return

        self.agregar_linea_cuota_front(cliente, producto, cuota, unidad, bodega)
        self.limpiar_form_cuota()

    def guardar_cuotas_lote_front(self):
        operacion_id = self.obtener_operacion_cuota_id()
        if not operacion_id:
            try:
                activa = self.api_get_operacion_activa()
                if activa:
                    operacion_id = self.safe_int(activa.get("id"), None)
            except Exception:
                operacion_id = None

        if not operacion_id:
            messagebox.showwarning("Sin operacion", "Seleccione una operacion en el combobox para crear las cuotas.")
            return

        filas = []
        if self.cuota_form_vars["cliente"].get().strip() or self.cuota_form_vars["cuota"].get().strip():
            datos = self.obtener_datos_cuota_form()
            if not datos:
                return
            filas.append(datos)

        for item in getattr(self, "cuota_lineas_vars", []):
            cliente = item["cliente"].get().strip()
            producto = item.get("producto", tk.StringVar(value="")).get().strip()
            bodega_numero = self.bodega_cuota_payload(item.get("bodega", tk.StringVar(value="Todas")).get())
            cuota = self.safe_number(item["cuota"].get().strip(), 0)
            unidad = item["unidad"].get().strip() or "MT"
            if cliente and producto and cuota > 0:
                filas.append((cliente, producto, bodega_numero, cuota, unidad))

        if not filas and self.cuotas_tree is not None:
            for item in self.cuotas_tree.get_children():
                valores = self.cuotas_tree.item(item, "values")
                cuota_id, cliente, producto, bodega, cuota_valor, unidad = self.normalizar_valores_cuota_tree(valores)
                if str(cuota_id).strip().upper() != "PENDIENTE":
                    continue
                cuota = self.safe_number(str(cuota_valor).replace(",", ""), 0)
                if cliente and producto and cuota > 0:
                    filas.append((cliente, producto, self.bodega_cuota_payload(bodega), cuota, unidad))

        if not filas:
            messagebox.showwarning("Sin cuotas", "Agregue al menos una cuota con el boton +.")
            return

        try:
            guardadas = []
            errores = []
            for cliente, producto, bodega_numero, cuota, unidad in filas:
                try:
                    payload_bodega = self.bodega_cuota_payload_api(bodega_numero)
                    respuesta = self.api_crear_cuota_buque({
                        "operacion_id": operacion_id,
                        "cliente": cliente,
                        "producto": producto,
                        **payload_bodega,
                        "cuota": cuota,
                        "unidad": unidad,
                        "observaciones": None,
                    })
                    cuota_db = respuesta.get("cuota", {})
                    guardadas.append({
                        "id": cuota_db.get("id", ""),
                        "cliente": cuota_db.get("cliente", cliente),
                        "producto": cuota_db.get("producto", producto),
                        "bodega_numeros": cuota_db.get("bodega_numeros", bodega_numero),
                        "cuota": cuota_db.get("cuota", cuota),
                        "unidad": cuota_db.get("unidad", unidad),
                    })
                except Exception as exc:
                    errores.append(f"{cliente}: {exc}")

            if errores:
                messagebox.showerror("Error creando cuotas", "\n".join(errores[:8]))
                return

            try:
                cuotas_db = self.cargar_cuotas_tabla_operacion_id(operacion_id)
            except Exception as exc:
                cuotas_db = []
                messagebox.showwarning("Cuotas creadas", f"Se crearon cuotas, pero no se pudo refrescar la tabla:\n{exc}")

            self.cuota_lineas_vars = []
            self.render_lineas_cuota_front()
            self.limpiar_form_cuota()
            alerta = self.alerta_cobertura_cuotas(operacion_id, cuotas_db or guardadas)
            messagebox.showinfo(
                "Cuotas creadas",
                f"Se enviaron {len(guardadas)} cuotas.\n"
                f"Cuotas visibles en base para esta operacion: {len(cuotas_db)}.",
            )
            if alerta:
                messagebox.showwarning("Cobertura de cuotas", alerta)
        except Exception as e:
            messagebox.showerror("Error creando cuotas", str(e))

    def guardar_cuota_front(self):
        if self.cuotas_tree is None:
            return

        datos = self.obtener_datos_cuota_form()
        if not datos:
            return
        cliente, producto, bodega_numero, cuota_num, unidad = datos

        # En apertura de operacion la cuota vive localmente hasta presionar Abrir operacion.
        if self.operaciones_tree is None:
            operacion_id = self.obtener_operacion_cuota_id()
            if operacion_id:
                payload_bodega = self.bodega_cuota_payload_api(bodega_numero)
                payload = {
                    "operacion_id": operacion_id,
                    "cliente": cliente,
                    "producto": producto,
                    **payload_bodega,
                    "cuota": cuota_num,
                    "unidad": unidad,
                    "observaciones": None,
                }
                try:
                    data = self.api_crear_cuota_buque(payload)
                    cuota_id = data.get("cuota", {}).get("id", "")
                    for item in self.cuotas_tree.get_children():
                        valores = self.cuotas_tree.item(item, "values")
                        _id, cliente_actual, producto_actual, bodega_actual, _cuota, _unidad = self.normalizar_valores_cuota_tree(valores)
                        if cliente_actual.strip().lower() == cliente.lower() and producto_actual.strip().lower() == producto.lower() and self.bodega_cuota_payload(bodega_actual) == bodega_numero:
                            self.cuotas_tree.item(item, values=(cuota_id, cliente, producto, self.bodega_cuota_visible(bodega_numero), self.formatear_numero(cuota_num, 2), unidad))
                            self.limpiar_form_cuota()
                            return
                    self.cuotas_tree.insert("", "end", values=(cuota_id, cliente, producto, self.bodega_cuota_visible(bodega_numero), self.formatear_numero(cuota_num, 2), unidad))
                    self.limpiar_form_cuota()
                    return
                except Exception as e:
                    messagebox.showerror("Error guardando cuota", str(e))
                    return

            if not self.operacion_activa:
                try:
                    self.operacion_activa = self.api_get_operacion_activa()
                    self.actualizar_operacion_activa_label()
                except Exception:
                    self.operacion_activa = None

            if self.operacion_activa:
                payload_bodega = self.bodega_cuota_payload_api(bodega_numero)
                payload = {
                    "operacion_id": self.operacion_activa.get("id"),
                    "cliente": cliente,
                    "producto": producto,
                    **payload_bodega,
                    "cuota": cuota_num,
                    "unidad": unidad,
                    "observaciones": None,
                }
                try:
                    data = self.api_crear_cuota_buque(payload)
                    cuota_id = data.get("cuota", {}).get("id", "")
                    for item in self.cuotas_tree.get_children():
                        valores = self.cuotas_tree.item(item, "values")
                        _id, cliente_actual, producto_actual, bodega_actual, _cuota, _unidad = self.normalizar_valores_cuota_tree(valores)
                        if cliente_actual.strip().lower() == cliente.lower() and producto_actual.strip().lower() == producto.lower() and self.bodega_cuota_payload(bodega_actual) == bodega_numero:
                            self.cuotas_tree.item(item, values=(cuota_id, cliente, producto, self.bodega_cuota_visible(bodega_numero), self.formatear_numero(cuota_num, 2), unidad))
                            self.limpiar_form_cuota()
                            return
                    self.cuotas_tree.insert("", "end", values=(cuota_id, cliente, producto, self.bodega_cuota_visible(bodega_numero), self.formatear_numero(cuota_num, 2), unidad))
                    self.limpiar_form_cuota()
                    return
                except Exception as e:
                    messagebox.showerror("Error guardando cuota", str(e))
                    return

            for item in self.cuotas_tree.get_children():
                valores = self.cuotas_tree.item(item, "values")
                _id, cliente_actual, producto_actual, bodega_actual, _cuota, _unidad = self.normalizar_valores_cuota_tree(valores)
                if cliente_actual.strip().lower() == cliente.lower() and producto_actual.strip().lower() == producto.lower() and self.bodega_cuota_payload(bodega_actual) == bodega_numero:
                    self.cuotas_tree.item(item, values=("", cliente, producto, self.bodega_cuota_visible(bodega_numero), self.formatear_numero(cuota_num, 2), unidad))
                    self.limpiar_form_cuota()
                    return

            self.cuotas_tree.insert("", "end", values=("", cliente, producto, self.bodega_cuota_visible(bodega_numero), self.formatear_numero(cuota_num, 2), unidad))
            self.limpiar_form_cuota()
            return

        operacion_id = self.obtener_operacion_seleccionada_id()
        if not operacion_id:
            messagebox.showwarning("Sin operacion", "Seleccione una operacion para agregar la cuota.")
            return

        payload_bodega = self.bodega_cuota_payload_api(bodega_numero)
        payload = {
            "operacion_id": operacion_id,
            "cliente": cliente,
            "producto": producto,
            **payload_bodega,
            "cuota": cuota_num,
            "unidad": unidad,
            "observaciones": None,
        }

        try:
            self.api_crear_cuota_buque(payload)
            self.limpiar_form_cuota()
            self.cargar_cuotas_operacion_seleccionada(silencioso=True)
            messagebox.showinfo("Cuota guardada", "La cuota fue guardada correctamente.")
        except Exception as e:
            messagebox.showerror("Error guardando cuota", str(e))

    def eliminar_cuota_seleccionada(self):
        if self.cuotas_tree is None:
            return

        lineas_pendientes = getattr(self, "cuota_lineas_vars", [])
        if lineas_pendientes:
            self.cuota_lineas_vars = lineas_pendientes[:-1]
            self.render_lineas_cuota_front()
            return

        seleccion = self.cuotas_tree.selection()
        if not seleccion:
            messagebox.showwarning("Sin seleccion", "Seleccione una cuota.")
            return

        valores = self.cuotas_tree.item(seleccion[0], "values")
        cuota_id = self.safe_int(valores[0], None) if valores else None

        if self.operaciones_tree is None:
            if cuota_id:
                if not messagebox.askyesno("Eliminar cuota", "Desea eliminar esta cuota de la operacion activa?"):
                    return
                try:
                    self.api_eliminar_cuota_buque(cuota_id)
                    self.cuotas_tree.delete(seleccion[0])
                    return
                except Exception as e:
                    messagebox.showerror("Error eliminando cuota", str(e))
                    return
            self.cuotas_tree.delete(seleccion[0])
            return

        if not cuota_id:
            self.cuotas_tree.delete(seleccion[0])
            return

        if not messagebox.askyesno("Eliminar cuota", "Â¿Desea eliminar esta cuota?"):
            return

        try:
            self.api_eliminar_cuota_buque(cuota_id)
            self.cargar_cuotas_operacion_seleccionada(silencioso=True)
            messagebox.showinfo("Cuota eliminada", "La cuota fue eliminada correctamente.")
        except Exception as e:
            messagebox.showerror("Error eliminando cuota", str(e))

    def crear_operacion_buque_front(self):
        nombre_buque = self.operacion_form_vars["nombre_buque"].get().strip()
        fecha_inicio = self.operacion_form_vars["fecha_inicio"].get().strip()
        productos = [var.get().strip() for var in self.producto_vars if var.get().strip()]

        if not nombre_buque:
            messagebox.showwarning("Dato requerido", "Debe indicar el nombre del buque.")
            return

        if not fecha_inicio:
            messagebox.showwarning("Dato requerido", "Debe seleccionar la fecha de inicio.")
            return

        if not productos:
            messagebox.showwarning("Dato requerido", "Debe agregar al menos un producto.")
            return

        cliente_pendiente = self.cuota_form_vars["cliente"].get().strip()
        producto_pendiente = self.cuota_form_vars["producto"].get().strip()
        bodega_pendiente = self.cuota_form_vars["bodega"].get().strip() or "Todas"
        cuota_pendiente = self.cuota_form_vars["cuota"].get().strip()
        unidad_pendiente = self.cuota_form_vars["unidad"].get().strip() or "MT"

        if cliente_pendiente or cuota_pendiente:
            if not cliente_pendiente:
                messagebox.showwarning("Dato requerido", "Debe indicar el cliente de la cuota.")
                return

            if not cuota_pendiente:
                messagebox.showwarning("Dato requerido", "Debe indicar la cuota del cliente.")
                return

            if not producto_pendiente:
                messagebox.showwarning("Dato requerido", "Debe seleccionar el producto de la cuota.")
                return

            cuota_num = self.safe_number(cuota_pendiente, None)
            if cuota_num is None or cuota_num <= 0:
                messagebox.showwarning("Cuota invalida", "La cuota debe ser mayor a cero.")
                return

            existe_cliente = False
            if self.cuotas_tree is not None:
                for item in self.cuotas_tree.get_children():
                    valores = self.cuotas_tree.item(item, "values")
                    _id, cliente_actual, producto_actual, bodega_actual, _cuota, _unidad = self.normalizar_valores_cuota_tree(valores)
                    if (
                        cliente_actual.strip().lower() == cliente_pendiente.lower()
                        and producto_actual.strip().lower() == producto_pendiente.lower()
                        and self.bodega_cuota_payload(bodega_actual) == self.bodega_cuota_payload(bodega_pendiente)
                    ):
                        self.cuotas_tree.item(item, values=("", cliente_pendiente, producto_pendiente, bodega_pendiente, self.formatear_numero(cuota_num, 2), unidad_pendiente))
                        existe_cliente = True
                        break

                if not existe_cliente:
                    self.cuotas_tree.insert(
                        "",
                        "end",
                        values=("", cliente_pendiente, producto_pendiente, bodega_pendiente, self.formatear_numero(cuota_num, 2), unidad_pendiente),
                    )

                self.limpiar_form_cuota()

        cuotas = []
        if self.cuotas_tree is not None:
            for item in self.cuotas_tree.get_children():
                valores = self.cuotas_tree.item(item, "values")
                cuota_id, cliente, producto, bodega, cuota_valor, unidad = self.normalizar_valores_cuota_tree(valores)
                cuota_num = self.safe_number(str(cuota_valor).replace(",", ""), 0)
                if cliente and producto and cuota_num > 0:
                    payload_bodega = self.bodega_cuota_payload_api(bodega)
                    cuotas.append({
                        "cliente": cliente,
                        "producto": producto,
                        **payload_bodega,
                        "cuota": cuota_num,
                        "unidad": unidad,
                    })

        bodegas = []
        for numero in range(1, 6):
            capacidad = self.safe_number(self.bodega_vars[numero].get(), 0)
            producto_bodega = ""
            if hasattr(self, "bodega_producto_vars") and numero in self.bodega_producto_vars:
                producto_bodega = self.bodega_producto_vars[numero].get().strip()
            producto_bodega = producto_bodega or self.producto_operacion_default()
            particiones = []
            for particion in self.bodega_particion_vars.get(numero, []):
                capacidad_part = self.safe_number(particion["capacidad_mt"].get(), 0)
                producto_part = particion.get("producto", tk.StringVar(value="")).get().strip()
                if capacidad_part > 0:
                    particiones.append({
                        "capacidad_mt": capacidad_part,
                        "producto": producto_part or producto_bodega,
                    })
            bodegas.append({
                "bodega_numero": numero,
                "capacidad_mt": capacidad,
                "producto": producto_bodega,
                "particiones": particiones,
            })

        if not any(item["capacidad_mt"] > 0 for item in bodegas):
            messagebox.showwarning("Dato requerido", "Debe indicar al menos una capacidad de bodega en MT.")
            return

        payload = {
            "nombre_buque": nombre_buque,
            "fecha_inicio": fecha_inicio,
            "producto": " / ".join(productos),
            "observaciones": None,
            "cerrar_operaciones_abiertas": True,
            "cuotas": cuotas,
            "bodegas": bodegas,
        }

        try:
            data = self.api_crear_operacion_buque(payload)
            operacion = data.get("operacion", {})
            self.limpiar_form_operacion()
            self.operacion_activa = operacion
            self.actualizar_operacion_activa_label()
            if hasattr(self, "cuota_operacion_combo") and operacion:
                etiqueta = (
                    f"{operacion.get('id')} | {operacion.get('nombre_buque', '')} | "
                    f"{self.fecha_larga_es(operacion.get('fecha_inicio'))} | {operacion.get('estado', '')}"
                )
                self.cuota_operaciones_map = {etiqueta: operacion}
                self.cuota_operacion_combo["values"] = [etiqueta]
                self.cuota_form_vars["operacion"].set(etiqueta)
            messagebox.showinfo("Operacion creada", f"Operacion abierta para {operacion.get('nombre_buque', nombre_buque)}")
        except Exception as e:
            messagebox.showerror("Error creando operacion", str(e))

    def show_buques(self):
        self.clear_content()
        self.highlight_sidebar_button("Historial de Buques")

        self.create_page_title(
            self.content,
            "Historial de Buques",
            "Consulte operaciones abiertas/cerradas, cierre o reabra operaciones y revise cuotas.",
        )

        scroll_host = tk.Frame(self.content, bg=self.colors["bg_main"])
        scroll_host.pack(fill="both", expand=True, padx=25, pady=(0, 20))

        canvas = tk.Canvas(scroll_host, bg=self.colors["bg_main"], highlightthickness=0)
        scroll_y = ttk.Scrollbar(scroll_host, orient="vertical", command=canvas.yview)
        scroll_x = ttk.Scrollbar(scroll_host, orient="horizontal", command=canvas.xview)
        body = tk.Frame(canvas, bg=self.colors["bg_main"])

        window_id = canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        self.bind_scroll_canvas(canvas, body, window_id, min_width=1040)

        canvas.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        scroll_host.grid_rowconfigure(0, weight=1)
        scroll_host.grid_columnconfigure(0, weight=1)

        panels = tk.Frame(body, bg=self.colors["bg_main"])
        panels.pack(fill="both", expand=True)

        ops_panel = tk.Frame(panels, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1)
        ops_panel.pack(side="left", fill="both", expand=True, padx=(0, 8))

        tk.Label(ops_panel, text="Operaciones", font=("Segoe UI", 13, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w", padx=14, pady=(12, 6))

        ops_actions = tk.Frame(ops_panel, bg=self.colors["bg_card"])
        ops_actions.pack(fill="x", padx=14, pady=(0, 8))
        ttk.Button(ops_actions, text="Cerrar", style="Gray.TButton", command=self.cerrar_operacion_seleccionada).pack(side="left", padx=(0, 8))
        ttk.Button(ops_actions, text="Reabrir", style="Olive.TButton", command=self.reabrir_operacion_seleccionada).pack(side="left", padx=(0, 8))
        ttk.Button(ops_actions, text="Ver cuotas", style="Gray.TButton", command=self.ver_cuotas_operacion_popup).pack(side="left", padx=(0, 8))
        ttk.Button(ops_actions, text="Buscar operaciones", style="Gray.TButton", command=self.refrescar_operaciones_buque).pack(side="left")

        ops_columns = ("id", "codigo", "buque", "inicio", "cierre", "producto", "estado")
        ops_table_frame = tk.Frame(ops_panel, bg=self.colors["bg_card"])
        ops_table_frame.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        self.operaciones_tree = ttk.Treeview(ops_table_frame, columns=ops_columns, show="headings", height=20)

        for col, heading, width in [
            ("id", "ID", 45), ("codigo", "Codigo", 180), ("buque", "Buque", 160),
            ("inicio", "Inicio", 130), ("cierre", "Cierre", 130), ("producto", "Producto", 180), ("estado", "Estado", 100),
        ]:
            self.operaciones_tree.heading(col, text=heading)
            self.operaciones_tree.column(col, width=width, anchor="center")

        ops_scroll_y = ttk.Scrollbar(ops_table_frame, orient="vertical", command=self.operaciones_tree.yview)
        ops_scroll_x = ttk.Scrollbar(ops_table_frame, orient="horizontal", command=self.operaciones_tree.xview)
        self.operaciones_tree.configure(yscrollcommand=ops_scroll_y.set, xscrollcommand=ops_scroll_x.set)
        self.operaciones_tree.grid(row=0, column=0, sticky="nsew")
        ops_scroll_y.grid(row=0, column=1, sticky="ns")
        ops_scroll_x.grid(row=1, column=0, sticky="ew")
        ops_table_frame.grid_rowconfigure(0, weight=1)
        ops_table_frame.grid_columnconfigure(0, weight=1)
        self.operaciones_tree.bind("<<TreeviewSelect>>", lambda _e: None)

        self.cuotas_tree = None
        self.operaciones_cache = []
        return
        cuotas_panel = tk.Frame(panels, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1, width=430)
        cuotas_panel.pack(side="right", fill="both")
        cuotas_panel.pack_propagate(False)
        tk.Label(cuotas_panel, text="Cuotas de la operacion", font=("Segoe UI", 13, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w", padx=14, pady=(12, 6))

        self.cuotas_tree = ttk.Treeview(cuotas_panel, columns=("id", "cliente", "cuota", "unidad"), show="headings", height=18)
        for col, heading, width in [("id", "ID", 45), ("cliente", "Cliente", 160), ("cuota", "Cuota", 110), ("unidad", "Unidad", 80)]:
            self.cuotas_tree.heading(col, text=heading)
            self.cuotas_tree.column(col, width=width, anchor="center")
        self.cuotas_tree.pack(fill="both", expand=True, padx=14, pady=(0, 14))

        self.operaciones_cache = []

    # =========================================================
    # APROBACIONES - FLUJO PENDING / APPROVED / REJECTED
    # =========================================================

    def api_cargar_aprobaciones_template(self):
        operacion = getattr(self, "operacion_activa", None) or self.asegurar_operacion_activa_boletas(mostrar_error=False) or {}
        operacion_id = self.safe_int(operacion.get("id"), None)
        if not operacion_id:
            raise RuntimeError("Busque una operacion activa antes de cargar aprobaciones extraordinarias.")
        template_path = self.obtener_template_boletas_trabajo_path(operacion_id)
        if template_path:
            with open(template_path, "rb") as archivo:
                contenido = archivo.read()
            usuario = (getattr(self, "current_user", {}) or {}).get("username") or os.getenv("USERNAME") or "desktop"
            respuesta_guardar = requests.post(
                f"{self.api_base}/aprobaciones/template-trabajo/guardar",
                params={"operacion_id": operacion_id, "usuario": usuario},
                data=contenido,
                headers={"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
                timeout=120,
            )
            if respuesta_guardar.status_code != 200:
                raise RuntimeError(self.obtener_detalle_error(respuesta_guardar))

        respuesta = requests.post(
            f"{self.api_base}/aprobaciones/cargar-template-trabajo",
            params={"operacion_id": operacion_id},
            timeout=120,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_aprobaciones_pendientes(self, params=None):
        respuesta = requests.get(
            f"{self.api_base}/aprobaciones/pendientes",
            params=params or {},
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_aprobaciones_filtros(self, params=None):
        respuesta = requests.get(
            f"{self.api_base}/aprobaciones/filtros",
            params=params or {},
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_accion_aprobaciones(self, accion, ids, comentario):
        endpoint = "aprobar" if accion == "APPROVED" else "rechazar"
        respuesta = requests.post(
            f"{self.api_base}/aprobaciones/{endpoint}",
            json={
                "ids": ids,
                "comentario": comentario,
            },
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_reasignar_aprobaciones(self, ids, chofer, placa, comentario):
        respuesta = requests.post(
            f"{self.api_base}/aprobaciones/reasignar",
            json={"ids": ids, "chofer": chofer, "placa": placa, "comentario": comentario},
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def show_aprobaciones(self):
        self.clear_content()
        self.highlight_sidebar_button("Aprobaciones")

        self.aprobaciones_tree = None
        self.aprobaciones_cache = []
        self.aprobaciones_checked_ids = set()
        self.aprobaciones_comentario_var = tk.StringVar()
        self.aprobaciones_reasignar_chofer_var = tk.StringVar()
        self.aprobaciones_reasignar_placa_var = tk.StringVar()
        self.aprobaciones_reasignar_guias_var = tk.StringVar()
        self.aprobaciones_accion_var = tk.StringVar(value="Aprobar")
        self.aprobaciones_filter_vars = {key: tk.StringVar() for key in ("guia", "empresa", "producto", "chofer", "placa")}
        self.aprobaciones_filter_widgets = {}

        self.create_page_title(
            self.content,
            "Aprobaciones",
            "Cargue el template de la operacion activa y apruebe o rechace registros pendientes.",
        )

        filtros_panel = tk.Frame(
            self.content,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        filtros_panel.pack(fill="x", padx=25, pady=(0, 10))
        tk.Label(
            filtros_panel,
            text="Filtros dinamicos",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=14, pady=(10, 4))
        filtros_grid = tk.Frame(filtros_panel, bg=self.colors["bg_card"])
        filtros_grid.pack(fill="x", padx=14, pady=(0, 8))
        for idx, (key, label) in enumerate([
            ("guia", "Guia"),
            ("empresa", "Empresa"),
            ("producto", "Producto"),
            ("chofer", "Chofer actual"),
            ("placa", "Placa"),
        ]):
            box = tk.Frame(filtros_grid, bg=self.colors["bg_card"])
            box.grid(row=0, column=idx, sticky="ew", padx=4)
            tk.Label(box, text=label, font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
            selector = self.crear_selector_filtrable_despacho(box, self.aprobaciones_filter_vars[key], width=18)
            selector.pack(fill="x")
            self.aprobaciones_filter_widgets[key] = selector
            filtros_grid.grid_columnconfigure(idx, weight=1)
        filtros_actions = tk.Frame(filtros_panel, bg=self.colors["bg_card"])
        filtros_actions.pack(fill="x", padx=14, pady=(0, 10))
        ttk.Button(filtros_actions, text="Cargar filtros", style="Gray.TButton", command=self.cargar_filtros_aprobaciones_front).pack(side="left", padx=(0, 8))
        ttk.Button(filtros_actions, text="Buscar pendientes", style="Olive.TButton", command=self.ver_aprobaciones_pendientes_front).pack(side="left", padx=(0, 8))
        ttk.Button(filtros_actions, text="Limpiar", style="Gray.TButton", command=self.limpiar_filtros_aprobaciones_front).pack(side="left")

        actions = tk.Frame(self.content, bg=self.colors["bg_main"])
        actions.pack(fill="x", padx=25, pady=(0, 10))

        ttk.Button(
            actions,
            text="Abrir Template",
            style="Gray.TButton",
            command=self.abrir_template_aprobaciones_remoto,
        ).pack(side="left", padx=(0, 8))

        ttk.Button(
            actions,
            text="Cargar Excel",
            style="Olive.TButton",
            command=self.cargar_template_aprobaciones_front,
        ).pack(side="left", padx=(0, 8))

        ttk.Button(
            actions,
            text="Ver datos cargados",
            style="Gray.TButton",
            command=self.ver_aprobaciones_pendientes_front,
        ).pack(side="left", padx=(0, 8))

        ttk.Combobox(
            actions,
            textvariable=self.aprobaciones_accion_var,
            values=["Aprobar", "Rechazar", "Seleccionar todo", "Desmarcar todo"],
            state="readonly",
            width=22,
        ).pack(side="left", padx=(0, 8), ipady=8)

        ttk.Button(
            actions,
            text="Aplicar accion",
            style="Olive.TButton",
            command=self.ejecutar_accion_aprobaciones,
        ).pack(side="left")

        comentario_panel = tk.Frame(
            self.content,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        comentario_panel.pack(fill="x", padx=25, pady=(0, 10))

        tk.Label(
            comentario_panel,
            text="Comentario para la accion",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=14, pady=(10, 3))

        ttk.Entry(
            comentario_panel,
            textvariable=self.aprobaciones_comentario_var,
        ).pack(fill="x", padx=14, pady=(0, 12))

        self.aprobaciones_table_host = tk.Frame(self.content, bg=self.colors["bg_main"])
        self.aprobaciones_table_host.pack(fill="both", expand=True, padx=25, pady=(0, 20))

        self.render_aprobaciones_placeholder()

    def render_aprobaciones_placeholder(self):
        for widget in self.aprobaciones_table_host.winfo_children():
            widget.destroy()

        panel = tk.Frame(
            self.aprobaciones_table_host,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        panel.pack(fill="x")

        tk.Label(
            panel,
            text="Los datos pendientes se muestran solo al presionar Ver datos cargados.",
            font=("Segoe UI", 11, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=14, pady=14)

    def cargar_template_aprobaciones_front(self):
        try:
            operacion_id = self.safe_int((getattr(self, "operacion_activa", {}) or {}).get("id"), None)
            if not self.obtener_template_boletas_trabajo_path(operacion_id):
                messagebox.showwarning(
                    "Template no preparado",
                    "Primero presione Abrir Template, complete el Excel y guardelo. Luego presione Cargar Excel.",
                )
                return
            data = self.api_cargar_aprobaciones_template()
            messagebox.showinfo(
                "Carga para aprobacion",
                f"{data.get('message', 'Template cargado.')}\n\n"
                f"Insertados: {data.get('insertados', 0)}\n"
                f"Omitidos: {data.get('omitidos', 0)}",
            )
        except Exception as e:
            messagebox.showerror("Error cargando aprobaciones", str(e))

    def obtener_params_aprobaciones(self):
        params = {}
        operacion_id = self.safe_int((getattr(self, "operacion_activa", {}) or {}).get("id"), None)
        if operacion_id:
            params["operacion_id"] = operacion_id
        for key, var in getattr(self, "aprobaciones_filter_vars", {}).items():
            value = var.get().strip()
            if value:
                params[key] = value
        return params

    def cargar_filtros_aprobaciones_front(self):
        try:
            try:
                data = self.api_get_aprobaciones_filtros(self.obtener_params_aprobaciones())
                opciones = data.get("opciones", {})
                if not self.opciones_tienen_datos(opciones):
                    raise RuntimeError("Filtros vacios")
            except Exception:
                data = self.api_get_aprobaciones_pendientes(self.obtener_params_aprobaciones())
                opciones = self.opciones_filtros_desde_filas(data.get("data", []))
            self.aplicar_opciones_filtros_aprobaciones(opciones)
        except Exception as e:
            messagebox.showerror("Error filtros aprobaciones", str(e))

    def aplicar_opciones_filtros_aprobaciones(self, opciones):
        mapa = {
            "guia": "guias",
            "empresa": "empresas",
            "producto": "productos",
            "chofer": "choferes",
            "placa": "placas",
        }
        for key, option_key in mapa.items():
            widget = getattr(self, "aprobaciones_filter_widgets", {}).get(key)
            var = getattr(self, "aprobaciones_filter_vars", {}).get(key)
            if widget is None or var is None:
                continue
            valores_limpios = [str(v) for v in opciones.get(option_key, []) if v not in (None, "")]
            valores = [""] + valores_limpios
            if hasattr(widget, "_xtravon_selector"):
                self.configurar_combo_filtrable_despacho(widget, valores_limpios, var)
            else:
                widget["values"] = valores
            if var.get() and var.get() not in valores:
                var.set("")

        choferes_limpios = [str(v) for v in opciones.get("choferes", []) if v not in (None, "")]
        placas_limpios = [str(v) for v in opciones.get("placas", []) if v not in (None, "")]
        choferes = [""] + choferes_limpios
        placas = [""] + placas_limpios
        if hasattr(self, "aprobaciones_reasignar_chofer_combo"):
            if hasattr(self.aprobaciones_reasignar_chofer_combo, "_xtravon_selector"):
                self.configurar_combo_filtrable_despacho(
                    self.aprobaciones_reasignar_chofer_combo,
                    choferes_limpios,
                    self.aprobaciones_reasignar_chofer_var,
                )
            else:
                self.aprobaciones_reasignar_chofer_combo["values"] = choferes
        if hasattr(self, "aprobaciones_reasignar_placa_combo"):
            if hasattr(self.aprobaciones_reasignar_placa_combo, "_xtravon_selector"):
                self.configurar_combo_filtrable_despacho(
                    self.aprobaciones_reasignar_placa_combo,
                    placas_limpios,
                    self.aprobaciones_reasignar_placa_var,
                )
            else:
                self.aprobaciones_reasignar_placa_combo["values"] = placas

    def limpiar_filtros_aprobaciones_front(self):
        for var in getattr(self, "aprobaciones_filter_vars", {}).values():
            var.set("")

    def ver_aprobaciones_pendientes_front(self):
        try:
            data = self.api_get_aprobaciones_pendientes(self.obtener_params_aprobaciones())
            self.aprobaciones_cache = data.get("data", [])
            self.aprobaciones_checked_ids = set()
            self.aplicar_opciones_filtros_aprobaciones(self.opciones_filtros_desde_filas(self.aprobaciones_cache))
            self.render_aprobaciones_tabla(self.aprobaciones_cache)
        except Exception as e:
            messagebox.showerror("Error consultando pendientes", str(e))

    def render_aprobaciones_tabla(self, datos):
        for widget in self.aprobaciones_table_host.winfo_children():
            widget.destroy()

        panel = tk.Frame(
            self.aprobaciones_table_host,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        panel.pack(fill="both", expand=True)

        tk.Label(
            panel,
            text=f"Registros PENDING: {len(datos)}",
            font=("Segoe UI", 13, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=14, pady=(12, 6))

        columns = (
            "check", "id", "guia", "empresa", "buque", "fecha", "producto",
            "chofer", "placa", "status"
        )

        table_frame = tk.Frame(panel, bg=self.colors["bg_card"])
        table_frame.pack(fill="both", expand=True, padx=14, pady=(0, 14))

        self.aprobaciones_tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            height=20,
            selectmode="extended",
        )

        headings = {
            "check": "Sel.",
            "id": "ID",
            "guia": "Guia",
            "empresa": "Empresa",
            "buque": "Buque",
            "fecha": "Fecha",
            "producto": "Producto",
            "chofer": "Chofer",
            "placa": "Placa",
            "status": "Status",
        }
        widths = {
            "check": 55,
            "id": 55,
            "guia": 90,
            "empresa": 140,
            "buque": 140,
            "fecha": 110,
            "producto": 140,
            "chofer": 150,
            "placa": 90,
            "status": 100,
        }

        for col in columns:
            self.aprobaciones_tree.heading(col, text=headings[col])
            self.aprobaciones_tree.column(col, width=widths[col], anchor="center")

        scroll_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.aprobaciones_tree.yview)
        scroll_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.aprobaciones_tree.xview)
        self.aprobaciones_tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        self.aprobaciones_tree.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        self.aprobaciones_tree.bind("<Button-1>", self.toggle_check_aprobacion_event)

        for fila in datos:
            self.aprobaciones_tree.insert(
                "",
                "end",
                values=(
                    "[ ]",
                    fila.get("id", ""),
                    fila.get("guia", ""),
                    fila.get("empresa", ""),
                    fila.get("buque", ""),
                    self.formatear_fecha(fila.get("fecha")),
                    fila.get("producto", ""),
                    fila.get("chofer", ""),
                    fila.get("placa", ""),
                    fila.get("aprobacion_estado", "PENDING"),
                ),
            )

    def toggle_check_aprobacion_event(self, event):
        if self.aprobaciones_tree is None:
            return
        region = self.aprobaciones_tree.identify("region", event.x, event.y)
        column = self.aprobaciones_tree.identify_column(event.x)
        if region != "cell" or column != "#1":
            return
        item = self.aprobaciones_tree.identify_row(event.y)
        if item:
            self.toggle_check_aprobacion_item(item)
            return "break"

    def toggle_check_aprobacion_item(self, item):
        valores = list(self.aprobaciones_tree.item(item, "values") or [])
        if len(valores) < 2:
            return
        registro_id = self.safe_int(valores[1], None)
        if not registro_id:
            return
        if registro_id in self.aprobaciones_checked_ids:
            self.aprobaciones_checked_ids.remove(registro_id)
            valores[0] = "[ ]"
        else:
            self.aprobaciones_checked_ids.add(registro_id)
            valores[0] = "[x]"
        self.aprobaciones_tree.item(item, values=valores)

    def marcar_todas_aprobaciones(self):
        if self.aprobaciones_tree is None:
            messagebox.showwarning("Sin datos", "Primero presione Ver datos cargados.")
            return

        items = self.aprobaciones_tree.get_children()
        if not items:
            return
        for item in items:
            valores = list(self.aprobaciones_tree.item(item, "values") or [])
            if len(valores) >= 2:
                registro_id = self.safe_int(valores[1], None)
                if registro_id:
                    self.aprobaciones_checked_ids.add(registro_id)
                    valores[0] = "[x]"
                    self.aprobaciones_tree.item(item, values=valores)

    def desmarcar_todas_aprobaciones(self):
        if self.aprobaciones_tree is None:
            messagebox.showwarning("Sin datos", "Primero presione Ver datos cargados.")
            return
        self.aprobaciones_checked_ids = set()
        for item in self.aprobaciones_tree.get_children():
            valores = list(self.aprobaciones_tree.item(item, "values") or [])
            if valores:
                valores[0] = "[ ]"
                self.aprobaciones_tree.item(item, values=valores)

    def ejecutar_accion_aprobaciones(self):
        accion = self.aprobaciones_accion_var.get().strip()
        if accion == "Seleccionar todo":
            self.marcar_todas_aprobaciones()
            return
        if accion == "Desmarcar todo":
            self.desmarcar_todas_aprobaciones()
            return
        if accion == "Aprobar":
            self.aplicar_accion_aprobaciones_front("APPROVED")
            return
        if accion == "Rechazar":
            self.aplicar_accion_aprobaciones_front("REJECTED")
            return
        messagebox.showwarning("Accion invalida", "Seleccione una accion valida.")

    def obtener_ids_aprobaciones_seleccionadas(self):
        if self.aprobaciones_tree is None:
            return []

        ids = list(getattr(self, "aprobaciones_checked_ids", set()))
        if ids:
            return ids

        ids = []
        for item in self.aprobaciones_tree.selection():
            valores = self.aprobaciones_tree.item(item, "values")
            if valores and len(valores) >= 2:
                registro_id = self.safe_int(valores[1], None)
                if registro_id:
                    ids.append(registro_id)

        return ids

    def obtener_ids_aprobaciones_por_guias(self):
        guias_texto = getattr(self, "aprobaciones_reasignar_guias_var", tk.StringVar()).get()
        guias = {
            item.strip()
            for item in guias_texto.replace("\n", ",").replace(" ", ",").split(",")
            if item.strip()
        }
        if not guias:
            return []
        ids = []
        for fila in getattr(self, "aprobaciones_cache", []):
            if str(fila.get("guia", "")).strip() in guias:
                registro_id = self.safe_int(fila.get("id"), None)
                if registro_id:
                    ids.append(registro_id)
        return ids

    def obtener_ids_aprobaciones_filtradas(self):
        ids = []
        for fila in getattr(self, "aprobaciones_cache", []):
            registro_id = self.safe_int(fila.get("id"), None)
            if registro_id:
                ids.append(registro_id)
        return ids

    def aplicar_accion_aprobaciones_front(self, accion):
        ids = self.obtener_ids_aprobaciones_seleccionadas()
        if not ids:
            ids = self.obtener_ids_aprobaciones_filtradas()
            if not ids:
                messagebox.showwarning("Sin seleccion", "Seleccione uno o mas registros o presione Buscar pendientes con filtros.")
                return
            if not messagebox.askyesno(
                "Usar registros visibles",
                f"No hay filas marcadas. Desea aplicar la accion a los {len(ids)} registro(s) visibles/filtrados?",
            ):
                return

        comentario = self.aprobaciones_comentario_var.get().strip()
        if not comentario:
            messagebox.showwarning("Comentario requerido", "Debe indicar un comentario para la accion.")
            return

        texto_accion = "aprobar" if accion == "APPROVED" else "rechazar"
        if not messagebox.askyesno("Confirmar accion", f"Desea {texto_accion} {len(ids)} registro(s)?"):
            return

        try:
            self.api_accion_aprobaciones(accion, ids, comentario)
            self.aprobaciones_comentario_var.set("")
            self.ver_aprobaciones_pendientes_front()
            messagebox.showinfo("Accion aplicada", "La accion fue aplicada correctamente.")
        except Exception as e:
            messagebox.showerror("Error aplicando accion", str(e))

    def reasignar_aprobaciones_front(self, modo="seleccion"):
        ids = self.obtener_ids_aprobaciones_por_guias() or self.obtener_ids_aprobaciones_seleccionadas()
        if not ids and modo == "filtros":
            if not getattr(self, "aprobaciones_cache", []):
                self.ver_aprobaciones_pendientes_front()
            ids = self.obtener_ids_aprobaciones_filtradas()
        if not ids:
            messagebox.showwarning("Sin seleccion", "Seleccione boletas, escriba guias exactas o busque pendientes filtradas.")
            return

        chofer = self.aprobaciones_reasignar_chofer_var.get().strip()
        placa = self.aprobaciones_reasignar_placa_var.get().strip()
        comentario = self.aprobaciones_comentario_var.get().strip()
        if not chofer and not placa:
            messagebox.showwarning("Dato requerido", "Indique el nuevo chofer, placa o ambos.")
            return

        if not messagebox.askyesno("Confirmar reasignaciÃ³n", f"Se reasignarÃ¡n {len(ids)} boleta(s). Â¿Desea continuar?"):
            return

        try:
            data = self.api_reasignar_aprobaciones(ids, chofer, placa, comentario)
            self.aprobaciones_reasignar_chofer_var.set("")
            self.aprobaciones_reasignar_placa_var.set("")
            self.aprobaciones_reasignar_guias_var.set("")
            self.ver_aprobaciones_pendientes_front()
            messagebox.showinfo(
                "ReasignaciÃ³n aplicada",
                f"Boletas reasignadas: {len(ids)}\nHashes regenerados: {data.get('hashes_regenerados', 0)}",
            )
        except Exception as e:
            messagebox.showerror("Error reasignando", str(e))

    def show_reasignacion_boletas(self):
        self.clear_content()
        self.highlight_sidebar_button("Reasignacion")

        self.aprobaciones_tree = None
        self.aprobaciones_cache = []
        self.aprobaciones_checked_ids = set()
        self.aprobaciones_comentario_var = tk.StringVar()
        self.aprobaciones_reasignar_chofer_var = tk.StringVar()
        self.aprobaciones_reasignar_placa_var = tk.StringVar()
        self.aprobaciones_reasignar_guias_var = tk.StringVar()
        self.aprobaciones_filter_vars = {key: tk.StringVar() for key in ("guia", "empresa", "producto", "chofer", "placa")}
        self.aprobaciones_filter_widgets = {}

        self.create_page_title(
            self.content,
            "Reasignacion de Boletas",
            "Filtre, marque guias y reasigne chofer o placa sin mezclarlo con el flujo de aprobacion.",
        )

        filtros_panel = tk.Frame(
            self.content,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        filtros_panel.pack(fill="x", padx=25, pady=(0, 10))
        tk.Label(
            filtros_panel,
            text="Filtros dinamicos",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=14, pady=(10, 4))
        filtros_grid = tk.Frame(filtros_panel, bg=self.colors["bg_card"])
        filtros_grid.pack(fill="x", padx=14, pady=(0, 8))
        for idx, (key, label) in enumerate([
            ("guia", "Guia"),
            ("empresa", "Empresa"),
            ("producto", "Producto"),
            ("chofer", "Chofer actual"),
            ("placa", "Placa"),
        ]):
            box = tk.Frame(filtros_grid, bg=self.colors["bg_card"])
            box.grid(row=0, column=idx, sticky="ew", padx=4)
            tk.Label(box, text=label, font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
            selector = self.crear_selector_filtrable_despacho(box, self.aprobaciones_filter_vars[key], width=18)
            selector.pack(fill="x")
            self.aprobaciones_filter_widgets[key] = selector
            filtros_grid.grid_columnconfigure(idx, weight=1)
        filtros_actions = tk.Frame(filtros_panel, bg=self.colors["bg_card"])
        filtros_actions.pack(fill="x", padx=14, pady=(0, 10))
        ttk.Button(filtros_actions, text="Cargar filtros", style="Gray.TButton", command=self.cargar_filtros_aprobaciones_front).pack(side="left", padx=(0, 8))
        ttk.Button(filtros_actions, text="Buscar boletas", style="Olive.TButton", command=self.buscar_reasignacion_pendientes_front).pack(side="left", padx=(0, 8))
        ttk.Button(filtros_actions, text="Marcar visibles", style="Gray.TButton", command=self.marcar_todas_aprobaciones).pack(side="left", padx=(0, 8))
        ttk.Button(filtros_actions, text="Limpiar", style="Gray.TButton", command=self.limpiar_filtros_aprobaciones_front).pack(side="left")

        reasignar_panel = tk.Frame(
            self.content,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        reasignar_panel.pack(fill="x", padx=25, pady=(0, 10))
        tk.Label(
            reasignar_panel,
            text="Nueva asignacion",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=14, pady=(10, 3))
        fila_reasignar = tk.Frame(reasignar_panel, bg=self.colors["bg_card"])
        fila_reasignar.pack(fill="x", padx=14, pady=(0, 10))
        tk.Label(fila_reasignar, text="Guias exactas", bg=self.colors["bg_card"], fg=self.colors["text_dark"], font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0, 5))
        ttk.Entry(fila_reasignar, textvariable=self.aprobaciones_reasignar_guias_var, width=26).pack(side="left", padx=(0, 8))
        tk.Label(fila_reasignar, text="Nuevo chofer", bg=self.colors["bg_card"], fg=self.colors["text_dark"], font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0, 5))
        self.aprobaciones_reasignar_chofer_combo = self.crear_selector_filtrable_despacho(
            fila_reasignar,
            self.aprobaciones_reasignar_chofer_var,
            width=34,
        )
        self.aprobaciones_reasignar_chofer_combo.pack(side="left", fill="x", expand=True, padx=(0, 8))
        tk.Label(fila_reasignar, text="Placa", bg=self.colors["bg_card"], fg=self.colors["text_dark"], font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0, 5))
        self.aprobaciones_reasignar_placa_combo = self.crear_selector_filtrable_despacho(
            fila_reasignar,
            self.aprobaciones_reasignar_placa_var,
            width=18,
        )
        self.aprobaciones_reasignar_placa_combo.pack(side="left", padx=(0, 8))
        ttk.Button(
            fila_reasignar,
            text="Reasignar marcadas",
            style="Olive.TButton",
            command=lambda: self.reasignar_aprobaciones_front("seleccion"),
        ).pack(side="left", padx=(0, 8))
        ttk.Button(
            fila_reasignar,
            text="Reasignar filtradas",
            style="Gray.TButton",
            command=lambda: self.reasignar_aprobaciones_front("filtros"),
        ).pack(side="left")

        comentario_frame = tk.Frame(reasignar_panel, bg=self.colors["bg_card"])
        comentario_frame.pack(fill="x", padx=14, pady=(0, 12))
        tk.Label(comentario_frame, text="Comentario", bg=self.colors["bg_card"], fg=self.colors["text_dark"], font=("Segoe UI", 9, "bold")).pack(anchor="w")
        ttk.Entry(comentario_frame, textvariable=self.aprobaciones_comentario_var).pack(fill="x")

        self.aprobaciones_table_host = tk.Frame(self.content, bg=self.colors["bg_main"])
        self.aprobaciones_table_host.pack(fill="both", expand=True, padx=25, pady=(0, 20))
        self.render_reasignacion_placeholder()

    def render_reasignacion_placeholder(self):
        for widget in self.aprobaciones_table_host.winfo_children():
            widget.destroy()
        panel = tk.Frame(
            self.aprobaciones_table_host,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        panel.pack(fill="x")
        tk.Label(
            panel,
            text="Busque boletas para reasignar. Puede marcar filas, escribir guias exactas o aplicar a todo lo filtrado.",
            font=("Segoe UI", 11, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=14, pady=14)

    def buscar_reasignacion_pendientes_front(self):
        try:
            data = self.api_get_aprobaciones_pendientes(self.obtener_params_aprobaciones())
            self.aprobaciones_cache = data.get("data", [])
            self.aprobaciones_checked_ids = set()
            self.aplicar_opciones_filtros_aprobaciones(self.opciones_filtros_desde_filas(self.aprobaciones_cache))
            self.render_reasignacion_tabla(self.aprobaciones_cache)
        except Exception as e:
            messagebox.showerror("Error consultando boletas", str(e))

    def render_reasignacion_tabla(self, datos):
        for widget in self.aprobaciones_table_host.winfo_children():
            widget.destroy()

        panel = tk.Frame(
            self.aprobaciones_table_host,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        panel.pack(fill="both", expand=True)

        tk.Label(
            panel,
            text=f"Boletas para reasignar: {len(datos)}",
            font=("Segoe UI", 13, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=14, pady=(12, 6))

        columns = (
            "check", "id", "guia", "empresa", "buque", "fecha", "producto",
            "chofer", "placa", "status"
        )

        split_frame = tk.Frame(panel, bg=self.colors["bg_card"])
        split_frame.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        table_frame = tk.Frame(split_frame, bg=self.colors["bg_card"])
        table_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))

        self.aprobaciones_tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            height=20,
            selectmode="extended",
        )

        headings = {
            "check": "Sel.",
            "id": "ID",
            "guia": "Guia",
            "empresa": "Empresa",
            "buque": "Buque",
            "fecha": "Fecha",
            "producto": "Producto",
            "chofer": "Chofer",
            "placa": "Placa",
            "status": "Status",
        }
        widths = {
            "check": 55,
            "id": 55,
            "guia": 90,
            "empresa": 140,
            "buque": 140,
            "fecha": 110,
            "producto": 140,
            "chofer": 180,
            "placa": 90,
            "status": 100,
        }

        for col in columns:
            self.aprobaciones_tree.heading(col, text=headings[col])
            self.aprobaciones_tree.column(col, width=widths[col], anchor="center")

        scroll_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.aprobaciones_tree.yview)
        scroll_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.aprobaciones_tree.xview)
        self.aprobaciones_tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        self.aprobaciones_tree.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        self.aprobaciones_tree.bind("<Button-1>", self.toggle_check_aprobacion_event)

        for fila in datos:
            self.aprobaciones_tree.insert(
                "",
                "end",
                values=(
                    "[ ]",
                    fila.get("id", ""),
                    fila.get("guia", ""),
                    fila.get("empresa", ""),
                    fila.get("buque", ""),
                    self.formatear_fecha(fila.get("fecha")),
                    fila.get("producto", ""),
                    fila.get("chofer", ""),
                    fila.get("placa", ""),
                    fila.get("aprobacion_estado", "PENDING"),
                ),
            )

        chofer_panel = tk.Frame(split_frame, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1, width=280)
        chofer_panel.pack(side="right", fill="y")
        chofer_panel.pack_propagate(False)
        tk.Label(
            chofer_panel,
            text="Choferes detectados",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=10, pady=(10, 4))
        tk.Label(
            chofer_panel,
            text="Doble click para usar como nuevo chofer.",
            font=("Segoe UI", 8),
            bg=self.colors["bg_card"],
            fg=self.colors["muted"],
        ).pack(anchor="w", padx=10, pady=(0, 6))
        lista_frame = tk.Frame(chofer_panel, bg=self.colors["bg_card"])
        lista_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        chofer_list = tk.Listbox(lista_frame, bg=self.colors["bg_main"], fg=self.colors["text_dark"], highlightthickness=1)
        chofer_scroll = ttk.Scrollbar(lista_frame, orient="vertical", command=chofer_list.yview)
        chofer_list.configure(yscrollcommand=chofer_scroll.set)
        chofer_list.pack(side="left", fill="both", expand=True)
        chofer_scroll.pack(side="right", fill="y")
        for chofer in sorted({str(fila.get("chofer") or "").strip() for fila in datos if str(fila.get("chofer") or "").strip()}):
            chofer_list.insert("end", chofer)

        def usar_chofer(_event=None):
            seleccion = chofer_list.curselection()
            if seleccion:
                self.aprobaciones_reasignar_chofer_var.set(chofer_list.get(seleccion[0]))

        chofer_list.bind("<Double-Button-1>", usar_chofer)

    # =========================================================
    # DESPACHO DE VIAJES
    # =========================================================

    def show_despacho_viajes(self):
        self.clear_content()
        self.highlight_sidebar_button("Despacho de Viajes")

        self.despacho_resumen = None
        self.despacho_operacion_id = None
        self.despacho_chofer_var = tk.StringVar()
        self.despacho_placa_var = tk.StringVar()
        self.despacho_canal_var = tk.StringVar(value="WHATSAPP")
        self.despacho_destino_var = tk.StringVar()
        self.despacho_solicitud_id = None
        self.despacho_filtro_vars = {
            "guia": tk.StringVar(),
            "empresa": tk.StringVar(),
            "producto": tk.StringVar(),
            "chofer": tk.StringVar(),
            "placa": tk.StringVar(),
            "estado_asignacion": tk.StringVar(),
            "bodega_numero": tk.StringVar(),
        }
        self.despacho_filtro_widgets = {}
        self.despacho_propuesta = None
        self.despacho_trees = {}
        self.despacho_pendientes_marcados = set()

        self.create_page_title(
            self.content,
            "Despacho de Viajes",
            "El ERP propone nuevas guias, pero despacho confirma antes de habilitar QR.",
        )

        scroll_wrapper = tk.Frame(self.content, bg=self.colors["bg_main"])
        scroll_wrapper.pack(fill="both", expand=True, padx=25, pady=(0, 20))
        canvas = tk.Canvas(scroll_wrapper, bg=self.colors["bg_main"], highlightthickness=0)
        scroll_y = ttk.Scrollbar(scroll_wrapper, orient="vertical", command=canvas.yview)
        scroll_x = ttk.Scrollbar(scroll_wrapper, orient="horizontal", command=canvas.xview)
        body = tk.Frame(canvas, bg=self.colors["bg_main"])
        window_id = canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        scroll_wrapper.grid_rowconfigure(0, weight=1)
        scroll_wrapper.grid_columnconfigure(0, weight=1)
        self.bind_scroll_canvas(canvas, body, window_id, min_width=1280)

        controls = tk.Frame(
            body,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        controls.pack(fill="x", pady=(0, 10))

        top = tk.Frame(controls, bg=self.colors["bg_card"])
        top.pack(fill="x", padx=14, pady=(12, 8))
        ttk.Button(top, text="Buscar operacion activa", style="Olive.TButton", command=self.cargar_despacho_resumen).pack(side="left", padx=(0, 10))
        self.despacho_estado_label = tk.Label(
            top,
            text="Presione Buscar operacion activa para consultar.",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        )
        self.despacho_estado_label.pack(side="left", fill="x", expand=True)

        form = tk.Frame(controls, bg=self.colors["bg_card"])
        form.pack(fill="x", padx=14, pady=(0, 8))
        for col, (label, var, width, kind) in enumerate([
            ("Chofer", self.despacho_chofer_var, 34, "combo"),
            ("Placa", self.despacho_placa_var, 16, "combo"),
            ("Destino WhatsApp/correo", self.despacho_destino_var, 28, "entry"),
        ]):
            tk.Label(form, text=label, bg=self.colors["bg_card"], fg=self.colors["text_dark"], font=("Segoe UI", 9, "bold")).grid(row=0, column=col, sticky="w", padx=4)
            if kind == "combo":
                combo = self.crear_selector_filtrable_despacho(
                    form,
                    var,
                    width=width,
                    on_select=self.autocompletar_placa_despacho if label == "Chofer" else None,
                )
                combo.grid(row=1, column=col, sticky="ew", padx=4)
                if label == "Chofer":
                    self.despacho_chofer_combo = combo
                else:
                    self.despacho_placa_combo = combo
            else:
                ttk.Entry(form, textvariable=var, width=width).grid(row=1, column=col, sticky="ew", padx=4)
            form.grid_columnconfigure(col, weight=1)

        actions = tk.Frame(controls, bg=self.colors["bg_card"])
        actions.pack(fill="x", padx=14, pady=(0, 12))
        ttk.Combobox(actions, textvariable=self.despacho_canal_var, values=["WHATSAPP", "CORREO"], state="readonly", width=12).pack(side="left", padx=(0, 8), ipady=8)
        ttk.Button(actions, text="Asignar guia", style="Olive.TButton", command=self.asignar_siguiente_viaje).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Enviar QR seleccion", style="Olive.TButton", command=self.entregar_qr_despacho).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Liberar seleccion", style="Gray.TButton", command=self.liberar_guias_despacho).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Cancelar seleccion", style="Gray.TButton", command=self.cancelar_guias_despacho).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Bloquear chofer/placa", style="Gray.TButton", command=self.bloquear_chofer_placa_despacho).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Marcar pendientes", style="Gray.TButton", command=self.marcar_todos_pendientes_reasignacion_despacho).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Reasignar marcadas", style="Olive.TButton", command=self.reasignar_pendientes_marcadas_despacho).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Desmarcar", style="Gray.TButton", command=self.desmarcar_todos_pendientes_reasignacion_despacho).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Limpiar filtros", style="Gray.TButton", command=self.limpiar_filtros_despacho).pack(side="left", padx=(0, 8))

        self.despacho_body = tk.Frame(body, bg=self.colors["bg_main"])
        self.despacho_body.pack(fill="both", expand=True)
        self.render_despacho_placeholder()

    def render_despacho_placeholder(self):
        for widget in self.despacho_body.winfo_children():
            widget.destroy()
        panel = tk.Frame(
            self.despacho_body,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        panel.pack(fill="x")
        tk.Label(
            panel,
            text="Sin consulta cargada. Esta pantalla no llama al backend hasta que presione Buscar operacion activa.",
            font=("Segoe UI", 11, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=14, pady=14)

    def obtener_filtros_despacho(self):
        filtros = {}
        for key, var in self.despacho_filtro_vars.items():
            value = var.get().strip()
            if value:
                filtros[key] = self.safe_int(value, value) if key == "bodega_numero" else value
        return filtros

    def aplicar_opciones_filtros_despacho(self, opciones):
        mapa = {
            "guia": "guias",
            "empresa": "empresas",
            "producto": "productos",
            "chofer": "choferes",
            "placa": "placas",
            "estado_asignacion": "estados_asignacion",
            "bodega_numero": "bodegas",
        }
        for key, option_key in mapa.items():
            widget = self.despacho_filtro_widgets.get(key)
            if widget is not None:
                self.configurar_combo_filtrable_despacho(widget, opciones.get(option_key, []), self.despacho_filtro_vars[key])
        if hasattr(self, "despacho_chofer_combo"):
            self.configurar_combo_filtrable_despacho(self.despacho_chofer_combo, opciones.get("choferes", []), self.despacho_chofer_var)
        if hasattr(self, "despacho_placa_combo"):
            self.configurar_combo_filtrable_despacho(self.despacho_placa_combo, opciones.get("placas", []), self.despacho_placa_var)

    def configurar_combo_filtrable_despacho(self, combo, valores, var):
        opciones = sorted({str(v) for v in valores if v not in (None, "")}, key=str.lower)
        combo._xtravon_values = opciones
        if hasattr(combo, "_xtravon_selector"):
            self.actualizar_selector_filtrable_despacho(combo)
            return
        combo["values"] = [""] + opciones

        def filtrar(_event=None):
            texto = var.get().strip().lower()
            if not texto:
                filtrados = opciones
            else:
                filtrados = [valor for valor in opciones if texto in valor.lower()]
            combo["values"] = [""] + filtrados
            if _event is not None and getattr(_event, "keysym", "") not in ("Up", "Down", "Left", "Right", "Return", "Escape", "Tab"):
                combo.after(10, lambda: combo.event_generate("<Down>"))

        combo.bind("<KeyRelease>", filtrar)

    def crear_selector_filtrable_despacho(self, parent, var, width=18, on_select=None):
        selector = tk.Frame(parent, bg=self.colors["bg_card"])
        selector._xtravon_selector = True
        selector._xtravon_values = []
        selector._xtravon_var = var
        selector._xtravon_popup = None
        selector._xtravon_on_select = on_select

        entry = ttk.Entry(selector, textvariable=var, width=width)
        button = tk.Button(
            selector,
            text="v",
            width=2,
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
            activebackground=self.colors["accent"],
            activeforeground=self.colors["bg_main"],
            relief="solid",
            bd=1,
            command=lambda s=selector: self.abrir_selector_filtrable_despacho(s),
        )
        entry.pack(side="left", fill="x", expand=True)
        button.pack(side="right", fill="y")

        selector.entry = entry
        selector.button = button

        def filtrar_en_vivo(event=None):
            if event is not None and getattr(event, "keysym", "") in ("Up", "Down", "Left", "Right", "Return", "Escape", "Tab"):
                if getattr(event, "keysym", "") == "Escape":
                    self.cerrar_selector_filtrable_despacho(selector)
                return
            self.abrir_selector_filtrable_despacho(selector)
            self.actualizar_selector_filtrable_despacho(selector)

        entry.bind("<KeyRelease>", filtrar_en_vivo)
        entry.bind("<FocusIn>", lambda _event, s=selector: self.actualizar_selector_filtrable_despacho(s))
        entry.bind("<Return>", lambda _event, s=selector: self.cerrar_selector_filtrable_despacho(s))
        return selector

    def abrir_selector_filtrable_despacho(self, selector):
        if not selector.winfo_exists():
            return
        popup_actual = getattr(self, "_despacho_selector_popup", None)
        if popup_actual is not None:
            try:
                owner = getattr(popup_actual, "_xtravon_owner", None)
                if popup_actual.winfo_exists() and owner is not selector:
                    popup_actual.destroy()
            except Exception:
                pass

        popup = getattr(selector, "_xtravon_popup", None)
        if popup is None or not popup.winfo_exists():
            popup = tk.Toplevel(self)
            popup.overrideredirect(True)
            popup.configure(bg=self.colors["bg_card"])
            popup._xtravon_owner = selector
            selector._xtravon_popup = popup
            self._despacho_selector_popup = popup

            listbox = tk.Listbox(
                popup,
                height=8,
                bg=self.colors["bg_topbar"],
                fg=self.colors["text_dark"],
                selectbackground=self.colors["accent"],
                selectforeground=self.colors["bg_main"],
                activestyle="none",
            )
            scroll = ttk.Scrollbar(popup, orient="vertical", command=listbox.yview)
            listbox.configure(yscrollcommand=scroll.set)
            listbox.pack(side="left", fill="both", expand=True)
            scroll.pack(side="right", fill="y")
            popup.listbox = listbox

            def seleccionar(_event=None):
                seleccion = listbox.curselection()
                if not seleccion:
                    return
                valor = listbox.get(seleccion[0])
                selector._xtravon_var.set("" if valor == "(Todos)" else valor)
                self.cerrar_selector_filtrable_despacho(selector)
                callback = getattr(selector, "_xtravon_on_select", None)
                if callback is not None:
                    callback()

            listbox.bind("<ButtonRelease-1>", seleccionar)
            listbox.bind("<Return>", seleccionar)
            popup.bind("<Escape>", lambda _event, s=selector: self.cerrar_selector_filtrable_despacho(s))

            def cerrar_por_click_externo(event, s=selector):
                pop = getattr(s, "_xtravon_popup", None)
                if pop is None or not pop.winfo_exists():
                    return
                x, y = event.x_root, event.y_root
                dentro_popup = (
                    pop.winfo_rootx() <= x <= pop.winfo_rootx() + pop.winfo_width()
                    and pop.winfo_rooty() <= y <= pop.winfo_rooty() + pop.winfo_height()
                )
                dentro_selector = (
                    s.winfo_rootx() <= x <= s.winfo_rootx() + s.winfo_width()
                    and s.winfo_rooty() <= y <= s.winfo_rooty() + s.winfo_height()
                )
                if not dentro_popup and not dentro_selector:
                    self.cerrar_selector_filtrable_despacho(s)

            bind_id = self.bind("<Button-1>", cerrar_por_click_externo, add="+")
            popup._xtravon_bind_id = bind_id

        x = selector.winfo_rootx()
        y = selector.winfo_rooty() + selector.winfo_height()
        width = max(selector.winfo_width(), 180)
        popup.geometry(f"{width}x180+{x}+{y}")
        self.actualizar_selector_filtrable_despacho(selector)
        popup.lift()

    def actualizar_selector_filtrable_despacho(self, selector):
        popup = getattr(selector, "_xtravon_popup", None)
        if popup is None or not popup.winfo_exists() or not hasattr(popup, "listbox"):
            return
        texto = selector._xtravon_var.get().strip().lower()
        opciones = list(getattr(selector, "_xtravon_values", []) or [])
        filtrados = [valor for valor in opciones if not texto or texto in valor.lower()]
        popup.listbox.delete(0, "end")
        popup.listbox.insert("end", "(Todos)")
        for valor in filtrados[:80]:
            popup.listbox.insert("end", valor)

    def cerrar_selector_filtrable_despacho(self, selector):
        popup = getattr(selector, "_xtravon_popup", None)
        if popup is None:
            return
        try:
            bind_id = getattr(popup, "_xtravon_bind_id", None)
            if bind_id:
                self.unbind("<Button-1>", bind_id)
        except Exception:
            pass
        try:
            if popup.winfo_exists():
                popup.destroy()
        except Exception:
            pass
        selector._xtravon_popup = None
        if getattr(self, "_despacho_selector_popup", None) is popup:
            self._despacho_selector_popup = None

    def autocompletar_placa_despacho(self, _event=None):
        chofer = self.despacho_chofer_var.get().strip().lower()
        if not chofer or not self.despacho_resumen:
            return
        for grupo in ("choferes_disponibles", "guias_disponibles", "guias_asignadas", "en_puerto", "completadas", "solicitudes_pendientes"):
            for item in self.despacho_resumen.get(grupo, []) or []:
                candidato = str(item.get("chofer_asignado") or item.get("chofer") or "").strip().lower()
                if candidato == chofer:
                    placa = str(item.get("placa_asignada") or item.get("placa") or "").strip()
                    if placa:
                        self.despacho_placa_var.set(placa)
                        return

    def cargar_filtros_despacho(self):
        try:
            data = self.api_get_despacho_filtros(self.despacho_operacion_id)
            operacion = data.get("operacion") or {}
            self.despacho_operacion_id = operacion.get("id") or self.despacho_operacion_id
            if hasattr(self, "despacho_estado_label"):
                self.despacho_estado_label.configure(
                    text=f"Operacion: {operacion.get('codigo_operacion', '')} | {operacion.get('nombre_buque', '')} | {operacion.get('estado', '')}"
                )
            self.aplicar_opciones_filtros_despacho(data.get("opciones", {}))
        except Exception as e:
            messagebox.showerror("Filtros despacho", str(e))

    def limpiar_filtros_despacho(self):
        for var in self.despacho_filtro_vars.values():
            var.set("")
        if hasattr(self, "despacho_chofer_var"):
            self.despacho_chofer_var.set("")
        if hasattr(self, "despacho_placa_var"):
            self.despacho_placa_var.set("")
        for tree in getattr(self, "despacho_trees", {}).values():
            for var in (getattr(tree, "_xtravon_filters", {}) or {}).values():
                var.set("")
            self.filtrar_tabla_despacho(tree)
        if self.despacho_resumen:
            self.cargar_despacho_resumen()

    def cargar_despacho_resumen(self):
        filtros = self.obtener_filtros_despacho()

        def tarea():
            activa = self.api_get_operacion_activa()
            oid = self.safe_int((activa or {}).get("id"), None)
            if not oid:
                raise RuntimeError("No hay operacion abierta.")
            data = self.api_get_despacho_resumen(oid, filtros)
            return {
                "operacion_activa": activa,
                "resumen": self.filtrar_resumen_despacho_local(data, filtros),
            }

        def al_terminar(resultado):
            data = resultado.get("resumen", {}) if isinstance(resultado, dict) else {}
            self.despacho_resumen = data
            operacion = data.get("operacion") or {}
            self.despacho_operacion_id = operacion.get("id")
            self.despacho_estado_label.configure(
                text=f"Operacion: {operacion.get('codigo_operacion', '')} | {operacion.get('nombre_buque', '')} | {operacion.get('estado', '')}"
            )
            self.aplicar_opciones_filtros_despacho(self.extraer_opciones_despacho(data))
            self.render_despacho_resumen(data)

        self.ejecutar_en_segundo_plano(
            "Despacho de viajes",
            "Buscando operacion activa y cargando tablero de despacho...",
            tarea,
            al_terminar,
        )

    def filtrar_resumen_despacho_local(self, data, filtros):
        filtros = {key: value for key, value in (filtros or {}).items() if value not in (None, "")}
        if not filtros or not isinstance(data, dict):
            return data

        def texto(item, *keys):
            for key in keys:
                value = item.get(key)
                if value not in (None, ""):
                    return str(value)
            return ""

        def coincide(item):
            reglas = {
                "guia": texto(item, "guia"),
                "empresa": texto(item, "empresa", "empresa_previa"),
                "producto": texto(item, "producto", "producto_previo"),
                "chofer": texto(item, "chofer_asignado", "chofer"),
                "placa": texto(item, "placa_asignada", "placa"),
                "estado_asignacion": texto(item, "estado_asignacion", "estado"),
            }
            for key, base in reglas.items():
                filtro = str(filtros.get(key) or "").strip().lower()
                if filtro and filtro not in base.lower():
                    return False
            bodega = str(filtros.get("bodega_numero") or "").strip()
            if bodega and str(item.get("bodega_numero") or "").strip() != bodega:
                return False
            return True

        grupos = (
            "guias_disponibles",
            "pendientes_reasignacion",
            "guias_asignadas",
            "primer_escaneo",
            "segundo_escaneo",
            "en_puerto",
            "completadas",
            "solicitudes_pendientes",
        )
        filtrado = dict(data)
        for grupo in grupos:
            filtrado[grupo] = [item for item in data.get(grupo, []) or [] if coincide(item)]

        chofer_filtro = str(filtros.get("chofer") or "").strip().lower()
        placa_filtro = str(filtros.get("placa") or "").strip().lower()
        if chofer_filtro or placa_filtro:
            filtrado["choferes_disponibles"] = [
                item for item in data.get("choferes_disponibles", []) or []
                if (not chofer_filtro or chofer_filtro in str(item.get("chofer") or "").lower())
                and (not placa_filtro or placa_filtro in str(item.get("placa") or "").lower())
            ]
        return filtrado

    def extraer_opciones_despacho(self, data):
        opciones = {
            "guias": set(),
            "empresas": set(),
            "productos": set(),
            "choferes": set(),
            "placas": set(),
            "estados_asignacion": set(),
            "bodegas": set(),
        }
        for grupo in (
            "guias_disponibles",
            "pendientes_reasignacion",
            "guias_asignadas",
            "primer_escaneo",
            "segundo_escaneo",
            "en_puerto",
            "completadas",
            "solicitudes_pendientes",
            "choferes_disponibles",
        ):
            for item in data.get(grupo, []) or []:
                valores = {
                    "guias": item.get("guia"),
                    "empresas": item.get("empresa") or item.get("empresa_previa"),
                    "productos": item.get("producto") or item.get("producto_previo"),
                    "choferes": item.get("chofer_asignado") or item.get("chofer"),
                    "placas": item.get("placa_asignada") or item.get("placa"),
                    "estados_asignacion": item.get("estado_asignacion") or item.get("estado"),
                    "bodegas": item.get("bodega_numero"),
                }
                for key, value in valores.items():
                    if value not in (None, ""):
                        opciones[key].add(str(value))
        return {key: sorted(values) for key, values in opciones.items()}

    def render_despacho_resumen(self, data):
        for widget in self.despacho_body.winfo_children():
            widget.destroy()
        self.despacho_trees = {}
        ids_pendientes = {
            self.safe_int(item.get("id"), None)
            for item in (data.get("pendientes_reasignacion", []) or [])
        }
        ids_pendientes.discard(None)
        self.despacho_pendientes_marcados = {
            item_id
            for item_id in (getattr(self, "despacho_pendientes_marcados", set()) or set())
            if item_id in ids_pendientes
        }

        kpis = tk.Frame(self.despacho_body, bg=self.colors["bg_main"])
        kpis.pack(fill="x", pady=(0, 10))
        self.create_card(kpis, "Solicitudes", self.formatear_numero(len(data.get("solicitudes_pendientes", [])), 0), self.colors["danger"])
        self.create_card(kpis, "Asignadas", self.formatear_numero(len(data.get("guias_asignadas", [])), 0), self.colors["info"])
        self.create_card(kpis, "Ingreso puerto", self.formatear_numero(len(data.get("primer_escaneo", [])), 0), self.colors["warning"])
        self.create_card(kpis, "Ingreso tolva", self.formatear_numero(len(data.get("segundo_escaneo", [])), 0), self.colors["success"])
        self.create_card(kpis, "Completadas", self.formatear_numero(len(data.get("completadas", [])), 0), self.colors["muted"])
        self.create_card(kpis, "Choferes disponibles", self.formatear_numero(len(data.get("choferes_disponibles", [])), 0), self.colors["success"])
        self.create_card(kpis, "Bloqueos", self.formatear_numero(len(data.get("bloqueos", [])), 0), self.colors["danger"])
        plan = data.get("plan_viajes") or {}
        self.create_card(kpis, "Viajes estimados", self.formatear_numero(plan.get("viajes_estimados"), 0), self.colors["accent"])

        propuesta_panel = tk.Frame(
            self.despacho_body,
            bg=self.colors["bg_card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        propuesta_panel.pack(fill="x", pady=(0, 10))
        self.despacho_propuesta_label = tk.Label(
            propuesta_panel,
            text=plan.get("lectura") or "Indique chofer, placa, cliente y producto; luego presione Asignar guia.",
            font=("Segoe UI", 11, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
            wraplength=1100,
            justify="left",
        )
        self.despacho_propuesta_label.pack(anchor="w", padx=14, pady=12)

        alertas = data.get("alertas", [])
        if alertas:
            alertas_text = " | ".join([f"{item.get('nivel')}: {item.get('mensaje')}" for item in alertas])
            tk.Label(
                propuesta_panel,
                text=f"Alertas: {alertas_text}",
                font=("Segoe UI", 10, "bold"),
                bg=self.colors["bg_card"],
                fg=self.colors["danger"],
                wraplength=1100,
                justify="left",
            ).pack(anchor="w", padx=14, pady=(0, 12))

        grid = tk.Frame(self.despacho_body, bg=self.colors["bg_main"])
        grid.pack(fill="both", expand=True)
        secciones = [
            ("solicitudes_pendientes", "Solicitudes de nuevo viaje"),
            ("pendientes_reasignacion", "Pendientes de reasignar"),
            ("guias_asignadas", "Guias asignadas"),
            ("primer_escaneo", "Primer escaneo - ingreso puerto"),
            ("segundo_escaneo", "Segundo escaneo - ingreso tolva"),
            ("completadas", "Tercer escaneo / ciclo completado"),
            ("choferes_disponibles", "Choferes disponibles"),
            ("bloqueos", "Bloqueos activos"),
        ]
        for idx, (key, title) in enumerate(secciones):
            row = idx // 2
            col = idx % 2
            panel = tk.Frame(grid, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1)
            panel.grid(row=row, column=col, sticky="nsew", padx=6, pady=6)
            tk.Label(panel, text=title, font=("Segoe UI", 12, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w", padx=10, pady=(10, 4))
            if key == "choferes_disponibles":
                tree = self.crear_tabla_despacho(panel, ("chofer", "placa", "estado"), {"chofer": 220, "placa": 110, "estado": 120})
                for item in data.get(key, []):
                    tree.insert("", "end", values=(item.get("chofer", ""), item.get("placa", ""), item.get("estado", "")))
                self.inicializar_filtros_tabla_despacho(tree)
                self.configurar_menu_despacho(tree, key)
            elif key == "solicitudes_pendientes":
                tree = self.crear_tabla_despacho(panel, ("id", "chofer", "placa", "guia", "producto"), {"id": 55, "chofer": 180, "placa": 90, "guia": 90, "producto": 140})
                for item in data.get(key, []):
                    tree.insert("", "end", values=(item.get("id", ""), item.get("chofer", ""), item.get("placa", ""), item.get("guia", ""), item.get("producto_previo", "")))
                self.inicializar_filtros_tabla_despacho(tree)

                def usar_solicitud(_event=None, tree_ref=tree):
                    seleccion = tree_ref.selection()
                    if not seleccion:
                        return
                    valores = tree_ref.item(seleccion[0], "values")
                    if not valores:
                        return
                    self.despacho_solicitud_id = self.safe_int(valores[0], None)
                    self.despacho_chofer_var.set(valores[1] or "")
                    self.despacho_placa_var.set(valores[2] or "")

                tree.bind("<Double-Button-1>", usar_solicitud)
                self.configurar_menu_despacho(tree, key)
            elif key == "bloqueos":
                tree = self.crear_tabla_despacho(panel, ("id", "chofer", "placa", "motivo"), {"id": 50, "chofer": 180, "placa": 90, "motivo": 240})
                for item in data.get(key, []):
                    tree.insert("", "end", values=(item.get("id", ""), item.get("chofer", ""), item.get("placa", ""), item.get("motivo", "")))
                self.inicializar_filtros_tabla_despacho(tree)
                self.configurar_menu_despacho(tree, key)
            else:
                if key == "pendientes_reasignacion":
                    acciones_pendientes = tk.Frame(panel, bg=self.colors["bg_card"])
                    acciones_pendientes.pack(fill="x", padx=10, pady=(0, 6))
                    ttk.Button(
                        acciones_pendientes,
                        text="Marcar visibles",
                        style="Olive.TButton",
                        command=self.marcar_visibles_pendientes_reasignacion_despacho,
                    ).pack(side="left", padx=(0, 6))
                    ttk.Button(
                        acciones_pendientes,
                        text="Desmarcar visibles",
                        style="Gray.TButton",
                        command=self.desmarcar_visibles_pendientes_reasignacion_despacho,
                    ).pack(side="left", padx=(0, 6))
                    ttk.Button(
                        acciones_pendientes,
                        text="Reasignar marcadas",
                        style="Olive.TButton",
                        command=self.reasignar_pendientes_marcadas_despacho,
                    ).pack(side="left")
                    columnas = ("sel", "id", "guia", "empresa", "producto", "chofer", "placa", "estado_asignacion")
                    anchos = {"sel": 48, "id": 55, "guia": 90, "empresa": 140, "producto": 120, "chofer": 160, "placa": 90, "estado_asignacion": 120}
                else:
                    columnas = ("id", "guia", "empresa", "producto", "chofer", "placa", "estado_asignacion")
                    anchos = {"id": 55, "guia": 90, "empresa": 140, "producto": 120, "chofer": 160, "placa": 90, "estado_asignacion": 120}
                tree = self.crear_tabla_despacho(
                    panel,
                    columnas,
                    anchos,
                )
                for item in data.get(key, []):
                    guia_id = self.safe_int(item.get("id"), None)
                    valores_base = (
                        item.get("id", ""),
                        item.get("guia", ""),
                        item.get("empresa", ""),
                        item.get("producto", ""),
                        item.get("chofer_asignado") or item.get("chofer", ""),
                        item.get("placa_asignada") or item.get("placa", ""),
                        item.get("estado_asignacion", ""),
                    )
                    valores = (
                        ("[x]" if guia_id in self.despacho_pendientes_marcados else "[ ]"),
                        *valores_base,
                    ) if key == "pendientes_reasignacion" else valores_base
                    tree.insert(
                        "",
                        "end",
                        values=valores,
                    )
                self.inicializar_filtros_tabla_despacho(tree)
                self.configurar_menu_despacho(tree, key)
                if key == "pendientes_reasignacion":
                    tree.bind(
                        "<Button-1>",
                        lambda event, tree_ref=tree: self.toggle_pendiente_reasignacion_checkbox(event, tree_ref),
                        add="+",
                    )
                if key in ("pendientes_reasignacion", "guias_asignadas"):
                    tree.bind("<Double-Button-1>", lambda _event, tree_ref=tree: self.usar_guia_disponible_despacho(tree_ref))
            self.despacho_trees[key] = tree

        for col in range(2):
            grid.grid_columnconfigure(col, weight=1, uniform="despacho")
        for row in range(4):
            grid.grid_rowconfigure(row, weight=1, minsize=230)

    def crear_tabla_despacho(self, parent, columns, widths):
        frame = tk.Frame(parent, bg=self.colors["bg_card"])
        frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        tree_filters = {}
        for col in columns:
            if col != "sel":
                tree_filters[col] = tk.StringVar()

        tree = ttk.Treeview(frame, columns=columns, show="headings", height=8, selectmode="extended")
        tree._xtravon_columns = columns
        tree._xtravon_filters = tree_filters
        tree._xtravon_all_rows = []
        for col in columns:
            if col == "sel":
                tree.heading(col, text="Sel.")
                tree.column(col, width=widths.get(col, 100), anchor="center")
                continue
            tree.heading(
                col,
                text=col.replace("_", " ").title(),
                command=lambda c=col, tree_ref=tree: self.abrir_filtro_columna_despacho(tree_ref, c),
            )
            tree.column(col, width=widths.get(col, 100), anchor="center")
        scroll_y = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        scroll_x = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        tree.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)
        return tree

    def inicializar_filtros_tabla_despacho(self, tree):
        columns = list(getattr(tree, "_xtravon_columns", []) or [])
        all_rows = [tuple(tree.item(item, "values") or []) for item in tree.get_children()]
        tree._xtravon_all_rows = all_rows
        for idx, col in enumerate(columns):
            var = tree._xtravon_filters.get(col)
            if var is None:
                continue
            valores = sorted({str(row[idx]) for row in all_rows if idx < len(row) and row[idx] not in (None, "")}, key=str.lower)
            tree._xtravon_filter_options = getattr(tree, "_xtravon_filter_options", {})
            tree._xtravon_filter_options[col] = valores

    def abrir_filtro_columna_despacho(self, tree, col):
        filtros = getattr(tree, "_xtravon_filters", {}) or {}
        var = filtros.get(col)
        if var is None:
            return
        opciones = list((getattr(tree, "_xtravon_filter_options", {}) or {}).get(col, []))

        popup_actual = getattr(self, "_despacho_filter_popup", None)
        if popup_actual is not None:
            try:
                if popup_actual.winfo_exists():
                    popup_actual.destroy()
            except Exception:
                pass

        popup = tk.Toplevel(self)
        self._despacho_filter_popup = popup
        popup.title(f"Filtrar {col.replace('_', ' ').title()}")
        popup.configure(bg=self.colors["bg_card"])
        popup.transient(self)
        popup.resizable(False, False)
        popup.geometry("+%d+%d" % (self.winfo_pointerx(), self.winfo_pointery()))
        popup.bind("<Escape>", lambda _event: popup.destroy())

        tk.Label(
            popup,
            text=f"Filtrar {col.replace('_', ' ').title()}",
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
            font=("Segoe UI", 10, "bold"),
        ).pack(anchor="w", padx=10, pady=(10, 4))

        search_var = tk.StringVar(value=var.get())
        entry = ttk.Entry(popup, textvariable=search_var, width=32)
        entry.pack(fill="x", padx=10, pady=(0, 6))

        listbox = tk.Listbox(
            popup,
            height=8,
            selectmode="extended",
            bg=self.colors["bg_topbar"],
            fg=self.colors["text_dark"],
            selectbackground=self.colors["accent"],
            selectforeground=self.colors["bg_main"],
            activestyle="none",
        )
        listbox.pack(fill="both", expand=True, padx=10, pady=(0, 8))

        def poblar():
            texto = search_var.get().strip().lower()
            listbox.delete(0, "end")
            listbox.insert("end", "(Todos)")
            for valor in opciones:
                if not texto or texto in valor.lower():
                    listbox.insert("end", valor)

        def aplicar(valor=None):
            if valor is None:
                seleccion = listbox.curselection()
                if seleccion:
                    seleccionados = [listbox.get(idx) for idx in seleccion]
                    if "(Todos)" in seleccionados:
                        valor = "(Todos)"
                    else:
                        valor = " || ".join(seleccionados)
                else:
                    valor = search_var.get()
            var.set("" if valor == "(Todos)" else valor)
            self.filtrar_tabla_despacho(tree)
            popup.destroy()

        def limpiar():
            var.set("")
            self.filtrar_tabla_despacho(tree)
            popup.destroy()

        def cerrar_si_fuera(_event=None):
            def verificar():
                if not popup.winfo_exists():
                    return
                focus = popup.focus_get()
                if focus is None or not str(focus).startswith(str(popup)):
                    popup.destroy()
            popup.after(120, verificar)

        def limpiar_referencia_popup(_event=None):
            if getattr(self, "_despacho_filter_popup", None) is popup:
                self._despacho_filter_popup = None

        def on_key(_event=None):
            poblar()

        entry.bind("<KeyRelease>", on_key)
        entry.bind("<Return>", lambda _event: aplicar(search_var.get()))
        listbox.bind("<Double-Button-1>", lambda _event: aplicar())
        popup.bind("<FocusOut>", cerrar_si_fuera, add="+")
        popup.bind("<Deactivate>", cerrar_si_fuera, add="+")
        popup.bind("<Destroy>", limpiar_referencia_popup, add="+")

        actions = tk.Frame(popup, bg=self.colors["bg_card"])
        actions.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(actions, text="Aplicar", style="Olive.TButton", command=aplicar).pack(side="left", padx=(0, 6))
        ttk.Button(actions, text="Limpiar", style="Gray.TButton", command=limpiar).pack(side="left")

        poblar()
        entry.focus_set()
        entry.selection_range(0, "end")

    def filtrar_tabla_despacho(self, tree):
        columns = list(getattr(tree, "_xtravon_columns", []) or [])
        filtros = getattr(tree, "_xtravon_filters", {}) or {}
        criterios = {
            col: [parte.strip().lower() for parte in filtros[col].get().split("||") if parte.strip()]
            for col in columns
            if col != "sel" and filtros.get(col) is not None and filtros[col].get().strip()
        }
        for col in columns:
            if col == "sel":
                tree.heading(col, text="Sel.")
                continue
            base = col.replace("_", " ").title()
            activo = bool(filtros.get(col) is not None and filtros[col].get().strip())
            tree.heading(
                col,
                text=f"{base} {'*' if activo else 'v'}",
                command=lambda c=col, tree_ref=tree: self.abrir_filtro_columna_despacho(tree_ref, c),
            )
        for item in tree.get_children():
            tree.delete(item)
        for row in getattr(tree, "_xtravon_all_rows", []) or []:
            incluir = True
            for idx, col in enumerate(columns):
                if col == "sel":
                    continue
                criterio = criterios.get(col)
                valor_celda = str(row[idx]).lower() if idx < len(row) else ""
                if criterio and not any(parte in valor_celda for parte in criterio):
                    incluir = False
                    break
            if incluir:
                tree.insert("", "end", values=row)

    def indice_columna_despacho(self, tree, col):
        columns = list(getattr(tree, "_xtravon_columns", []) or [])
        try:
            return columns.index(col)
        except ValueError:
            return None

    def valor_columna_despacho(self, tree, valores, col, default=""):
        idx = self.indice_columna_despacho(tree, col)
        if idx is None or idx >= len(valores):
            return default
        return valores[idx]

    def id_fila_despacho(self, tree, valores):
        return self.safe_int(self.valor_columna_despacho(tree, valores, "id", None), None)

    def ids_seleccionados_tree_despacho(self, tree, incluir_marcados=False):
        ids = []
        if incluir_marcados:
            for item_id in sorted(getattr(self, "despacho_pendientes_marcados", set()) or set()):
                if item_id not in ids:
                    ids.append(item_id)
        for item in tree.selection():
            gid = self.id_fila_despacho(tree, tree.item(item, "values") or [])
            if gid is not None and gid not in ids:
                ids.append(gid)
        return ids

    def actualizar_fila_marcada_pendiente_reasignacion(self, tree, guia_id, marcado):
        idx_sel = self.indice_columna_despacho(tree, "sel")
        if idx_sel is None:
            return
        marca = "[x]" if marcado else "[ ]"
        for item in tree.get_children():
            vals = list(tree.item(item, "values") or [])
            if self.id_fila_despacho(tree, vals) == guia_id and idx_sel < len(vals):
                vals[idx_sel] = marca
                tree.item(item, values=tuple(vals))
        nuevas = []
        for row in getattr(tree, "_xtravon_all_rows", []) or []:
            vals = list(row)
            if self.id_fila_despacho(tree, vals) == guia_id and idx_sel < len(vals):
                vals[idx_sel] = marca
            nuevas.append(tuple(vals))
        tree._xtravon_all_rows = nuevas

    def toggle_pendiente_reasignacion_checkbox(self, event, tree):
        if tree.identify("region", event.x, event.y) != "cell":
            return
        column_id = tree.identify_column(event.x)
        try:
            idx = int(column_id.replace("#", "")) - 1
        except ValueError:
            return
        columns = list(getattr(tree, "_xtravon_columns", []) or [])
        if idx >= len(columns) or columns[idx] != "sel":
            return
        item = tree.identify_row(event.y)
        if not item:
            return "break"
        gid = self.id_fila_despacho(tree, tree.item(item, "values") or [])
        if gid is None:
            return "break"
        marcados = set(getattr(self, "despacho_pendientes_marcados", set()) or set())
        if gid in marcados:
            marcados.remove(gid)
        else:
            marcados.add(gid)
        self.despacho_pendientes_marcados = marcados
        self.actualizar_fila_marcada_pendiente_reasignacion(tree, gid, gid in marcados)
        return "break"

    def marcar_visibles_pendientes_reasignacion_despacho(self):
        tree = (getattr(self, "despacho_trees", {}) or {}).get("pendientes_reasignacion")
        if tree is None:
            return
        marcados = set(getattr(self, "despacho_pendientes_marcados", set()) or set())
        for item in tree.get_children():
            gid = self.id_fila_despacho(tree, tree.item(item, "values") or [])
            if gid is not None:
                marcados.add(gid)
                self.actualizar_fila_marcada_pendiente_reasignacion(tree, gid, True)
        self.despacho_pendientes_marcados = marcados

    def marcar_todos_pendientes_reasignacion_despacho(self):
        tree = (getattr(self, "despacho_trees", {}) or {}).get("pendientes_reasignacion")
        if tree is None:
            return
        marcados = set(getattr(self, "despacho_pendientes_marcados", set()) or set())
        for row in getattr(tree, "_xtravon_all_rows", []) or []:
            gid = self.id_fila_despacho(tree, row)
            if gid is not None:
                marcados.add(gid)
                self.actualizar_fila_marcada_pendiente_reasignacion(tree, gid, True)
        self.despacho_pendientes_marcados = marcados

    def desmarcar_visibles_pendientes_reasignacion_despacho(self):
        tree = (getattr(self, "despacho_trees", {}) or {}).get("pendientes_reasignacion")
        if tree is None:
            return
        marcados = set(getattr(self, "despacho_pendientes_marcados", set()) or set())
        for item in tree.get_children():
            gid = self.id_fila_despacho(tree, tree.item(item, "values") or [])
            if gid is not None:
                marcados.discard(gid)
                self.actualizar_fila_marcada_pendiente_reasignacion(tree, gid, False)
        self.despacho_pendientes_marcados = marcados

    def desmarcar_todos_pendientes_reasignacion_despacho(self):
        tree = (getattr(self, "despacho_trees", {}) or {}).get("pendientes_reasignacion")
        if tree is None:
            return
        marcados = set(getattr(self, "despacho_pendientes_marcados", set()) or set())
        for row in getattr(tree, "_xtravon_all_rows", []) or []:
            gid = self.id_fila_despacho(tree, row)
            if gid is not None:
                marcados.discard(gid)
                self.actualizar_fila_marcada_pendiente_reasignacion(tree, gid, False)
        self.despacho_pendientes_marcados = marcados

    def reasignar_pendientes_marcadas_despacho(self):
        tree = (getattr(self, "despacho_trees", {}) or {}).get("pendientes_reasignacion")
        if tree is None:
            return
        ids = self.ids_seleccionados_tree_despacho(tree, incluir_marcados=True)
        if not ids:
            messagebox.showwarning("Sin seleccion", "Marque una o varias guias pendientes de reasignar.")
            return
        self.abrir_popup_reasignar_despacho(tree, guia_ids=ids)
    def configurar_menu_despacho(self, tree, key):
        menu = tk.Menu(tree, tearoff=0, bg=self.colors["bg_card"], fg=self.colors["text_dark"])
        if key == "solicitudes_pendientes":
            menu.add_command(label="Aprobar / asignar guia", command=lambda: self.aprobar_solicitud_despacho(tree))
            menu.add_command(label="Rechazar solicitud", command=lambda: self.rechazar_solicitud_despacho(tree))
            menu.add_command(label="Bloquear chofer/placa", command=lambda: self.bloquear_solicitud_despacho(tree))
        if key == "choferes_disponibles":
            menu.add_command(label="Bloquear chofer/placa", command=lambda: self.usar_chofer_despacho_y_bloquear(tree))
        if key == "bloqueos":
            menu.add_command(label="Desbloquear chofer/placa", command=lambda: self.desbloquear_bloqueo_despacho(tree))
        if key in ("pendientes_reasignacion", "guias_asignadas", "primer_escaneo", "segundo_escaneo", "completadas"):
            menu.add_command(label="Ver / asignar o reasignar", command=lambda: self.abrir_popup_reasignar_despacho(tree))
        if key == "pendientes_reasignacion":
            menu.add_command(label="Asignar guia", command=self.asignar_siguiente_viaje)
            menu.add_separator()
            menu.add_command(label="Marcar todos", command=self.marcar_todos_pendientes_reasignacion_despacho)
            menu.add_command(label="Marcar visibles", command=self.marcar_visibles_pendientes_reasignacion_despacho)
            menu.add_command(label="Desmarcar todos", command=self.desmarcar_todos_pendientes_reasignacion_despacho)
            menu.add_command(label="Desmarcar visibles", command=self.desmarcar_visibles_pendientes_reasignacion_despacho)
            menu.add_command(label="Reasignar marcadas", command=self.reasignar_pendientes_marcadas_despacho)
        if key in ("pendientes_reasignacion", "guias_asignadas", "primer_escaneo", "segundo_escaneo"):
            menu.add_command(label="Enviar QR", command=self.entregar_qr_despacho)
            menu.add_command(label="Liberar seleccion", command=self.liberar_guias_despacho)
            menu.add_command(label="Cancelar seleccion", command=self.cancelar_guias_despacho)

        def abrir_menu(event):
            item = tree.identify_row(event.y)
            if item:
                if item not in tree.selection():
                    tree.selection_set(item)
                tree.focus(item)
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()

        tree.bind("<Button-3>", abrir_menu)

    def usar_chofer_despacho_y_bloquear(self, tree):
        seleccion = tree.selection()
        if not seleccion:
            messagebox.showwarning("Sin chofer", "Seleccione un chofer.")
            return
        valores = tree.item(seleccion[0], "values")
        if not valores:
            return
        self.despacho_chofer_var.set(valores[0] if len(valores) > 0 else "")
        self.despacho_placa_var.set(valores[1] if len(valores) > 1 else "")
        self.bloquear_chofer_placa_despacho(
            chofer=valores[0] if len(valores) > 0 else "",
            placa=valores[1] if len(valores) > 1 else "",
        )

    def desbloquear_bloqueo_despacho(self, tree):
        seleccion = tree.selection()
        if not seleccion:
            messagebox.showwarning("Sin bloqueo", "Seleccione un bloqueo activo.")
            return
        valores = tree.item(seleccion[0], "values") or []
        bloqueo_id = self.safe_int(valores[0] if len(valores) > 0 else None, None)
        chofer = valores[1] if len(valores) > 1 else ""
        placa = valores[2] if len(valores) > 2 else ""
        if bloqueo_id is None:
            messagebox.showwarning("Bloqueo invalido", "No se pudo leer el ID del bloqueo.")
            return
        if not messagebox.askyesno(
            "Confirmar desbloqueo",
            f"Desbloquear chofer/placa?\n\nChofer: {chofer or '-'}\nPlaca: {placa or '-'}",
        ):
            return
        try:
            self.api_post_despacho_desbloquear(bloqueo_id)
            self.cargar_despacho_resumen()
            messagebox.showinfo("Bloqueo levantado", "El chofer o placa quedo habilitado nuevamente.")
        except Exception as e:
            messagebox.showerror("Desbloquear chofer/placa", str(e))

    def obtener_solicitud_despacho(self, tree):
        seleccion = tree.selection()
        if not seleccion:
            messagebox.showwarning("Sin solicitud", "Seleccione una solicitud de nuevo viaje.")
            return None
        valores = tree.item(seleccion[0], "values")
        if not valores:
            return None
        return {
            "id": self.safe_int(valores[0], None),
            "chofer": valores[1] if len(valores) > 1 else "",
            "placa": valores[2] if len(valores) > 2 else "",
            "guia": valores[3] if len(valores) > 3 else "",
            "producto": valores[4] if len(valores) > 4 else "",
        }

    def usar_solicitud_despacho(self, tree):
        solicitud = self.obtener_solicitud_despacho(tree)
        if not solicitud:
            return
        self.despacho_solicitud_id = solicitud["id"]
        self.despacho_chofer_var.set(solicitud["chofer"] or "")
        self.despacho_placa_var.set(solicitud["placa"] or "")
        if solicitud["producto"]:
            self.despacho_filtro_vars["producto"].set(solicitud["producto"])

    def aprobar_solicitud_despacho(self, tree):
        solicitud = self.obtener_solicitud_despacho(tree)
        if not solicitud:
            return
        self.usar_solicitud_despacho(tree)
        self.asignar_siguiente_viaje()

    def rechazar_solicitud_despacho(self, tree):
        solicitud = self.obtener_solicitud_despacho(tree)
        if not solicitud or not solicitud["id"]:
            return
        comentario = simpledialog.askstring(
            "Rechazar solicitud",
            "Comentario opcional:",
            parent=self,
        )
        if comentario is None:
            return
        if not messagebox.askyesno("Confirmar rechazo", "Desea rechazar esta solicitud de nuevo viaje?"):
            return
        try:
            self.api_post_despacho_rechazar_solicitud(solicitud["id"], comentario)
            messagebox.showinfo("Solicitud rechazada", "La solicitud fue rechazada correctamente.")
            self.cargar_despacho_resumen()
        except Exception as e:
            messagebox.showerror("Rechazar solicitud", str(e))

    def bloquear_solicitud_despacho(self, tree):
        solicitud = self.obtener_solicitud_despacho(tree)
        if not solicitud:
            return
        self.despacho_chofer_var.set(solicitud["chofer"] or "")
        self.despacho_placa_var.set(solicitud["placa"] or "")
        self.bloquear_chofer_placa_despacho()

    def abrir_popup_reasignar_despacho(self, tree, guia_ids=None):
        ids = list(guia_ids or self.ids_seleccionados_tree_despacho(tree))
        if not ids:
            messagebox.showwarning("Sin seleccion", "Seleccione o marque una o varias guias.")
            return
        valores = None
        for item in tree.selection():
            row = tuple(tree.item(item, "values") or [])
            if self.id_fila_despacho(tree, row) in ids:
                valores = row
                break
        if valores is None:
            for row in getattr(tree, "_xtravon_all_rows", []) or []:
                if self.id_fila_despacho(tree, row) in ids:
                    valores = tuple(row)
                    break
        if valores is None:
            messagebox.showwarning("Sin seleccion", "No se pudo leer la guia seleccionada.")
            return

        ventana = tk.Toplevel(self)
        ventana.title("Asignar / reasignar guia")
        ventana.geometry("700x560")
        ventana.minsize(640, 500)
        ventana.configure(bg=self.colors["bg_card"])
        ventana.transient(self)
        self.aplicar_icono_app(ventana)

        labels = (
            ("ID", "id"),
            ("Guia", "guia"),
            ("Cliente", "empresa"),
            ("Producto", "producto"),
            ("Chofer actual", "chofer"),
            ("Placa actual", "placa"),
            ("Estado", "estado_asignacion"),
        )
        tk.Label(
            ventana,
            text=f"Datos actuales de la guia{'s' if len(ids) > 1 else ''} ({len(ids)} seleccionada(s))",
            font=("Segoe UI", 16, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=18, pady=(18, 8))
        info = tk.Frame(ventana, bg=self.colors["bg_card"])
        info.pack(fill="x", padx=18, pady=(0, 12))
        for label, col in labels:
            value = self.valor_columna_despacho(tree, valores, col, "")
            row = tk.Frame(info, bg=self.colors["bg_card"])
            row.pack(fill="x", pady=2)
            tk.Label(row, text=f"{label}:", width=14, anchor="w", bg=self.colors["bg_card"], fg=self.colors["text_secondary"], font=("Segoe UI", 9, "bold")).pack(side="left")
            tk.Label(row, text=str(value), anchor="w", bg=self.colors["bg_card"], fg=self.colors["text_dark"], font=("Segoe UI", 9)).pack(side="left", fill="x", expand=True)

        chofer_var = tk.StringVar()
        placa_var = tk.StringVar()
        politica_var = tk.StringVar(value="Misma empresa")
        empresa_actual = str(self.valor_columna_despacho(tree, valores, "empresa", "") or "")
        empresa_destino_var = tk.StringVar(value=empresa_actual)
        campos = tk.Frame(ventana, bg=self.colors["bg_card"])
        campos.pack(fill="x", padx=18, pady=(0, 12))
        tk.Label(campos, text="Nuevo chofer", bg=self.colors["bg_card"], fg=self.colors["text_dark"], font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w", padx=(0, 8))
        tk.Label(campos, text="Nueva placa", bg=self.colors["bg_card"], fg=self.colors["text_dark"], font=("Segoe UI", 10, "bold")).grid(row=0, column=1, sticky="w")
        chofer_combo = self.crear_selector_filtrable_despacho(campos, chofer_var, width=30)
        placa_combo = self.crear_selector_filtrable_despacho(campos, placa_var, width=18)
        chofer_combo.grid(row=1, column=0, sticky="ew", padx=(0, 8))
        placa_combo.grid(row=1, column=1, sticky="ew")
        campos.grid_columnconfigure(0, weight=2)
        campos.grid_columnconfigure(1, weight=1)
        opciones = self.extraer_opciones_despacho(self.despacho_resumen or {})
        self.configurar_combo_filtrable_despacho(chofer_combo, opciones.get("choferes", []), chofer_var)
        self.configurar_combo_filtrable_despacho(placa_combo, opciones.get("placas", []), placa_var)

        empresa_frame = tk.Frame(ventana, bg=self.colors["bg_card"])
        empresa_frame.pack(fill="x", padx=18, pady=(0, 12))
        tk.Label(
            empresa_frame,
            text="Confirmacion empresa",
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
            font=("Segoe UI", 10, "bold"),
        ).grid(row=0, column=0, sticky="w", padx=(0, 8))
        tk.Label(
            empresa_frame,
            text="Empresa destino",
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
            font=("Segoe UI", 10, "bold"),
        ).grid(row=0, column=1, sticky="w")
        politica_combo = ttk.Combobox(
            empresa_frame,
            textvariable=politica_var,
            values=("Misma empresa", "Otra empresa"),
            state="readonly",
        )
        empresa_combo = self.crear_selector_filtrable_despacho(empresa_frame, empresa_destino_var, width=30)
        politica_combo.grid(row=1, column=0, sticky="ew", padx=(0, 8))
        empresa_combo.grid(row=1, column=1, sticky="ew")
        empresa_frame.grid_columnconfigure(0, weight=1)
        empresa_frame.grid_columnconfigure(1, weight=2)
        self.configurar_combo_filtrable_despacho(empresa_combo, opciones.get("empresas", []), empresa_destino_var)

        nota = tk.Label(
            ventana,
            text="Puede reasignar la guia al nuevo chofer/placa y confirmar si permanece en la misma empresa o se mueve a otra. Si el ciclo ya empezo, se conserva su etapa operativa.",
            bg=self.colors["bg_card"],
            fg=self.colors["text_aux"],
            font=("Segoe UI", 9),
            wraplength=560,
            justify="left",
        )
        nota.pack(anchor="w", padx=18, pady=(0, 12))

        def ejecutar():
            chofer = chofer_var.get().strip()
            placa = placa_var.get().strip()
            if not chofer or not placa:
                messagebox.showwarning("Dato requerido", "Indique nuevo chofer y placa.")
                return
            politica = "OTRA_EMPRESA" if politica_var.get().upper().startswith("OTRA") else "MISMA_EMPRESA"
            empresa_destino = empresa_destino_var.get().strip()
            if politica == "OTRA_EMPRESA" and not empresa_destino:
                messagebox.showwarning("Dato requerido", "Indique la empresa destino.")
                return
            empresa_texto = empresa_destino if politica == "OTRA_EMPRESA" else empresa_actual
            guia_texto = str(self.valor_columna_despacho(tree, valores, "guia", ids[0]) or ids[0])
            if len(ids) > 1:
                guia_texto = f"{len(ids)} guias marcadas"
            if not messagebox.askyesno(
                "Confirmar reasignacion",
                (
                    f"Se reasignara {guia_texto} a {chofer} / {placa}.\n\n"
                    f"Politica: {politica_var.get()}\n"
                    f"Empresa: {empresa_texto}\n\n"
                    "Desea continuar?"
                ),
            ):
                return
            try:
                self.api_post_despacho_reasignar(
                    ids,
                    chofer,
                    placa,
                    "Reasignacion desde tablero de despacho.",
                    politica_empresa=politica,
                    empresa_destino=empresa_destino if politica == "OTRA_EMPRESA" else None,
                )
                marcados = set(getattr(self, "despacho_pendientes_marcados", set()) or set())
                marcados.difference_update(ids)
                self.despacho_pendientes_marcados = marcados
                ventana.destroy()
                self.cargar_despacho_resumen()
                messagebox.showinfo("Reasignacion", f"{len(ids)} guia(s) reasignada(s) correctamente.")
            except Exception as e:
                messagebox.showerror("Reasignacion", str(e))

        actions = tk.Frame(ventana, bg=self.colors["bg_card"])
        actions.pack(fill="x", padx=18, pady=(0, 18))
        ttk.Button(actions, text="Reasignar por chofer", style="Olive.TButton", command=ejecutar).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Reasignar por guia", style="Olive.TButton", command=ejecutar).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Cancelar", style="Gray.TButton", command=ventana.destroy).pack(side="left")

    def usar_guia_disponible_despacho(self, tree):
        seleccion = tree.selection()
        if not seleccion:
            return
        valores = tree.item(seleccion[0], "values")
        if not valores or len(valores) < 4:
            return
        self.despacho_filtro_vars["empresa"].set(self.valor_columna_despacho(tree, valores, "empresa", "") or "")
        self.despacho_filtro_vars["producto"].set(self.valor_columna_despacho(tree, valores, "producto", "") or "")
        chofer = self.valor_columna_despacho(tree, valores, "chofer", "")
        placa = self.valor_columna_despacho(tree, valores, "placa", "")
        if chofer:
            self.despacho_chofer_var.set(chofer)
        if placa:
            self.despacho_placa_var.set(placa)

    def guia_disponible_seleccionada_despacho(self):
        for key in ("pendientes_reasignacion", "guias_asignadas"):
            tree = self.despacho_trees.get(key)
            if tree is None:
                continue
            seleccion = tree.selection()
            if not seleccion:
                continue
            valores = tree.item(seleccion[0], "values")
            if valores:
                return self.id_fila_despacho(tree, valores)
        return None

    def proponer_siguiente_viaje(self):
        if not self.despacho_operacion_id:
            messagebox.showwarning("Sin operacion", "Primero presione Buscar operacion activa.")
            return
        chofer = self.despacho_chofer_var.get().strip()
        placa = self.despacho_placa_var.get().strip()
        if not chofer or not placa:
            messagebox.showwarning("Dato requerido", "Indique chofer y placa.")
            return
        try:
            data = self.api_get_despacho_propuesta(self.despacho_operacion_id, chofer, placa, self.obtener_filtros_despacho())
            self.despacho_propuesta = data.get("propuesta")
            guia = (self.despacho_propuesta or {}).get("guia", {})
            plan = (self.despacho_propuesta or {}).get("plan_viajes") or {}
            plan_texto = plan.get("lectura") or ""
            self.despacho_propuesta_label.configure(
                text=(
                    f"Chofer {chofer} disponible para nuevo viaje. "
                    f"Guia sugerida: {guia.get('guia', '')} | Cliente: {guia.get('empresa', '')} | "
                    f"Producto: {guia.get('producto', '')} | Bodega: {guia.get('bodega_numero') or 'N/D'} | "
                    f"Guias disponibles: {self.despacho_propuesta.get('guias_disponibles', 0)}. "
                    f"{plan_texto} Accion: confirme con Asignar siguiente guia."
                )
            )
        except Exception as e:
            messagebox.showerror("Propuesta despacho", str(e))

    def asignar_siguiente_viaje(self):
        if not self.despacho_operacion_id:
            messagebox.showwarning("Sin operacion", "Primero presione Buscar operacion activa.")
            return
        chofer = self.despacho_chofer_var.get().strip()
        placa = self.despacho_placa_var.get().strip()
        if not chofer or not placa:
            messagebox.showwarning("Dato requerido", "Indique chofer y placa.")
            return
        empresa = self.despacho_filtro_vars["empresa"].get().strip()
        producto = self.despacho_filtro_vars["producto"].get().strip()
        base_id = None
        for key in ("pendientes_reasignacion", "guias_asignadas"):
            tree = self.despacho_trees.get(key)
            if tree is not None and tree.selection():
                valores = tree.item(tree.selection()[0], "values")
                if valores:
                    base_id = self.id_fila_despacho(tree, valores)
                    empresa = empresa or self.valor_columna_despacho(tree, valores, "empresa", "")
                    producto = producto or self.valor_columna_despacho(tree, valores, "producto", "")
                break
        if not base_id and not producto:
            messagebox.showwarning("Dato requerido", "Seleccione una guia pendiente/asignada o indique al menos el producto.")
            return
        if self.despacho_propuesta and self.despacho_propuesta.get("guia"):
            base_id = self.despacho_propuesta["guia"].get("id")
        base_id = base_id or self.guia_disponible_seleccionada_despacho()
        if not messagebox.askyesno(
            "Confirmar despacho",
            "Se asignara una guia disponible, se habilitara su QR y se enviara por el canal seleccionado. Desea continuar?",
        ):
            return
        try:
            data = self.api_post_despacho_asignar({
                "operacion_id": self.despacho_operacion_id,
                "base_operacion_id": base_id,
                "chofer": chofer,
                "placa": placa,
                "empresa": empresa,
                "producto": producto,
                "bodega_numero": self.safe_int(self.despacho_filtro_vars["bodega_numero"].get().strip(), None),
                "solicitud_id": self.despacho_solicitud_id,
                "canal_entrega": self.despacho_canal_var.get().strip(),
                "destinatario": self.despacho_destino_var.get().strip(),
                "comentario": "",
                "creado_por": "desktop",
            })
            guia = data.get("guia", {})
            canal = self.despacho_canal_var.get().strip()
            destino = self.despacho_destino_var.get().strip()
            entrega_msg = ""
            if canal in ("WHATSAPP", "CORREO"):
                try:
                    entrega_payload = {
                        "canal": canal,
                        "formato": "jpg",
                        "operacion_id": self.despacho_operacion_id,
                        "ids": [guia.get("id")],
                        "whatsapp_destino": destino if canal == "WHATSAPP" else None,
                        "email_destino": destino if canal == "CORREO" else None,
                    }
                    entrega = requests.post(
                        f"{self.api_base}/base-operaciones-camiones/entregar-qr",
                        json=entrega_payload,
                        timeout=90,
                    )
                    if entrega.status_code == 200:
                        entrega_msg = f"\nEntrega QR: {canal}"
                    else:
                        entrega_msg = f"\nEntrega QR pendiente: {self.obtener_detalle_error(entrega)}"
                except Exception as exc:
                    entrega_msg = f"\nEntrega QR pendiente: {exc}"
            entrega_msg = (entrega_msg or "") + "\nApp/handheld: guia disponible al sincronizar."
            messagebox.showinfo("Guia asignada", f"Guia asignada: {guia.get('guia', '')}\nChofer: {chofer}\nPlaca: {placa}")
            if entrega_msg:
                messagebox.showinfo("Entrega despacho", entrega_msg.strip())
            self.despacho_propuesta = None
            self.despacho_solicitud_id = None
            self.cargar_despacho_resumen()
        except Exception as e:
            messagebox.showerror("Asignacion despacho", str(e))

    def ids_seleccionados_despacho(self):
        ids = []
        for key in ("pendientes_reasignacion", "guias_asignadas", "primer_escaneo", "segundo_escaneo", "completadas"):
            tree = self.despacho_trees.get(key)
            if tree is None:
                continue
            for item_id in tree.selection():
                valores = tree.item(item_id, "values")
                if valores:
                    item_num = self.id_fila_despacho(tree, valores)
                    if item_num is not None:
                        ids.append(item_num)
        return ids

    def liberar_guias_despacho(self):
        if not self.asegurar_operacion_despacho():
            return
        ids = self.ids_seleccionados_despacho()
        if not ids:
            messagebox.showwarning("Sin seleccion", "Seleccione una o varias guias.")
            return
        if not messagebox.askyesno("Liberar guias", f"Se liberaran {len(ids)} guia(s) sin lecturas. Desea continuar?"):
            return
        try:
            self.api_post_despacho_liberar(ids, "")
            self.cargar_despacho_resumen()
        except Exception as e:
            messagebox.showerror("Liberar guias", str(e))

    def cancelar_guias_despacho(self):
        if not self.asegurar_operacion_despacho():
            return
        ids = self.ids_seleccionados_despacho()
        if not ids:
            messagebox.showwarning("Sin seleccion", "Seleccione una o varias guias.")
            return
        if not messagebox.askyesno("Cancelar guias", f"Se cancelaran {len(ids)} guia(s) sin lecturas. Desea continuar?"):
            return
        try:
            self.api_post_despacho_cancelar(ids, "")
            self.cargar_despacho_resumen()
        except Exception as e:
            messagebox.showerror("Cancelar guias", str(e))

    def entregar_qr_despacho(self):
        if not self.asegurar_operacion_despacho():
            return
        ids = self.ids_seleccionados_despacho()
        if not ids:
            messagebox.showwarning("Sin seleccion", "Seleccione una o varias guias asignadas.")
            return

        ventana = tk.Toplevel(self)
        ventana.title("Enviar QR despacho")
        ventana.geometry("520x360")
        ventana.minsize(500, 320)
        ventana.configure(bg=self.colors["bg_card"])
        ventana.transient(self)

        canal_var = tk.StringVar(value="WHATSAPP")
        destino_var = tk.StringVar()
        formato_var = tk.StringVar(value="jpg")

        tk.Label(
            ventana,
            text=f"Enviar QR seleccionados: {len(ids)} guia(s)",
            font=("Segoe UI", 16, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=18, pady=(18, 8))

        body = tk.Frame(ventana, bg=self.colors["bg_card"])
        body.pack(fill="both", expand=True, padx=18, pady=8)

        for label, var, values in [
            ("Canal", canal_var, ["WHATSAPP", "CORREO"]),
            ("Formato", formato_var, ["jpg", "pdf"]),
        ]:
            tk.Label(body, text=label, font=("Segoe UI", 10, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
            ttk.Combobox(body, textvariable=var, values=values, state="readonly").pack(fill="x", pady=(0, 10))

        destino_label = tk.Label(body, text="Destino", font=("Segoe UI", 10, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"])
        destino_label.pack(anchor="w")
        destino_entry = ttk.Entry(body, textvariable=destino_var)
        destino_entry.pack(fill="x", pady=(0, 10))

        def actualizar_destino(*_args):
            canal = canal_var.get()
            if canal == "WHATSAPP":
                destino_label.configure(text="WhatsApp destino")
            else:
                destino_label.configure(text="Email destino")

        canal_var.trace_add("write", actualizar_destino)
        actualizar_destino()

        def ejecutar():
            canal = canal_var.get()
            destino = destino_var.get().strip()
            payload = {
                "canal": canal,
                "formato": formato_var.get(),
                "operacion_id": self.despacho_operacion_id,
                "ids": ids,
                "whatsapp_destino": destino if canal == "WHATSAPP" else None,
                "email_destino": destino if canal == "CORREO" else None,
            }
            if canal in ("WHATSAPP", "CORREO") and not destino:
                if not messagebox.askyesno("Destino vacio", "No indico destino. Se usara el contacto registrado si existe. Desea continuar?"):
                    return
            ventana.destroy()

            def tarea():
                respuesta = requests.post(
                    f"{self.api_base}/base-operaciones-camiones/entregar-qr",
                    json=payload,
                    timeout=120,
                )
                if respuesta.status_code != 200:
                    raise RuntimeError(self.obtener_detalle_error(respuesta))
                return respuesta.json()

            def al_terminar(data):
                enviados = len([item for item in data.get("entregas", []) if item.get("estado") == "ENVIADO"])
                errores = len([item for item in data.get("entregas", []) if item.get("estado") == "ERROR_ENVIO"])
                messagebox.showinfo(
                    "Entrega QR despacho",
                    f"Lote: {data.get('lote_codigo')}\nQR preparados: {data.get('total_qr', 0)}\nEnviados: {enviados}\nErrores: {errores}",
                )

            self.ejecutar_en_segundo_plano("Entrega QR", "Enviando QR seleccionados...", tarea, al_terminar)

        actions = tk.Frame(ventana, bg=self.colors["bg_card"])
        actions.pack(fill="x", padx=18, pady=(0, 18))
        ttk.Button(actions, text="Enviar", style="Olive.TButton", command=ejecutar).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Cancelar", style="Gray.TButton", command=ventana.destroy).pack(side="left")

    def bloquear_chofer_placa_despacho(self, chofer=None, placa=None):
        if not self.despacho_operacion_id:
            messagebox.showwarning("Sin operacion", "Primero presione Buscar operacion activa.")
            return
        chofer = (chofer if chofer is not None else self.despacho_chofer_var.get()).strip()
        placa = (placa if placa is not None else self.despacho_placa_var.get()).strip()
        motivo = ""
        if not chofer and not placa:
            messagebox.showwarning("Dato requerido", "Indique chofer o placa.")
            return
        if not messagebox.askyesno(
            "Confirmar bloqueo",
            f"Bloquear para nuevas asignaciones?\n\nChofer: {chofer or '-'}\nPlaca: {placa or '-'}",
        ):
            return
        try:
            data = self.api_post_despacho_bloquear(self.despacho_operacion_id, chofer, placa, motivo)
            self.cargar_despacho_resumen()
            movidas = self.safe_int(data.get("guias_movidas", 0), 0) if isinstance(data, dict) else 0
            messagebox.showinfo(
                "Bloqueo registrado",
                "Chofer o placa bloqueado para nuevas asignaciones.\n"
                f"Guias movidas a pendientes de reasignar: {movidas}",
            )
        except Exception as e:
            messagebox.showerror("Bloquear chofer/placa", str(e))

    # =========================================================
    # ROLES Y PERMISOS
    # =========================================================

    def api_get_rbac_catalogo(self):
        respuesta = requests.get(f"{self.api_base}/rbac/catalogo", timeout=60)
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_rbac_usuario(self, usuario_id):
        respuesta = requests.get(f"{self.api_base}/rbac/usuarios/{usuario_id}", timeout=60)
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_crear_rbac_usuario(self, payload):
        respuesta = requests.post(f"{self.api_base}/rbac/usuarios", json=payload, timeout=60)
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_crear_rbac_rol(self, payload):
        respuesta = requests.post(f"{self.api_base}/rbac/roles", json=payload, timeout=60)
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_asignar_rbac(self, payload):
        respuesta = requests.post(f"{self.api_base}/rbac/asignar", json=payload, timeout=60)
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def comparar_versiones(self, actual, remota):
        def partes(version):
            salida = []
            for item in str(version or "0").split("."):
                try:
                    salida.append(int("".join(ch for ch in item if ch.isdigit()) or 0))
                except Exception:
                    salida.append(0)
            return (salida + [0, 0, 0])[:3]

        a = partes(actual)
        b = partes(remota)
        return (b > a) - (b < a)

    def normalizar_version_actualizacion(self, version):
        texto = str(version or "").strip().lower()
        if texto.startswith("v"):
            texto = texto[1:]
        return texto or "0"

    def ruta_marca_actualizacion(self):
        return app_user_data_path("desktop_update_attempt.txt")

    def leer_marca_actualizacion(self):
        try:
            with open(self.ruta_marca_actualizacion(), "r", encoding="utf-8") as archivo:
                return archivo.read().strip()
        except Exception:
            return ""

    def guardar_marca_actualizacion(self, version):
        try:
            with open(self.ruta_marca_actualizacion(), "w", encoding="utf-8") as archivo:
                archivo.write(self.normalizar_version_actualizacion(version))
        except Exception:
            pass

    def descargar_e_instalar_actualizacion(self, download_url, remote_version):
        if not download_url:
            messagebox.showwarning("Actualizaciones", "La version existe, pero no tiene URL de descarga configurada.")
            return

        version_key = self.normalizar_version_actualizacion(remote_version)
        if self._update_installing_version == version_key:
            return
        self._update_installing_version = version_key
        self.guardar_marca_actualizacion(version_key)

        if not str(download_url).lower().split("?")[0].endswith(".exe"):
            webbrowser.open(download_url)
            return

        def worker():
            destino = os.path.join(tempfile.gettempdir(), f"XTRAVON_ONE_Update_{remote_version}.exe")
            try:
                with requests.get(download_url, stream=True, timeout=120) as respuesta:
                    respuesta.raise_for_status()
                    with open(destino, "wb") as archivo:
                        for chunk in respuesta.iter_content(chunk_size=1024 * 256):
                            if chunk:
                                archivo.write(chunk)

                args = [
                    destino,
                    "/SILENT",
                    "/SUPPRESSMSGBOXES",
                    "/CLOSEAPPLICATIONS",
                    "/FORCECLOSEAPPLICATIONS",
                    "/RESTARTAPPLICATIONS",
                ]
                subprocess.Popen(args, close_fds=True)
                self.after(
                    300,
                    lambda: (
                        messagebox.showinfo(
                            "Actualizacion",
                            "XTRAVON ONE cerrara para instalar la nueva version. Si Windows solicita permisos, acepte la instalacion.",
                        ),
                        self.destroy(),
                    ),
                )
            except Exception as exc:
                self._update_installing_version = None
                self.after(0, lambda: messagebox.showerror("Actualizaciones", f"No se pudo instalar la actualizacion:\n{exc}"))

        messagebox.showinfo("Actualizacion", "Descargando actualizacion. XTRAVON continuara en segundo plano hasta iniciar el instalador.")
        threading.Thread(target=worker, daemon=True).start()

    def verificar_actualizacion_remota(self, silencioso=False):
        try:
            respuesta = requests.get(f"{self.api_base}/releases/desktop/latest", timeout=8)
            if respuesta.status_code != 200:
                if not silencioso:
                    messagebox.showwarning("Actualizaciones", self.obtener_detalle_error(respuesta))
                return
            data = respuesta.json()
            remote_version = data.get("version")
            download_url = data.get("download_url")
            notes = data.get("notes") or ""
            remote_key = self.normalizar_version_actualizacion(remote_version)
            marca_instalada = self.leer_marca_actualizacion()
            local_key = self.normalizar_version_actualizacion(self.app_version)

            if remote_key and remote_key == local_key:
                self.guardar_marca_actualizacion(remote_key)
                if not silencioso:
                    messagebox.showinfo(
                        "Actualizaciones",
                        f"XTRAVON ONE ya esta actualizado.\nVersion actual: {self.app_version}",
                    )
                return

            if marca_instalada == remote_key:
                if not silencioso:
                    messagebox.showinfo(
                        "Actualizaciones",
                        f"La version {remote_version} ya fue marcada como instalada.\n"
                        "Reinicie XTRAVON ONE si aun ve la version anterior.",
                    )
                return

            if self.comparar_versiones(self.app_version, remote_version) <= 0:
                if not silencioso:
                    messagebox.showinfo("Actualizaciones", f"XTRAVON ONE ya esta actualizado.\nVersion actual: {self.app_version}")
                return
            if self._update_prompted_version == remote_key or self._update_installing_version == remote_key:
                return

            mensaje = (
                f"Hay una nueva version disponible.\n\n"
                f"Actual: {self.app_version}\n"
                f"Nueva: {remote_version}\n\n"
                f"{notes}\n\n"
                "Desea descargar e instalar la actualizacion ahora?"
            )
            self._update_prompted_version = remote_key
            if messagebox.askyesno("Actualizacion disponible", mensaje):
                self.descargar_e_instalar_actualizacion(download_url, remote_version)
        except Exception as exc:
            if not silencioso:
                messagebox.showerror("Actualizaciones", str(exc))

    def api_get_resumen_control_operativo(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/control-operativo/operaciones/{operacion_id}/resumen",
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_control_plan_operativo(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/control-operativo/operaciones/{operacion_id}/control-plan",
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_salud_operativa(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/control-operativo/operaciones/{operacion_id}/salud-operativa",
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_spc_operativo(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/control-operativo/operaciones/{operacion_id}/spc",
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_bloqueos_inteligentes(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/control-operativo/operaciones/{operacion_id}/bloqueos-inteligentes",
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_excepciones_operativas(self, operacion_id, params=None):
        respuesta = requests.get(
            f"{self.api_base}/control-operativo/operaciones/{operacion_id}/excepciones",
            params=params or {},
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_auditoria_senior(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/control-operativo/operaciones/{operacion_id}/auditoria-senior",
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_cierre_guiado(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/control-operativo/operaciones/{operacion_id}/cierre-guiado",
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_modo_offline(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/control-operativo/operaciones/{operacion_id}/modo-offline",
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_get_productividad_operativa(self, operacion_id):
        respuesta = requests.get(
            f"{self.api_base}/control-operativo/operaciones/{operacion_id}/productividad",
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_ejecutar_cierre_guiado(self, operacion_id, payload=None):
        respuesta = requests.post(
            f"{self.api_base}/control-operativo/operaciones/{operacion_id}/cierre-guiado/ejecutar",
            json=payload or {},
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_generar_excepciones_desde_bloqueos(self, operacion_id):
        respuesta = requests.post(
            f"{self.api_base}/control-operativo/operaciones/{operacion_id}/excepciones/generar-desde-bloqueos",
            timeout=90,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def api_cerrar_excepcion_operativa(self, excepcion_id, comentario, usuario="desktop"):
        respuesta = requests.post(
            f"{self.api_base}/control-operativo/excepciones/{excepcion_id}/cerrar",
            json={"comentario": comentario, "usuario": usuario},
            timeout=60,
        )
        if respuesta.status_code != 200:
            raise RuntimeError(self.obtener_detalle_error(respuesta))
        return respuesta.json()

    def show_roles_permisos(self):
        self.clear_content()
        self.highlight_sidebar_button("Roles y Permisos")

        self.rbac_catalogo = {"usuarios": [], "roles": [], "permisos": []}
        self.rbac_usuario_var = tk.StringVar()
        self.rbac_rol_var = tk.StringVar()
        self.rbac_nombre_var = tk.StringVar()
        self.rbac_usuario_login_var = tk.StringVar()
        self.rbac_email_var = tk.StringVar()
        self.rbac_nuevo_rol_var = tk.StringVar()
        self.rbac_nuevo_rol_desc_var = tk.StringVar()
        self.rbac_permisos_tree = None

        self.create_page_title(
            self.content,
            "Roles y Permisos",
            "Administre usuarios, roles y permisos de acceso por mÃ³dulo.",
        )

        scroll_host = tk.Frame(self.content, bg=self.colors["bg_main"])
        scroll_host.pack(fill="both", expand=True, padx=25, pady=(0, 20))

        canvas = tk.Canvas(scroll_host, bg=self.colors["bg_main"], highlightthickness=0)
        scroll_y = ttk.Scrollbar(scroll_host, orient="vertical", command=canvas.yview)
        scroll_x = ttk.Scrollbar(scroll_host, orient="horizontal", command=canvas.xview)
        body = tk.Frame(canvas, bg=self.colors["bg_main"])

        window_id = canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        self.bind_scroll_canvas(canvas, body, window_id, min_width=1040)

        canvas.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        scroll_host.grid_rowconfigure(0, weight=1)
        scroll_host.grid_columnconfigure(0, weight=1)

        top = tk.Frame(body, bg=self.colors["bg_main"])
        top.pack(fill="x", pady=(0, 10))

        usuarios_panel = tk.Frame(top, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1)
        usuarios_panel.pack(side="left", fill="both", expand=True, padx=(0, 8))

        tk.Label(usuarios_panel, text="Usuario", font=("Segoe UI", 13, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w", padx=14, pady=(12, 6))

        usuario_form = tk.Frame(usuarios_panel, bg=self.colors["bg_card"])
        usuario_form.pack(fill="x", padx=14, pady=(0, 10))

        for idx, (var, label) in enumerate([
            (self.rbac_nombre_var, "Nombre"),
            (self.rbac_usuario_login_var, "Usuario"),
            (self.rbac_email_var, "Email"),
        ]):
            box = tk.Frame(usuario_form, bg=self.colors["bg_card"])
            box.grid(row=0, column=idx, sticky="ew", padx=5, pady=4)
            tk.Label(box, text=label, font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
            ttk.Entry(box, textvariable=var).pack(fill="x")
            usuario_form.grid_columnconfigure(idx, weight=1)

        usuario_actions = tk.Frame(usuarios_panel, bg=self.colors["bg_card"])
        usuario_actions.pack(fill="x", padx=14, pady=(0, 12))
        ttk.Button(usuario_actions, text="Crear/Actualizar usuario", style="Olive.TButton", command=self.crear_rbac_usuario_front).pack(side="left", padx=(0, 8))
        ttk.Button(usuario_actions, text="Cargar catalogo", style="Gray.TButton", command=self.cargar_rbac_catalogo_front).pack(side="left")

        roles_panel = tk.Frame(top, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1, width=420)
        roles_panel.pack(side="right", fill="both")
        roles_panel.pack_propagate(False)

        tk.Label(roles_panel, text="Rol", font=("Segoe UI", 13, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w", padx=14, pady=(12, 6))
        rol_form = tk.Frame(roles_panel, bg=self.colors["bg_card"])
        rol_form.pack(fill="x", padx=14, pady=(0, 10))

        tk.Label(rol_form, text="Nombre rol", font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
        ttk.Entry(rol_form, textvariable=self.rbac_nuevo_rol_var).pack(fill="x", pady=(0, 5))
        tk.Label(rol_form, text="DescripciÃ³n", font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
        ttk.Entry(rol_form, textvariable=self.rbac_nuevo_rol_desc_var).pack(fill="x")

        ttk.Button(roles_panel, text="Crear/Actualizar rol", style="Olive.TButton", command=self.crear_rbac_rol_front).pack(anchor="w", padx=14, pady=(0, 12))

        assign_panel = tk.Frame(body, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1)
        assign_panel.pack(fill="x", pady=(0, 10))

        tk.Label(assign_panel, text="AsignaciÃ³n SAP-like", font=("Segoe UI", 13, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w", padx=14, pady=(12, 6))

        combo_grid = tk.Frame(assign_panel, bg=self.colors["bg_card"])
        combo_grid.pack(fill="x", padx=14, pady=(0, 12))

        usuario_box = tk.Frame(combo_grid, bg=self.colors["bg_card"])
        usuario_box.grid(row=0, column=0, sticky="ew", padx=5)
        tk.Label(usuario_box, text="Usuario", font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
        self.rbac_usuario_combo = ttk.Combobox(usuario_box, textvariable=self.rbac_usuario_var, state="readonly")
        self.rbac_usuario_combo.pack(fill="x")
        self.rbac_usuario_combo.bind("<<ComboboxSelected>>", lambda _e: None)

        rol_box = tk.Frame(combo_grid, bg=self.colors["bg_card"])
        rol_box.grid(row=0, column=1, sticky="ew", padx=5)
        tk.Label(rol_box, text="Rol", font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
        self.rbac_rol_combo = ttk.Combobox(rol_box, textvariable=self.rbac_rol_var, state="readonly")
        self.rbac_rol_combo.pack(fill="x")

        combo_grid.grid_columnconfigure(0, weight=1)
        combo_grid.grid_columnconfigure(1, weight=1)

        permisos_panel = tk.Frame(body, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1)
        permisos_panel.pack(fill="both", expand=True)

        header = tk.Frame(permisos_panel, bg=self.colors["bg_card"])
        header.pack(fill="x", padx=14, pady=(12, 6))
        tk.Label(header, text="Permisos", font=("Segoe UI", 13, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(side="left")
        ttk.Button(header, text="Guardar permisos", style="Olive.TButton", command=self.guardar_rbac_asignacion_front).pack(side="right", padx=(8, 0))
        ttk.Button(header, text="Marcar todos", style="Gray.TButton", command=self.marcar_todos_rbac_permisos).pack(side="right")

        table_frame = tk.Frame(permisos_panel, bg=self.colors["bg_card"])
        table_frame.pack(fill="both", expand=True, padx=14, pady=(0, 14))

        columns = ("id", "modulo", "accion", "codigo", "descripcion")
        self.rbac_permisos_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=18, selectmode="extended")

        for col, heading, width in [
            ("id", "ID", 55),
            ("modulo", "Modulo", 170),
            ("accion", "Accion", 120),
            ("codigo", "Codigo", 190),
            ("descripcion", "Descripcion", 420),
        ]:
            self.rbac_permisos_tree.heading(col, text=heading)
            self.rbac_permisos_tree.column(col, width=width, anchor="center" if col != "descripcion" else "w")

        scroll_tree_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.rbac_permisos_tree.yview)
        scroll_tree_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.rbac_permisos_tree.xview)
        self.rbac_permisos_tree.configure(yscrollcommand=scroll_tree_y.set, xscrollcommand=scroll_tree_x.set)
        self.rbac_permisos_tree.grid(row=0, column=0, sticky="nsew")
        scroll_tree_y.grid(row=0, column=1, sticky="ns")
        scroll_tree_x.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        self.rbac_catalogo = {"usuarios": [], "roles": [], "permisos": []}

    def cargar_rbac_catalogo_front(self):
        try:
            self.rbac_catalogo = self.api_get_rbac_catalogo()
            self.render_rbac_catalogo()
        except Exception as e:
            messagebox.showerror("Error RBAC", str(e))

    def render_rbac_catalogo(self):
        usuarios = self.rbac_catalogo.get("usuarios", [])
        roles = self.rbac_catalogo.get("roles", [])
        permisos = self.rbac_catalogo.get("permisos", [])

        self.rbac_usuario_combo["values"] = [f"{u['id']} | {u['nombre']} ({u['usuario']})" for u in usuarios]
        self.rbac_rol_combo["values"] = [f"{r['id']} | {r['nombre']}" for r in roles]

        for item in self.rbac_permisos_tree.get_children():
            self.rbac_permisos_tree.delete(item)

        for permiso in permisos:
            self.rbac_permisos_tree.insert(
                "",
                "end",
                values=(
                    permiso.get("id", ""),
                    permiso.get("modulo", ""),
                    permiso.get("accion", ""),
                    permiso.get("codigo", ""),
                    permiso.get("descripcion", ""),
                ),
            )

    def extraer_id_combo(self, valor):
        try:
            return int(str(valor).split("|")[0].strip())
        except Exception:
            return None

    def crear_rbac_usuario_front(self):
        nombre = self.rbac_nombre_var.get().strip()
        usuario = self.rbac_usuario_login_var.get().strip()

        if not nombre or not usuario:
            messagebox.showwarning("Datos requeridos", "Debe indicar nombre y usuario.")
            return

        try:
            self.api_crear_rbac_usuario({
                "nombre": nombre,
                "usuario": usuario,
                "email": self.rbac_email_var.get().strip() or None,
                "activo": True,
            })
            self.rbac_nombre_var.set("")
            self.rbac_usuario_login_var.set("")
            self.rbac_email_var.set("")
            self.cargar_rbac_catalogo_front()
            messagebox.showinfo("Usuario guardado", "Usuario creado/actualizado correctamente.")
        except Exception as e:
            messagebox.showerror("Error usuario", str(e))

    def crear_rbac_rol_front(self):
        nombre = self.rbac_nuevo_rol_var.get().strip()
        if not nombre:
            messagebox.showwarning("Dato requerido", "Debe indicar el nombre del rol.")
            return

        try:
            self.api_crear_rbac_rol({
                "nombre": nombre,
                "descripcion": self.rbac_nuevo_rol_desc_var.get().strip() or None,
                "activo": True,
            })
            self.rbac_nuevo_rol_var.set("")
            self.rbac_nuevo_rol_desc_var.set("")
            self.cargar_rbac_catalogo_front()
            messagebox.showinfo("Rol guardado", "Rol creado/actualizado correctamente.")
        except Exception as e:
            messagebox.showerror("Error rol", str(e))

    def cargar_rbac_usuario_seleccionado(self):
        usuario_id = self.extraer_id_combo(self.rbac_usuario_var.get())
        if not usuario_id:
            return

        try:
            data = self.api_get_rbac_usuario(usuario_id)
            rol = data.get("rol")
            permisos = data.get("permisos", [])
            permiso_ids = {p.get("id") for p in permisos}

            if rol:
                self.rbac_rol_var.set(f"{rol['id']} | {rol['nombre']}")

            self.rbac_permisos_tree.selection_remove(self.rbac_permisos_tree.selection())
            for item in self.rbac_permisos_tree.get_children():
                valores = self.rbac_permisos_tree.item(item, "values")
                if valores and self.safe_int(valores[0], None) in permiso_ids:
                    self.rbac_permisos_tree.selection_add(item)

        except Exception as e:
            messagebox.showerror("Error usuario", str(e))

    def marcar_todos_rbac_permisos(self):
        if self.rbac_permisos_tree is None:
            return
        self.rbac_permisos_tree.selection_set(self.rbac_permisos_tree.get_children())

    def obtener_rbac_permiso_ids(self):
        ids = []
        for item in self.rbac_permisos_tree.selection():
            valores = self.rbac_permisos_tree.item(item, "values")
            if valores:
                permiso_id = self.safe_int(valores[0], None)
                if permiso_id:
                    ids.append(permiso_id)
        return ids

    def guardar_rbac_asignacion_front(self):
        usuario_id = self.extraer_id_combo(self.rbac_usuario_var.get())
        rol_id = self.extraer_id_combo(self.rbac_rol_var.get())

        if not usuario_id:
            messagebox.showwarning("Usuario requerido", "Debe seleccionar un usuario.")
            return

        if not rol_id:
            messagebox.showwarning("Rol requerido", "Debe seleccionar un rol.")
            return

        permiso_ids = self.obtener_rbac_permiso_ids()
        if not permiso_ids:
            messagebox.showwarning("Permisos requeridos", "Debe seleccionar al menos un permiso.")
            return

        try:
            self.api_asignar_rbac({
                "usuario_id": usuario_id,
                "rol_id": rol_id,
                "permiso_ids": permiso_ids,
            })
            messagebox.showinfo("Permisos guardados", "Rol y permisos asignados correctamente.")
        except Exception as e:
            messagebox.showerror("Error permisos", str(e))

    # =========================================================
    # INFORMES - CARGA BAJO DEMANDA
    # =========================================================

    def show_informes(self):
        self.clear_content()
        self.highlight_sidebar_button("Informes")

        self.informes_tree = None
        self.informes_cache = []
        self.informes_exportar_formato_var = tk.StringVar(value="PDF")
        self.informes_tipo_reporte_var = tk.StringVar(value="Cliente ejecutivo")
        self.informes_filter_vars = {
            "empresa": tk.StringVar(),
            "producto": tk.StringVar(),
            "bodega_numero": tk.StringVar(),
            "fecha_desde": tk.StringVar(),
            "fecha_hasta": tk.StringVar(),
        }
        self.informes_fecha_display_vars = {
            "fecha_desde": tk.StringVar(value="All dates"),
            "fecha_hasta": tk.StringVar(value="All dates"),
        }
        self.informes_filter_widgets = {}

        self.create_page_title(
            self.content,
            "Informes",
            "Busque operaciones por buque para ver o descargar informes cuando el modulo PDF este activo.",
        )

        scroll_host = tk.Frame(self.content, bg=self.colors["bg_main"])
        scroll_host.pack(fill="both", expand=True, padx=25, pady=(0, 20))

        canvas = tk.Canvas(scroll_host, bg=self.colors["bg_main"], highlightthickness=0)
        scroll_y = ttk.Scrollbar(scroll_host, orient="vertical", command=canvas.yview)
        scroll_x = ttk.Scrollbar(scroll_host, orient="horizontal", command=canvas.xview)
        body = tk.Frame(canvas, bg=self.colors["bg_main"])

        window_id = canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        self.bind_scroll_canvas(canvas, body, window_id, min_width=1040)

        canvas.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        scroll_host.grid_rowconfigure(0, weight=1)
        scroll_host.grid_columnconfigure(0, weight=1)

        panel = tk.Frame(body, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1)
        panel.pack(fill="both", expand=True)

        header = tk.Frame(panel, bg=self.colors["bg_card"])
        header.pack(fill="x", padx=14, pady=(12, 8))
        tk.Label(header, text="Informes por buque", font=("Segoe UI", 13, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(side="left")
        ttk.Button(header, text="Buscar informes", style="Olive.TButton", command=self.cargar_informes_operaciones).pack(side="right")

        filtros = tk.Frame(panel, bg=self.colors["bg_card"])
        filtros.pack(fill="x", padx=14, pady=(0, 8))
        campos_filtro = [
            ("empresa", "Cliente", "combo"),
            ("producto", "Producto", "combo"),
            ("bodega_numero", "Bodega", "combo"),
            ("fecha_desde", "Desde", "date"),
            ("fecha_hasta", "Hasta", "date"),
        ]
        for idx, (key, label, kind) in enumerate(campos_filtro):
            box = tk.Frame(filtros, bg=self.colors["bg_card"])
            box.grid(row=idx // 5, column=idx % 5, sticky="ew", padx=5, pady=4)
            tk.Label(box, text=label, font=("Segoe UI", 9, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w")
            if kind == "date":
                row = tk.Frame(box, bg=self.colors["bg_card"])
                row.pack(fill="x")
                ttk.Entry(row, textvariable=self.informes_fecha_display_vars[key], state="readonly").pack(side="left", fill="x", expand=True)
                ttk.Button(row, text="...", style="Gray.TButton", width=4, command=lambda campo=key: self.abrir_selector_fecha_informe(campo)).pack(side="left", padx=(4, 0))
            else:
                widget = ttk.Combobox(box, textvariable=self.informes_filter_vars[key], values=[], state="readonly")
                widget.pack(fill="x")
                self.informes_filter_widgets[key] = widget
            filtros.grid_columnconfigure(idx % 5, weight=1)

        actions = tk.Frame(panel, bg=self.colors["bg_card"])
        actions.pack(fill="x", padx=14, pady=(0, 8))
        ttk.Combobox(
            actions,
            textvariable=self.informes_tipo_reporte_var,
            values=["Cliente ejecutivo", "Operativo sintetizado"],
            state="readonly",
            width=24,
        ).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Cargar filtros", style="Gray.TButton", command=self.cargar_filtros_informe_seleccionado).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Limpiar filtros", style="Gray.TButton", command=self.limpiar_filtros_informes).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Ver informe", style="Gray.TButton", command=self.ver_informe_seleccionado).pack(side="left", padx=(0, 8))
        ttk.Combobox(actions, textvariable=self.informes_exportar_formato_var, values=["PDF", "Excel"], state="readonly", width=8).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Exportar", style="Gray.TButton", command=self.descargar_informe_seleccionado).pack(side="left")

        table_frame = tk.Frame(panel, bg=self.colors["bg_card"])
        table_frame.pack(fill="both", expand=True, padx=14, pady=(0, 14))

        columns = ("id", "codigo", "buque", "inicio", "cierre", "producto", "estado")
        self.informes_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=22)

        for col, heading, width in [
            ("id", "ID", 55),
            ("codigo", "Codigo", 180),
            ("buque", "Buque", 220),
            ("inicio", "Inicio", 130),
            ("cierre", "Cierre", 130),
            ("producto", "Producto", 220),
            ("estado", "Estado", 120),
        ]:
            self.informes_tree.heading(col, text=heading)
            self.informes_tree.column(col, width=width, anchor="center")

        table_scroll_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.informes_tree.yview)
        table_scroll_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.informes_tree.xview)
        self.informes_tree.configure(yscrollcommand=table_scroll_y.set, xscrollcommand=table_scroll_x.set)
        self.informes_tree.grid(row=0, column=0, sticky="nsew")
        table_scroll_y.grid(row=0, column=1, sticky="ns")
        table_scroll_x.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

    def tipo_reporte_informes_api(self):
        valor = self.informes_tipo_reporte_var.get().strip().lower() if hasattr(self, "informes_tipo_reporte_var") else ""
        return "ejecutivo" if "operativo" in valor else "cliente"

    def obtener_parametros_informe(self):
        params = {"tipo_reporte": self.tipo_reporte_informes_api()}
        for key, var in getattr(self, "informes_filter_vars", {}).items():
            value = var.get().strip()
            if value:
                params[key] = self.safe_int(value, value) if key == "bodega_numero" else value
        return params

    def limpiar_filtros_informes(self):
        for var in getattr(self, "informes_filter_vars", {}).values():
            var.set("")
        if hasattr(self, "informes_fecha_display_vars"):
            self.informes_fecha_display_vars["fecha_desde"].set("All dates")
            self.informes_fecha_display_vars["fecha_hasta"].set("All dates")

    def aplicar_opciones_filtros_informes(self, opciones):
        mapa = {
            "empresa": "empresas",
            "producto": "productos",
            "bodega_numero": "bodegas",
        }
        for key, option_key in mapa.items():
            widget = self.informes_filter_widgets.get(key) if hasattr(self, "informes_filter_widgets") else None
            if widget is not None:
                valores = [""] + [str(v) for v in opciones.get(option_key, []) if v not in (None, "")]
                widget["values"] = valores

    def cargar_filtros_informe_seleccionado(self):
        operacion = self.obtener_operacion_informe_seleccionada()
        if not operacion:
            return

        def tarea():
            return self.api_get_reporte_buque_filtros(operacion.get("id"))

        def al_terminar(data):
            opciones = data.get("opciones", {}) if isinstance(data, dict) else {}
            self.aplicar_opciones_filtros_informes(opciones)

        self.ejecutar_en_segundo_plano(
            "Filtros de informe",
            "Cargando opciones de cliente, producto, bodega y fechas...",
            tarea,
            al_terminar,
        )

    def abrir_selector_fecha_informe(self, campo):
        popup = tk.Toplevel(self)
        popup.title("Select date")
        popup.geometry("390x220")
        popup.configure(bg=self.colors["bg_card"])
        popup.transient(self)
        popup.grab_set()

        hoy = date.today()
        actual = self.informes_filter_vars.get(campo).get() if hasattr(self, "informes_filter_vars") else ""
        try:
            base = datetime.strptime(actual, "%Y-%m-%d").date() if actual else hoy
        except Exception:
            base = hoy

        dia_var = tk.StringVar(value=str(base.day))
        mes_var = tk.StringVar(value=self.meses_en()[base.month - 1])
        anio_var = tk.StringVar(value=str(base.year))

        contenido = tk.Frame(popup, bg=self.colors["bg_card"])
        contenido.pack(fill="both", expand=True, padx=18, pady=18)
        tk.Label(contenido, text="Report date range", font=("Segoe UI", 13, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w", pady=(0, 12))

        fila = tk.Frame(contenido, bg=self.colors["bg_card"])
        fila.pack(fill="x")
        ttk.Combobox(fila, textvariable=dia_var, values=[str(i) for i in range(1, 32)], state="readonly", width=6).pack(side="left", padx=(0, 8))
        ttk.Combobox(fila, textvariable=mes_var, values=self.meses_en(), state="readonly", width=16).pack(side="left", padx=(0, 8))
        ttk.Combobox(fila, textvariable=anio_var, values=[str(i) for i in range(hoy.year - 5, hoy.year + 6)], state="readonly", width=8).pack(side="left")

        def aplicar():
            try:
                mes = self.meses_en().index(mes_var.get()) + 1
                fecha = date(int(anio_var.get()), mes, int(dia_var.get()))
            except Exception:
                messagebox.showerror("Invalid date", "Select a valid date.")
                return
            self.informes_filter_vars[campo].set(fecha.isoformat())
            self.informes_fecha_display_vars[campo].set(self.fecha_larga_en(fecha.isoformat()))
            popup.destroy()

        def limpiar():
            self.informes_filter_vars[campo].set("")
            self.informes_fecha_display_vars[campo].set("All dates")
            popup.destroy()

        acciones = tk.Frame(contenido, bg=self.colors["bg_card"])
        acciones.pack(fill="x", pady=(18, 0))
        ttk.Button(acciones, text="Apply", style="Olive.TButton", command=aplicar).pack(side="left", padx=(0, 8))
        ttk.Button(acciones, text="Clear", style="Gray.TButton", command=limpiar).pack(side="left", padx=(0, 8))
        ttk.Button(acciones, text="Cancel", style="Gray.TButton", command=popup.destroy).pack(side="left")

    def cargar_informes_operaciones(self):
        if self.informes_tree is None:
            return

        def tarea():
            return self.api_get_operaciones_buque()

        def al_terminar(data):
            self.informes_cache = data.get("data", []) if isinstance(data, dict) else []

            for item in self.informes_tree.get_children():
                self.informes_tree.delete(item)

            for operacion in self.informes_cache:
                self.informes_tree.insert(
                    "",
                    "end",
                    values=(
                        operacion.get("id", ""),
                        operacion.get("codigo_operacion", ""),
                        operacion.get("nombre_buque", ""),
                        operacion.get("fecha_inicio", ""),
                        operacion.get("fecha_cierre", ""),
                        operacion.get("producto", ""),
                        operacion.get("estado", ""),
                    ),
                )

        self.ejecutar_en_segundo_plano(
            "Informes",
            "Buscando operaciones para informes...",
            tarea,
            al_terminar,
        )

    def obtener_operacion_informe_seleccionada(self):
        if self.informes_tree is None:
            return None

        seleccion = self.informes_tree.selection()
        if not seleccion:
            messagebox.showwarning("Sin seleccion", "Seleccione una operacion para el informe.")
            return None

        valores = self.informes_tree.item(seleccion[0], "values")
        operacion_id = self.safe_int(valores[0], None) if valores else None
        if not operacion_id:
            messagebox.showwarning("Seleccion invalida", "No se pudo leer la operacion seleccionada.")
            return None

        for operacion in self.informes_cache:
            if self.safe_int(operacion.get("id"), None) == operacion_id:
                return operacion

        return {"id": operacion_id}

    def ver_informe_seleccionado(self):
        operacion = self.obtener_operacion_informe_seleccionada()
        if not operacion:
            return

        operacion_id = operacion.get("id")
        params = self.obtener_parametros_informe()

        def tarea():
            return self.api_get_reporte_buque(operacion_id, params)

        def al_terminar(data):
            self.mostrar_ventana_informe_operacion(data)

        self.ejecutar_en_segundo_plano(
            "Informe operativo",
            "Generando vista del informe...",
            tarea,
            al_terminar,
        )

    def descargar_informe_seleccionado(self):
        operacion = self.obtener_operacion_informe_seleccionada()
        if not operacion:
            return

        formato_ui = self.informes_exportar_formato_var.get().strip().lower() if hasattr(self, "informes_exportar_formato_var") else "pdf"
        formato = "excel" if formato_ui == "excel" else "pdf"
        extension = ".xlsx" if formato == "excel" else ".pdf"
        filetypes = [("Excel", "*.xlsx")] if formato == "excel" else [("PDF", "*.pdf")]
        ruta = filedialog.asksaveasfilename(
            title=f"Exportar informe {formato_ui.upper()}",
            defaultextension=extension,
            filetypes=filetypes,
        )
        if not ruta:
            return

        operacion_id = operacion.get("id")
        params = self.obtener_parametros_informe()

        def tarea():
            self.api_descargar_reporte_buque(operacion_id, formato, ruta, params)
            return ruta

        def al_terminar(ruta_generada):
            messagebox.showinfo("Informe exportado", f"Archivo generado correctamente:\n{ruta}")

        self.ejecutar_en_segundo_plano(
            "Exportar informe",
            f"Generando informe {formato_ui.upper()}...",
            tarea,
            al_terminar,
        )

    def mostrar_ventana_informe_operacion(self, data):
        operacion = data.get("operacion", {})
        kpis = data.get("kpis", {})
        cuotas_raw = data.get("cuotas_vs_retiro", data.get("clientes", []))
        productos_raw = data.get("por_producto", data.get("graficos", {}).get("retiro_por_producto", []))
        cuotas = []
        for row in cuotas_raw or []:
            if not isinstance(row, dict):
                continue
            producto = str(row.get("producto") or "").strip()
            cuotas.append({
                "cliente": row.get("cliente") or row.get("empresa"),
                "producto": "Cuota general" if producto.upper() in ("", "TODOS", "ALL") else producto,
                "cuota_mt": row.get("cuota_mt", self.safe_number(row.get("cuota_kg"), 0)),
                "retirado_mt": row.get("retirado_mt", self.safe_number(row.get("retirado_kg"), 0)),
                "saldo_mt": row.get("saldo_mt", row.get("faltante_mt", self.safe_number(row.get("diferencia_kg") or row.get("faltante_kg"), 0))),
                "faltante_mt": row.get("faltante_mt", self.safe_number(row.get("diferencia_kg") or row.get("faltante_kg"), 0)),
                "sobre_descarga_mt": row.get("sobre_descarga_mt"),
                "avance_pct": row.get("avance_pct"),
                "guias": row.get("guias", row.get("camiones")),
            })
        productos = []
        for row in productos_raw or []:
            if not isinstance(row, dict):
                continue
            producto = str(row.get("producto") or "").strip()
            productos.append({
                "producto": "Cuota general" if producto.upper() in ("", "TODOS", "ALL") else producto,
                "guias": row.get("guias"),
                "retirado_mt": row.get("retirado_mt", self.safe_number(row.get("retirado_kg"), 0)),
            })
        alertas = data.get("alertas", [])

        ventana = tk.Toplevel(self)
        ventana.title(f"Informe {operacion.get('nombre_buque', '')}")
        ventana.geometry("980x720")
        ventana.minsize(860, 620)
        ventana.resizable(True, True)
        ventana.configure(bg=self.colors["bg_main"])

        header = tk.Frame(ventana, bg=self.colors["bg_topbar"], height=78)
        header.pack(fill="x")
        header.pack_propagate(False)

        tk.Label(
            header,
            text=f"Informe de Buque: {operacion.get('nombre_buque', '')}",
            font=("Segoe UI", 18, "bold"),
            bg=self.colors["bg_topbar"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=18, pady=(12, 0))
        tk.Label(
            header,
            text=f"Codigo: {operacion.get('codigo_operacion', '')} | Estado: {operacion.get('estado', '')} | Inicio: {operacion.get('fecha_inicio', '')}",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors["bg_topbar"],
            fg=self.colors["text_secondary"],
        ).pack(anchor="w", padx=18)

        host = tk.Frame(ventana, bg=self.colors["bg_main"])
        host.pack(fill="both", expand=True, padx=18, pady=18)

        canvas = tk.Canvas(host, bg=self.colors["bg_main"], highlightthickness=0)
        scroll_y = ttk.Scrollbar(host, orient="vertical", command=canvas.yview)
        scroll_x = ttk.Scrollbar(host, orient="horizontal", command=canvas.xview)
        body = tk.Frame(canvas, bg=self.colors["bg_main"])

        window_id = canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        self.bind_scroll_canvas(canvas, body, window_id, min_width=900)
        canvas.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        host.grid_rowconfigure(0, weight=1)
        host.grid_columnconfigure(0, weight=1)

        cards = tk.Frame(body, bg=self.colors["bg_main"])
        cards.pack(fill="x", pady=(0, 12))
        self.create_card(cards, "Guias", self.formatear_numero(kpis.get("total_guias")), self.colors["accent"])
        self.create_card(cards, "Completas", self.formatear_numero(kpis.get("completas")), self.colors["success"])
        self.create_card(cards, "Descargado MT", self.formatear_numero(kpis.get("retirado_mt", self.safe_number(kpis.get("retirado_kg"), 0)), 2), self.colors["info"])
        self.create_card(cards, "Alertas", self.formatear_numero(kpis.get("alertas")), self.colors["warning"])

        corte_cliente = data.get("corte_cliente", {}) if isinstance(data.get("corte_cliente"), dict) else {}
        corte_totales = corte_cliente.get("totales", {}) if isinstance(corte_cliente.get("totales"), dict) else {}
        cards_cliente = tk.Frame(body, bg=self.colors["bg_main"])
        cards_cliente.pack(fill="x", pady=(0, 12))
        self.create_card(cards_cliente, "Cuota total MT", self.formatear_numero(corte_totales.get("cuota_tm"), 2), self.colors["accent_light"])
        self.create_card(cards_cliente, "Pendiente MT", self.formatear_numero(corte_totales.get("pendiente_tm"), 2), self.colors["warning"])
        self.create_card(cards_cliente, "Avance cuota", f"{self.safe_number(corte_totales.get('avance_pct')):,.2f}%", self.colors["success"])
        self.create_card(cards_cliente, "Promedio MT/viaje", self.formatear_numero(kpis.get("promedio_mt_camion"), 2), self.colors["info"])

        corte_rows = []
        for row in corte_cliente.get("rows", []) or []:
            corte_rows.append({
                "empresa": row.get("empresa"),
                "cuota_pct": row.get("cuota_pct"),
                "cuota_tm": row.get("cuota_tm"),
                "cuota_viajes": row.get("cuota_viajes"),
                "retirado_tm": row.get("retirado_tm"),
                "retirado_pct": row.get("retirado_pct"),
                "retirado_viajes": row.get("retirado_viajes"),
                "promedio_x_viaje": row.get("promedio_x_viaje"),
                "pendiente_tm": row.get("pendiente_tm"),
                "pendiente_viajes": row.get("pendiente_viajes"),
            })

        visual_panel = tk.Frame(body, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1)
        visual_panel.pack(fill="x", pady=(0, 12))
        tk.Label(visual_panel, text="Vista ejecutiva", font=("Segoe UI", 13, "bold"), bg=self.colors["bg_card"], fg=self.colors["text_dark"]).pack(anchor="w", padx=14, pady=(12, 6))
        visual_grid = tk.Frame(visual_panel, bg=self.colors["bg_card"])
        visual_grid.pack(fill="x", padx=14, pady=(0, 14))
        grafico_clientes = tk.Canvas(visual_grid, bg=self.colors["bg_card"], height=250, highlightthickness=0)
        grafico_clientes.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        silueta_bodegas = tk.Canvas(visual_grid, bg=self.colors["bg_card"], height=250, highlightthickness=0)
        silueta_bodegas.grid(row=0, column=1, sticky="ew", padx=(8, 0))
        visual_grid.grid_columnconfigure(0, weight=1)
        visual_grid.grid_columnconfigure(1, weight=1)
        grafico_clientes.after(80, lambda c=grafico_clientes, rows=corte_rows: self.dibujar_grafico_corte_cliente(c, rows))
        silueta_bodegas.after(80, lambda c=silueta_bodegas, d=data: self.dibujar_silueta_bodegas_informe(c, d))

        self.crear_tabla_informe(
            body,
            "CORTE FINAL - CUOTA VS DESCARGADO",
            ("empresa", "cuota_pct", "cuota_tm", "cuota_viajes", "retirado_tm", "retirado_pct", "retirado_viajes", "promedio_x_viaje", "pendiente_tm", "pendiente_viajes"),
            {
                "empresa": "EMPRESA",
                "cuota_pct": "CUOTA %",
                "cuota_tm": "CUOTA T.M.",
                "cuota_viajes": "CUOTA # VIAJES",
                "retirado_tm": "RETIRADO T.M.",
                "retirado_pct": "RETIRADO %",
                "retirado_viajes": "RETIRADO # VIAJES",
                "promedio_x_viaje": "PROMEDIO X VIAJE",
                "pendiente_tm": "PENDIENTE T.M.",
                "pendiente_viajes": "PENDIENTE VIAJES",
            },
            corte_rows if corte_rows else cuotas,
        )

        bodegas_cliente = data.get("reporte_bodegas_cliente", {}) if isinstance(data.get("reporte_bodegas_cliente"), dict) else {}
        bodega_headers = list(bodegas_cliente.get("headers", []) or [])
        bodega_columns = ["concepto"] + [f"col_{idx}" for idx, _header in enumerate(bodega_headers)]
        bodega_headings = {"concepto": "CONCEPTO"}
        for idx, header_text in enumerate(bodega_headers):
            bodega_headings[f"col_{idx}"] = header_text
        bodega_rows = []
        for row in bodegas_cliente.get("rows", []) or []:
            item = {"concepto": row.get("concepto")}
            for idx, value in enumerate(row.get("valores", []) or []):
                item[f"col_{idx}"] = value
            bodega_rows.append(item)
        if bodega_rows:
            self.crear_tabla_informe(
                body,
                "REPORTE DE SALDOS POR BODEGA",
                tuple(bodega_columns),
                bodega_headings,
                bodega_rows,
                height=11,
            )

        self.crear_tabla_informe(
            body,
            "Resumen por producto",
            ("producto", "guias", "retirado_mt"),
            {
                "producto": "Producto",
                "guias": "Guias",
                "retirado_mt": "Descargado MT",
            },
            productos,
        )

        self.crear_tabla_informe(
            body,
            "Alertas operativas",
            ("severidad", "tipo", "mensaje"),
            {
                "severidad": "Severidad",
                "tipo": "Tipo",
                "mensaje": "Mensaje",
            },
            alertas,
            height=8,
        )

    def crear_tabla_informe(self, parent, titulo, columns, headings, data, height=10):
        panel = tk.Frame(parent, bg=self.colors["bg_card"], highlightbackground=self.colors["border"], highlightthickness=1)
        panel.pack(fill="both", expand=True, pady=(0, 12))

        tk.Label(
            panel,
            text=titulo,
            font=("Segoe UI", 13, "bold"),
            bg=self.colors["bg_card"],
            fg=self.colors["text_dark"],
        ).pack(anchor="w", padx=14, pady=(12, 6))

        table_frame = tk.Frame(panel, bg=self.colors["bg_card"])
        table_frame.pack(fill="both", expand=True, padx=14, pady=(0, 14))

        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=height)
        for col in columns:
            tree.heading(col, text=headings.get(col, col))
            tree.column(col, width=180 if col != "mensaje" else 520, anchor="center" if col != "mensaje" else "w")

        for row in data or []:
            values = []
            for col in columns:
                value = row.get(col, "")
                if isinstance(value, float):
                    value = f"{value:,.2f}"
                values.append(value)
            tree.insert("", "end", values=values)

        scroll_y = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        scroll_x = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        tree.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

    def dibujar_grafico_corte_cliente(self, canvas, rows):
        canvas.delete("all")
        width = max(canvas.winfo_width(), 420)
        height = max(canvas.winfo_height(), 230)
        canvas.create_text(12, 14, anchor="w", text="Descargado vs pendiente por cliente", fill=self.colors["text_dark"], font=("Segoe UI", 10, "bold"))
        items = sorted(rows or [], key=lambda row: self.safe_number(row.get("pendiente_tm"), 0), reverse=True)[:8]
        if not items:
            canvas.create_text(width / 2, height / 2, text="Sin datos para graficar", fill=self.colors["text_secondary"], font=("Segoe UI", 10, "bold"))
            return
        max_total = max(self.safe_number(row.get("cuota_tm"), 0) for row in items) or 1
        left = 145
        top = 42
        bar_h = 16
        gap = 9
        right = width - 18
        for idx, row in enumerate(items):
            y = top + idx * (bar_h + gap)
            if y + bar_h > height - 18:
                break
            cliente = str(row.get("empresa") or "SIN CLIENTE")[:18]
            descargado = max(self.safe_number(row.get("retirado_tm"), 0), 0)
            pendiente = max(self.safe_number(row.get("pendiente_tm"), 0), 0)
            cuota = max(self.safe_number(row.get("cuota_tm"), 0), descargado + pendiente, 1)
            canvas.create_text(10, y + bar_h / 2, anchor="w", text=cliente, fill=self.colors["text_dark"], font=("Segoe UI", 8, "bold"))
            total_w = max(1, int((right - left) * cuota / max_total))
            descargado_w = int(total_w * min(descargado / cuota, 1)) if cuota else 0
            canvas.create_rectangle(left, y, left + total_w, y + bar_h, fill=self.colors["bg_main"], outline=self.colors["border"])
            canvas.create_rectangle(left, y, left + max(descargado_w, 1), y + bar_h, fill=self.colors["success"], outline="")
            if pendiente > 0:
                canvas.create_rectangle(left + descargado_w, y, left + total_w, y + bar_h, fill=self.colors["warning"], outline="")
            canvas.create_text(left + total_w + 6, y + bar_h / 2, anchor="w", text=f"{descargado:,.0f}/{cuota:,.0f}", fill=self.colors["text_secondary"], font=("Segoe UI", 8))

    def dibujar_silueta_bodegas_informe(self, canvas, data):
        detalle = data.get("graficos", {}).get("avance_bodegas_detalle") or data.get("bodegas", []) or []
        rows = []
        for row in detalle:
            capacidad = self.safe_number(row.get("capacidad_mt"), 0)
            retirado = self.safe_number(row.get("retirado_mt"), 0)
            rows.append({
                **row,
                "descargado_mt": retirado,
                "pendiente_mt": self.safe_number(row.get("faltante_mt"), max(capacidad - retirado, 0)),
            })
        self.dibujar_silueta_pendiente_bodegas(canvas, rows)

install_frontend_modules(ERPElSurcoApp)


def center_window(window, width, height):
    window.update_idletasks()
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = max((screen_width - width) // 2, 0)
    y = max((screen_height - height) // 2, 0)
    window.geometry(f"{width}x{height}+{x}+{y}")


def show_startup_splash(on_finish):
    splash = tk.Tk()
    try:
        current_scaling = float(splash.tk.call("tk", "scaling"))
        splash.tk.call("tk", "scaling", current_scaling * 1.08)
    except Exception:
        pass
    splash.title("XTRAVON ONE")
    try:
        splash_icon = tk.PhotoImage(file=APP_ICON_PATH)
        splash.iconphoto(True, splash_icon)
        splash._xtravon_icon = splash_icon
    except Exception:
        pass
    splash.configure(bg="#050B14")
    splash.overrideredirect(True)
    splash.resizable(False, False)
    splash.attributes("-topmost", True)

    width = 820
    height = 700
    center_window(splash, width, height)

    shell = tk.Frame(splash, bg="#050B14", highlightbackground="#14283D", highlightthickness=1)
    shell.pack(fill="both", expand=True)

    image_frame = tk.Frame(shell, bg="#050B14")
    image_frame.pack(fill="both", expand=True, padx=20, pady=(20, 12))

    try:
        photo = tk.PhotoImage(file=SPLASH_IMAGE_PATH)
        if photo.width() > 760:
            factor = max((photo.width() + 759) // 760, 1)
            photo = photo.subsample(factor, factor)
        image_label = tk.Label(image_frame, image=photo, bg="#050B14")
        image_label.image = photo
        splash.splash_photo = photo
        image_label.pack(expand=True)
    except Exception:
        tk.Label(
            image_frame,
            text="XTRAVON ONE",
            font=("Segoe UI", 28, "bold"),
            bg="#050B14",
            fg="#F4F8FF",
        ).pack(expand=True)

    progress_wrap = tk.Frame(shell, bg="#050B14")
    progress_wrap.pack(fill="x", padx=70, pady=(0, 22))

    tk.Label(
        progress_wrap,
        text="Preparando sistema...",
        font=("Segoe UI", 10, "bold"),
        bg="#050B14",
        fg="#DCEBFF",
    ).pack(anchor="w", pady=(0, 6))

    progress_bg = tk.Canvas(progress_wrap, height=36, bg="#050B14", highlightthickness=0)
    progress_bg.pack(fill="x")
    progress_track_glow = progress_bg.create_line(0, 18, 1, 18, fill="#061A2C", width=10)
    progress_track = progress_bg.create_line(0, 18, 1, 18, fill="#0B2B3D", width=2)
    flare_tail_deep = progress_bg.create_line(-260, 18, -80, 18, fill="#00D1FF", width=1)
    flare_tail = progress_bg.create_line(-185, 18, -36, 18, fill="#00D1FF", width=3)
    flare_core = progress_bg.create_line(-38, 18, 22, 18, fill="#00D1FF", width=6)
    flare_head = progress_bg.create_line(22, 18, 130, 18, fill="#00D1FF", width=2)
    flare_vertical = progress_bg.create_line(0, 7, 0, 29, fill="#00D1FF", width=2)
    flare_star_a = progress_bg.create_line(-18, 8, 18, 28, fill="#00D1FF", width=1)
    flare_star_b = progress_bg.create_line(-18, 28, 18, 8, fill="#00D1FF", width=1)
    flare_halo = progress_bg.create_oval(-16, 2, 16, 34, outline="#00D1FF", width=1)

    started_at = datetime.now()
    finished = {"done": False}
    after_jobs = []

    def splash_after(delay, callback):
        try:
            if splash.winfo_exists():
                after_jobs.append(splash.after(delay, callback))
        except tk.TclError:
            pass

    def destroy_splash():
        for job in list(after_jobs):
            try:
                splash.after_cancel(job)
            except Exception:
                pass
        after_jobs.clear()
        try:
            if splash.winfo_exists():
                splash.destroy()
        except tk.TclError:
            pass

    def keep_splash_visible():
        try:
            if finished["done"] or not splash.winfo_exists():
                return
            splash.lift()
            splash.attributes("-topmost", True)
        except Exception:
            return
        splash_after(250, keep_splash_visible)

    def update_progress():
        try:
            if finished["done"] or not splash.winfo_exists():
                return
        except tk.TclError:
            return

        elapsed = (datetime.now() - started_at).total_seconds() * 1000
        ratio = min(elapsed / SPLASH_DURATION_MS, 1)
        total_width = max(progress_bg.winfo_width(), 1)
        progress_bg.coords(progress_track_glow, 0, 18, total_width, 18)
        progress_bg.coords(progress_track, 0, 18, total_width, 18)
        shine_x = int(-90 + (total_width + 180) * ratio)
        progress_bg.coords(flare_tail_deep, shine_x - 260, 18, shine_x - 65, 18)
        progress_bg.coords(flare_tail, shine_x - 170, 18, shine_x - 36, 18)
        progress_bg.coords(flare_core, shine_x - 34, 18, shine_x + 18, 18)
        progress_bg.coords(flare_head, shine_x + 18, 18, shine_x + 112, 18)
        progress_bg.coords(flare_vertical, shine_x, 2, shine_x, 34)
        progress_bg.coords(flare_star_a, shine_x - 22, 5, shine_x + 22, 31)
        progress_bg.coords(flare_star_b, shine_x - 22, 31, shine_x + 22, 5)
        progress_bg.coords(flare_halo, shine_x - 18, 0, shine_x + 18, 36)
        for item in (flare_tail_deep, flare_tail, flare_head, flare_core, flare_vertical, flare_star_a, flare_star_b, flare_halo):
            progress_bg.tag_raise(item)
        splash.update_idletasks()

        if ratio >= 1:
            finished["done"] = True
            destroy_splash()
            return

        splash_after(50, update_progress)

    splash.lift()
    splash.focus_force()
    splash.update()
    splash_after(250, keep_splash_visible)
    splash_after(100, update_progress)
    splash.mainloop()

    if finished["done"]:
        on_finish()


def _credential_store_available():
    return os.name == "nt"


def _advapi32():
    advapi = ctypes.windll.advapi32
    advapi.CredWriteW.argtypes = [ctypes.POINTER(_WinCredential), ctypes.c_uint32]
    advapi.CredWriteW.restype = ctypes.c_bool
    advapi.CredReadW.argtypes = [ctypes.c_wchar_p, ctypes.c_uint32, ctypes.c_uint32, ctypes.POINTER(ctypes.c_void_p)]
    advapi.CredReadW.restype = ctypes.c_bool
    advapi.CredFree.argtypes = [ctypes.c_void_p]
    advapi.CredFree.restype = None
    return advapi


def _windows_credential_target(usuario):
    return f"{WINDOWS_CREDENTIAL_SERVICE}:{usuario}"


class _WinCredential(ctypes.Structure):
    _fields_ = [
        ("Flags", ctypes.c_uint32),
        ("Type", ctypes.c_uint32),
        ("TargetName", ctypes.c_wchar_p),
        ("Comment", ctypes.c_wchar_p),
        ("LastWritten", ctypes.c_uint64),
        ("CredentialBlobSize", ctypes.c_uint32),
        ("CredentialBlob", ctypes.c_void_p),
        ("Persist", ctypes.c_uint32),
        ("AttributeCount", ctypes.c_uint32),
        ("Attributes", ctypes.c_void_p),
        ("TargetAlias", ctypes.c_wchar_p),
        ("UserName", ctypes.c_wchar_p),
    ]


def _cred_write(target, username, secret):
    blob = secret.encode("utf-16-le")
    blob_buffer = ctypes.create_string_buffer(blob)
    credential = _WinCredential()
    credential.Type = 1  # CRED_TYPE_GENERIC
    credential.TargetName = target
    credential.UserName = username
    credential.CredentialBlobSize = len(blob)
    credential.CredentialBlob = ctypes.cast(blob_buffer, ctypes.c_void_p)
    credential.Persist = 2  # CRED_PERSIST_LOCAL_MACHINE
    if not _advapi32().CredWriteW(ctypes.byref(credential), 0):
        raise ctypes.WinError()


def _cred_read(target):
    credential_ptr = ctypes.c_void_p()
    advapi = _advapi32()
    if not advapi.CredReadW(target, 1, 0, ctypes.byref(credential_ptr)):
        return ""
    try:
        credential = ctypes.cast(credential_ptr, ctypes.POINTER(_WinCredential)).contents
        if not credential.CredentialBlob or credential.CredentialBlobSize <= 0:
            return ""
        blob = ctypes.string_at(credential.CredentialBlob, credential.CredentialBlobSize)
        return blob.decode("utf-16-le")
    finally:
        advapi.CredFree(credential_ptr)


def _save_windows_credential(usuario, password):
    if not _credential_store_available() or not usuario or not password:
        return False
    try:
        _cred_write(_windows_credential_target(usuario), usuario, password)
        _cred_write(_windows_credential_target("_last_user"), "_last_user", usuario)
        return True
    except Exception:
        return False


def _load_windows_credential(usuario=None):
    if not _credential_store_available():
        return "", ""
    try:
        user = usuario or _cred_read(_windows_credential_target("_last_user")) or ""
        password = _cred_read(_windows_credential_target(user)) if user else ""
        return user or "", password or ""
    except Exception:
        return "", ""


class _CredUiInfo(ctypes.Structure):
    _fields_ = [
        ("cbSize", ctypes.c_uint32),
        ("hwndParent", ctypes.c_void_p),
        ("pszMessageText", ctypes.c_wchar_p),
        ("pszCaptionText", ctypes.c_wchar_p),
        ("hbmBanner", ctypes.c_void_p),
    ]


def _force_window_front(window):
    try:
        window.update_idletasks()
        window.lift()
        window.attributes("-topmost", True)
        window.focus_force()
        window.after(350, lambda: window.attributes("-topmost", False))
    except Exception:
        pass


def _bring_windows_security_prompt_to_front(stop_event):
    if os.name != "nt":
        return

    user32 = ctypes.windll.user32
    titles = (
        "seguridad de windows",
        "windows security",
        "windows hello",
    )
    SW_RESTORE = 9
    HWND_TOPMOST = -1
    HWND_NOTOPMOST = -2
    SWP_NOMOVE = 0x0002
    SWP_NOSIZE = 0x0001
    SWP_SHOWWINDOW = 0x0040

    def try_focus_once():
        found = []

        @ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p)
        def enum_proc(hwnd, _lparam):
            try:
                if not user32.IsWindowVisible(hwnd):
                    return True
                length = user32.GetWindowTextLengthW(hwnd)
                if length <= 0:
                    return True
                buffer = ctypes.create_unicode_buffer(length + 1)
                user32.GetWindowTextW(hwnd, buffer, length + 1)
                title = (buffer.value or "").strip().lower()
                if title and any(part in title for part in titles):
                    found.append(hwnd)
            except Exception:
                pass
            return True

        try:
            user32.EnumWindows(enum_proc, 0)
            for hwnd in found:
                user32.ShowWindow(hwnd, SW_RESTORE)
                user32.SetWindowPos(hwnd, HWND_TOPMOST, 0, 0, 0, 0, SWP_NOMOVE | SWP_NOSIZE | SWP_SHOWWINDOW)
                user32.SetForegroundWindow(hwnd)
                user32.SetWindowPos(hwnd, HWND_NOTOPMOST, 0, 0, 0, 0, SWP_NOMOVE | SWP_NOSIZE | SWP_SHOWWINDOW)
                return True
        except Exception:
            return False
        return False

    end_time = datetime.now().timestamp() + 18
    while not stop_event.is_set() and datetime.now().timestamp() < end_time:
        try_focus_once()
        stop_event.wait(0.25)


def _verify_windows_presence(parent=None, reason="Confirme su identidad de Windows para usar la credencial guardada."):
    if os.name != "nt":
        return True

    try:
        if parent is not None:
            try:
                parent.attributes("-topmost", False)
            except Exception:
                pass
            parent.update()
        try:
            ctypes.windll.user32.AllowSetForegroundWindow(-1)
        except Exception:
            pass
    except Exception:
        pass

    powershell = os.path.join(
        os.environ.get("SystemRoot", r"C:\Windows"),
        "System32",
        "WindowsPowerShell",
        "v1.0",
        "powershell.exe",
    )
    if not os.path.exists(powershell):
        try:
            messagebox.showerror(
                "Windows Hello/PIN",
                "No se encontro PowerShell para abrir Windows Hello/PIN.",
                parent=parent,
            )
        except Exception:
            pass
        return False

    safe_reason = str(reason or "XTRAVON ONE requiere confirmar Windows Hello/PIN.").replace("'", "''")
    script = f"""
Add-Type -AssemblyName System.Runtime.WindowsRuntime
[void][Windows.Security.Credentials.UI.UserConsentVerifier, Windows.Security.Credentials.UI, ContentType=WindowsRuntime]

function Await-WinRtOperation($Operation, $ResultType) {{
    $method = [System.WindowsRuntimeSystemExtensions].GetMethods() |
        Where-Object {{
            $_.Name -eq 'AsTask' -and
            $_.IsGenericMethodDefinition -and
            $_.GetParameters().Count -eq 1
        }} |
        Select-Object -First 1

    if (-not $method) {{
        throw 'AsTask generic method not found'
    }}

    $task = $method.MakeGenericMethod($ResultType).Invoke($null, @($Operation))
    $task.Wait()
    return $task.Result
}}

$availability = Await-WinRtOperation `
    ([Windows.Security.Credentials.UI.UserConsentVerifier]::CheckAvailabilityAsync()) `
    ([Windows.Security.Credentials.UI.UserConsentVerifierAvailability])

if ($availability.ToString() -ne 'Available') {{
    Write-Output ('UNAVAILABLE:' + $availability.ToString())
    exit 2
}}

$verification = Await-WinRtOperation `
    ([Windows.Security.Credentials.UI.UserConsentVerifier]::RequestVerificationAsync('{safe_reason}')) `
    ([Windows.Security.Credentials.UI.UserConsentVerificationResult])

Write-Output $verification.ToString()
if ($verification.ToString() -eq 'Verified') {{
    exit 0
}}
if ($verification.ToString() -eq 'Canceled') {{
    exit 3
}}
exit 4
"""

    script_path = None
    focus_stop = threading.Event()
    focus_thread = None
    try:
        with tempfile.NamedTemporaryFile("w", suffix=".ps1", delete=False, encoding="utf-8") as ps_file:
            ps_file.write(script)
            script_path = ps_file.name

        focus_thread = threading.Thread(
            target=_bring_windows_security_prompt_to_front,
            args=(focus_stop,),
            daemon=True,
        )
        focus_thread.start()
        creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        result = subprocess.run(
            [
                powershell,
                "-NoProfile",
                "-STA",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                script_path,
            ],
            capture_output=True,
            text=True,
            timeout=120,
            creationflags=creationflags,
        )
        focus_stop.set()
        output = (result.stdout or result.stderr or "").strip()
        if result.returncode == 0 and "Verified" in output:
            if parent is not None:
                _force_window_front(parent)
            return True
        if result.returncode == 3 or "Canceled" in output:
            return False

        if output.startswith("UNAVAILABLE:"):
            detalle = output.replace("UNAVAILABLE:", "", 1)
            mensaje = (
                "Windows Hello/PIN no esta disponible para este usuario.\n\n"
                f"Estado: {detalle}\n\n"
                "Configure un PIN, huella o rostro en Windows Hello e intente de nuevo."
            )
        else:
            mensaje = (
                "No se pudo confirmar con Windows Hello/PIN.\n\n"
                f"Detalle: {output or 'Sin detalle'}"
            )
        messagebox.showerror("Windows Hello/PIN", mensaje, parent=parent)
        return False
    except subprocess.TimeoutExpired:
        focus_stop.set()
        messagebox.showerror(
            "Windows Hello/PIN",
            "La confirmacion de Windows Hello/PIN tardo demasiado.",
            parent=parent,
        )
        return False
    except Exception as exc:
        focus_stop.set()
        messagebox.showerror(
            "Windows Hello/PIN",
            f"No se pudo abrir Windows Hello/PIN.\n\nDetalle: {exc}",
            parent=parent,
        )
        return False
    finally:
        focus_stop.set()
        if focus_thread is not None:
            try:
                focus_thread.join(timeout=0.5)
            except Exception:
                pass
        if parent is not None:
            try:
                _force_window_front(parent)
            except Exception:
                pass
        if script_path:
            try:
                os.remove(script_path)
            except OSError:
                pass


def _show_mfa_setup_dialog(parent, payload):
    dialog = tk.Toplevel(parent)
    dialog.title("Microsoft Authenticator")
    dialog.configure(bg="#050B14")
    dialog.geometry("470x560")
    dialog.transient(parent)
    dialog.grab_set()
    center_window(dialog, 470, 560)

    tk.Label(
        dialog,
        text="Configurar Microsoft Authenticator",
        font=("Segoe UI", 16, "bold"),
        bg="#050B14",
        fg="#F4F8FF",
    ).pack(anchor="w", padx=18, pady=(18, 4))
    tk.Label(
        dialog,
        text="Escanee este QR una sola vez y escriba el codigo de 6 digitos.",
        font=("Segoe UI", 10),
        bg="#050B14",
        fg="#B9D8FF",
        wraplength=420,
        justify="left",
    ).pack(anchor="w", padx=18, pady=(0, 12))

    qr_data = str(payload.get("mfa_qr_data_url") or "")
    qr_label = tk.Label(dialog, bg="#FFFFFF")
    qr_label.pack(padx=18, pady=(0, 12))
    try:
        if "," in qr_data:
            qr_b64 = qr_data.split(",", 1)[1]
            qr_raw_img = tk.PhotoImage(data=qr_b64)
            max_side = max(qr_raw_img.width(), qr_raw_img.height())
            factor = max(1, (max_side + 259) // 260)
            qr_img = qr_raw_img.subsample(factor, factor)
            qr_label.configure(image=qr_img)
            qr_label._qr_raw_img = qr_raw_img
            qr_label._qr_img = qr_img
        else:
            qr_label.configure(text="QR no disponible", fg="#050B14")
    except Exception:
        qr_label.configure(text="QR no disponible", fg="#050B14")

    code_var = tk.StringVar()
    tk.Label(dialog, text="Codigo Authenticator", font=("Segoe UI", 9, "bold"), bg="#050B14", fg="#F4F8FF").pack(anchor="w", padx=18)
    entry = ttk.Entry(dialog, textvariable=code_var, justify="center", font=("Segoe UI", 14, "bold"))
    entry.pack(fill="x", padx=18, pady=(4, 14))

    result = {"code": None}

    def confirm():
        code = "".join(ch for ch in code_var.get() if ch.isdigit())
        if len(code) != 6:
            messagebox.showwarning("Codigo requerido", "Ingrese el codigo de 6 digitos.", parent=dialog)
            return
        result["code"] = code
        dialog.destroy()

    actions = tk.Frame(dialog, bg="#050B14")
    actions.pack(fill="x", padx=18)
    ttk.Button(actions, text="Confirmar", style="Olive.TButton", command=confirm).pack(side="left", fill="x", expand=True, padx=(0, 8))
    ttk.Button(actions, text="Cancelar", style="Gray.TButton", command=dialog.destroy).pack(side="left", fill="x", expand=True)
    entry.bind("<Return>", lambda _event: confirm())
    entry.focus_set()
    parent.wait_window(dialog)
    return result["code"]


def show_login_window(on_success):
    login = tk.Tk()
    try:
        current_scaling = float(login.tk.call("tk", "scaling"))
        login.tk.call("tk", "scaling", current_scaling * 1.08)
    except Exception:
        pass
    login.title("XTRAVON ONE - Login")
    try:
        login_icon = tk.PhotoImage(file=APP_ICON_PATH)
        login.iconphoto(True, login_icon)
        login._xtravon_icon = login_icon
    except Exception:
        pass
    login.geometry("640x560")
    login.minsize(600, 520)
    login.configure(bg="#050B14")
    center_window(login, 640, 560)
    login_style = ttk.Style(login)
    try:
        login_style.theme_use("clam")
    except Exception:
        pass
    login_style.configure("Olive.TButton", background="#00D1FF", foreground="#050B14", font=("Segoe UI", 10, "bold"), padding=10)
    login_style.map("Olive.TButton", background=[("active", "#0877D9")])
    login_style.configure("Gray.TButton", background="#14283D", foreground="#F4F8FF", font=("Segoe UI", 10, "bold"), padding=10)
    login_style.map("Gray.TButton", background=[("active", "#2979FF")])

    shell = tk.Frame(login, bg="#0B1B2E", highlightbackground="#14283D", highlightthickness=1)
    shell.pack(fill="both", expand=True, padx=24, pady=24)

    tk.Label(
        shell,
        text="XTRAVON ONE",
        font=("Segoe UI", 20, "bold"),
        bg="#0B1B2E",
        fg="#F4F8FF",
    ).pack(anchor="w", padx=24, pady=(22, 2))

    tk.Label(
        shell,
        text="Ingreso al sistema",
        font=("Segoe UI", 10, "bold"),
        bg="#0B1B2E",
        fg="#DCEBFF",
    ).pack(anchor="w", padx=24, pady=(0, 18))

    form = tk.Frame(shell, bg="#0B1B2E")
    form.pack(fill="x", padx=24)

    usuario_var = tk.StringVar()
    password_var = tk.StringVar()
    mfa_var = tk.StringVar()
    mfa_state = {"pending": False}
    recordar_var = tk.BooleanVar(value=False)
    stored_user, stored_password = _load_windows_credential()
    if stored_user:
        usuario_var.set(stored_user)

    tk.Label(form, text="Usuario", font=("Segoe UI", 9, "bold"), bg="#0B1B2E", fg="#F4F8FF").pack(anchor="w")
    usuario_entry = ttk.Entry(form, textvariable=usuario_var)
    usuario_entry.pack(fill="x", pady=(4, 12))

    tk.Label(form, text="Contraseña", font=("Segoe UI", 9, "bold"), bg="#0B1B2E", fg="#F4F8FF").pack(anchor="w")
    password_entry = ttk.Entry(form, textvariable=password_var, show="*")
    password_entry.pack(fill="x", pady=(4, 10))

    mfa_frame = tk.Frame(form, bg="#0B1B2E")
    tk.Label(mfa_frame, text="Codigo Microsoft Authenticator", font=("Segoe UI", 9, "bold"), bg="#0B1B2E", fg="#F4F8FF").pack(anchor="w")
    mfa_entry = ttk.Entry(mfa_frame, textvariable=mfa_var)
    mfa_entry.pack(fill="x", pady=(4, 10))

    def mostrar_mfa(mensaje=None):
        mfa_state["pending"] = True
        if not mfa_frame.winfo_ismapped():
            mfa_frame.pack(fill="x", before=options)
        if mensaje:
            messagebox.showinfo("Microsoft Authenticator", mensaje, parent=login)
        mfa_entry.focus_set()

    options = tk.Frame(form, bg="#0B1B2E")
    options.pack(fill="x", pady=(0, 10))
    tk.Checkbutton(
        options,
        text="Recordar contrasena en Windows",
        variable=recordar_var,
        bg="#0B1B2E",
        fg="#DCEBFF",
        selectcolor="#14283D",
        activebackground="#0B1B2E",
        activeforeground="#F4F8FF",
        font=("Segoe UI", 9, "bold"),
    ).pack(side="left")

    def usar_guardada():
        if not _verify_windows_presence(login, "Confirme su PIN, rostro, huella o clave de Windows para usar la contrasena guardada."):
            messagebox.showinfo("Credencial Windows", "Operacion cancelada. No se completaron datos guardados.", parent=login)
            return
        user, saved = _load_windows_credential(usuario_var.get().strip() or None)
        if user:
            usuario_var.set(user)
        if saved:
            password_var.set(saved)
            password_entry.focus_set()
        else:
            messagebox.showinfo(
                "Credencial Windows",
                "No hay contrasena guardada para este usuario o el almacÃ©n seguro no esta disponible.",
            )

    ttk.Button(options, text="Usar guardada", style="Gray.TButton", command=usar_guardada).pack(side="right")

    def ingresar():
        usuario = usuario_var.get().strip()
        password = password_var.get()
        if not usuario or not password:
            messagebox.showwarning("Dato requerido", "Ingrese usuario y contrasena.")
            return
        try:
            payload = {"usuario": usuario, "password": password}
            if mfa_state["pending"]:
                code = mfa_var.get().strip()
                if not code:
                    mostrar_mfa("Ingrese el codigo de 6 digitos para continuar.")
                    return
                payload["mfa_code"] = code
            respuesta = post_json_with_retry(
                f"{API_BASE_DEFAULT}/rbac/login",
                payload,
                timeout=45,
                retries=1,
            )
            if respuesta.status_code != 200:
                try:
                    detalle = respuesta.json().get("detail", respuesta.text)
                except Exception:
                    detalle = respuesta.text
                messagebox.showerror("Ingreso no autorizado", str(detalle))
                return
            auth_data = respuesta.json()
            if auth_data.get("mfa_setup_required"):
                code = _show_mfa_setup_dialog(login, auth_data)
                if not code:
                    return
                mfa_state["pending"] = True
                mfa_var.set(code)
                ingresar()
                return
            if auth_data.get("mfa_required"):
                mostrar_mfa(auth_data.get("message") or "Ingrese el codigo de 6 digitos.")
                return
            if recordar_var.get():
                if not _verify_windows_presence(login, "Confirme su identidad de Windows antes de guardar esta contrasena en el equipo."):
                    messagebox.showwarning(
                        "Credencial Windows",
                        "Ingreso correcto, pero no se guardo la contrasena porque no se confirmo Windows.",
                        parent=login,
                    )
                    login.destroy()
                    try:
                        on_success(auth_data)
                    except TypeError:
                        on_success()
                    return
                if not _save_windows_credential(usuario, password):
                    messagebox.showwarning(
                        "Credencial Windows",
                        "Ingreso correcto, pero no se pudo guardar en el almacÃ©n seguro de Windows.",
                    )
            login.destroy()
            try:
                on_success(auth_data)
            except TypeError:
                on_success()
        except requests.exceptions.Timeout:
            messagebox.showerror("Login", "La API tardo demasiado en responder. Intente de nuevo en unos segundos.")
        except requests.exceptions.ConnectionError:
            messagebox.showerror("Login", f"No se pudo conectar con la API:\n{API_BASE_DEFAULT}")
        except Exception as exc:
            messagebox.showerror("Login", str(exc))

    actions = tk.Frame(shell, bg="#0B1B2E")
    actions.pack(fill="x", padx=24, pady=(4, 0))
    ttk.Button(actions, text="Ingresar", style="Olive.TButton", command=ingresar).pack(side="left", fill="x", expand=True, padx=(0, 8))
    ttk.Button(actions, text="Salir", style="Gray.TButton", command=login.destroy).pack(side="left", fill="x", expand=True)

    password_entry.bind("<Return>", lambda _event: ingresar())
    mfa_entry.bind("<Return>", lambda _event: ingresar())
    usuario_entry.bind("<Return>", lambda _event: password_entry.focus_set())
    usuario_entry.focus_set()
    login.mainloop()


if __name__ == "__main__":
    show_startup_splash(lambda: show_login_window(lambda auth=None: ERPElSurcoApp(auth).mainloop()))
